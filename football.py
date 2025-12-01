import sqlite3
import requests
import json
import time
import random
from datetime import datetime, timedelta, timezone
from typing import Optional, Dict, Any, List, Tuple
import os
import re
from dotenv import load_dotenv
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    import sys
    print(f"[Football] openpyxl не доступен для экспорта в Excel")
    print(f"[Football] Python: {sys.executable}")
    print(f"[Football] Ошибка импорта: {e}")
    print(f"[Football] Путь Python: {sys.path[:3]}")



from config import FOOTBALL_DB_FILE
from workflow_model_config import get_model_for_operation

# Попытка импортировать telegram_notifier (может отсутствовать)
try:
    from telegram_notifier import telegram_notifier
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    print("[Football] Telegram notifier не доступен")

load_dotenv()

FOOTBALL_DATABASE_FILE = str(FOOTBALL_DB_FILE)
# Список ключей Odds API для ротации
ODDS_API_KEYS = [
    os.getenv("ODDS_API_KEY_1"),
    os.getenv("ODDS_API_KEY_2"),
    os.getenv("ODDS_API_KEY_3"),
    os.getenv("ODDS_API_KEY_4"),
]
# Фильтруем None значения
ODDS_API_KEYS = [key for key in ODDS_API_KEYS if key]
# Для обратной совместимости
ODDS_API_KEY = ODDS_API_KEYS[0] if ODDS_API_KEYS else os.getenv("ODDS_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
ODDS_API_URL = "https://api.the-odds-api.com/v4"
# Порог для переключения на следующий ключ (осталось запросов)
ODDS_API_SWITCH_THRESHOLD = 10
SOFASCORE_API_URL = "https://api.sofascore1.com/api/v1"

# Список User-Agent'ов для SofaScore (случайный выбор, чтобы уменьшить шанс бана)
SOFASCORE_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/605.1"
]

SOFASCORE_DEFAULT_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.sofascore.com/",
    "Origin": "https://www.sofascore.com",
    "Connection": "keep-alive",
}

# Полный список всех доступных футбольных лиг из The Odds API
# Источник: https://api.the-odds-api.com/v4/sports/
ALL_AVAILABLE_FOOTBALL_LEAGUES = [
    # --- Европейские топ-лиги ---
    "soccer_epl",                    # Английская Премьер-лига (EPL)
    "soccer_spain_la_liga",          # Ла Лига (Испания)
    "soccer_italy_serie_a",          # Серия A (Италия)
    "soccer_germany_bundesliga",     # Бундеслига (Германия)
    "soccer_france_ligue_one",       # Лига 1 (Франция)
    "soccer_netherlands_eredivisie", # Эредивизи (Нидерланды)
    "soccer_portugal_primeira_liga", # Примейра Лига (Португалия)
    "soccer_spl",                    # Премьершип (Шотландия)
    
    # --- Европейские вторые лиги ---
    "soccer_efl_champ",              # Чемпионшип (Англия)
    "soccer_spain_segunda_division", # Ла Лига 2 (Испания)
    "soccer_italy_serie_b",          # Серия B (Италия)
    "soccer_germany_bundesliga2",    # Бундеслига 2 (Германия)
    "soccer_germany_liga3",          # 3. Лига (Германия)
    "soccer_france_ligue_two",       # Лига 2 (Франция)
    "soccer_england_league1",        # Лига 1 (Англия)
    "soccer_england_league2",        # Лига 2 (Англия)
    "soccer_sweden_superettan",      # Суперэттан (Швеция)
    
    # --- Другие европейские лиги ---
    "soccer_belgium_first_div",      # Первый дивизион (Бельгия)
    "soccer_austria_bundesliga",     # Бундеслига (Австрия)
    "soccer_switzerland_superleague", # Суперлига (Швейцария)
    "soccer_greece_super_league",    # Суперлига (Греция)
    "soccer_turkey_super_league",    # Суперлига (Турция)
    "soccer_poland_ekstraklasa",     # Экстракласса (Польша)
    "soccer_denmark_superliga",      # Суперлига (Дания)
    "soccer_norway_eliteserien",     # Элитсериен (Норвегия)
    "soccer_sweden_allsvenskan",     # Алльсвенскан (Швеция)
    "soccer_finland_veikkausliiga",  # Вейккауслига (Финляндия)
    "soccer_germany_liga3",          # 3. Лига (Германия) - дубликат?
    
    # --- Европейские клубные турниры ---
    "soccer_uefa_champs_league",     # Лига Чемпионов
    "soccer_uefa_europa_league",     # Лига Европы
    "soccer_uefa_europa_conference_league", # Лига Конференций
    "soccer_fifa_world_cup_qualifiers_europe", # Отборочные ЧМ (Европа)
    
    # --- Южноамериканские лиги ---
    "soccer_argentina_primera_division", # Примера Дивизион (Аргентина)
    "soccer_brazil_campeonato",      # Серия A (Бразилия)
    "soccer_brazil_serie_b",         # Серия B (Бразилия)
    "soccer_chile_campeonato",       # Примера Дивизион (Чили)
    "soccer_conmebol_copa_libertadores", # Копа Либертадорес
    "soccer_conmebol_copa_sudamericana", # Копа Судамерикана
    
    # --- Североамериканские лиги ---
    "soccer_usa_mls",                # MLS (США/Канада)
    "soccer_mexico_ligamx",          # Лига MX (Мексика)
    
    # --- Азиатские лиги ---
    "soccer_japan_j_league",         # J League (Япония)
    "soccer_korea_kleague1",         # K League 1 (Корея)
    "soccer_china_superleague",      # Суперлига (Китай)
    
    # --- Океания ---
    "soccer_australia_aleague",      # A-League (Австралия)
]

# Список лиг для сбора матчей (можно переопределить через FOOTBALL_LEAGUES в .env)
# Формат: "soccer_epl,soccer_spain_la_liga,soccer_germany_bundesliga" и т.д.
# Если не указано, используется список по умолчанию ниже
# 
# ВАЖНО: Для отладки используем только 3 лиги, чтобы не выйти за лимит запросов API
# Чтобы включить все лиги, раскомментируйте нужные строки ниже
DEFAULT_FOOTBALL_LEAGUES = [
    "soccer_epl",                    # Английская Премьер-лига
     "soccer_uefa_champs_league",     # Лига Чемпионов
     "soccer_uefa_europa_league",     # Лига Европы
    # --- Раскомментируйте для включения остальных лиг ---
     "soccer_spain_la_liga",          # Ла Лига (Испания)
     "soccer_italy_serie_a",          # Серия A (Италия)
     "soccer_germany_bundesliga",     # Бундеслига (Германия)
     "soccer_france_ligue_one",       # Лига 1 (Франция)
     "soccer_netherlands_eredivisie", # Эредивизи (Нидерланды)
     "soccer_portugal_primeira_liga", # Примейра Лига (Португалия)
    # "soccer_spl",                    # Премьершип (Шотландия)
    # "soccer_efl_champ",              # Чемпионшип (Англия)
    # "soccer_spain_segunda_division", # Ла Лига 2 (Испания)
    # "soccer_italy_serie_b",          # Серия B (Италия)
    # "soccer_germany_bundesliga2",    # Бундеслига 2 (Германия)
    # "soccer_germany_liga3",          # 3. Лига (Германия)
    # "soccer_france_ligue_two",       # Лига 2 (Франция)
    # "soccer_england_league1",        # Лига 1 (Англия)
    # "soccer_england_league2",        # Лига 2 (Англия)
     "soccer_belgium_first_div",      # Первый дивизион (Бельгия)
     "soccer_austria_bundesliga",     # Бундеслига (Австрия)
     "soccer_switzerland_superleague", # Суперлига (Швейцария)
    # "soccer_greece_super_league",    # Суперлига (Греция)
     "soccer_turkey_super_league",    # Суперлига (Турция)
    # "soccer_poland_ekstraklasa",     # Экстракласса (Польша)
    # "soccer_denmark_superliga",      # Суперлига (Дания)
     "soccer_norway_eliteserien",     # Элитсериен (Норвегия)
    # "soccer_sweden_allsvenskan",     # Алльсвенскан (Швеция)
    # "soccer_sweden_superettan",      # Суперэттан (Швеция)
    # "soccer_finland_veikkausliiga",  # Вейккауслига (Финляндия)
     "soccer_uefa_europa_conference_league", # Лига Конференций
     "soccer_fifa_world_cup_qualifiers_europe", # Отборочные ЧМ (Европа)
     "soccer_argentina_primera_division", # Примера Дивизион (Аргентина)
     "soccer_brazil_campeonato",      # Серия A (Бразилия)
    # "soccer_brazil_serie_b",         # Серия B (Бразилия)
    # "soccer_chile_campeonato",       # Примера Дивизион (Чили)
     "soccer_conmebol_copa_libertadores", # Копа Либертадорес
     "soccer_conmebol_copa_sudamericana", # Копа Судамерикана
     "soccer_usa_mls",                # MLS (США/Канада)
    # "soccer_mexico_ligamx",          # Лига MX (Мексика)
    # "soccer_japan_j_league",         # J League (Япония)
    # "soccer_korea_kleague1",         # K League 1 (Корея)
    # "soccer_china_superleague",      # Суперлига (Китай)
    # "soccer_australia_aleague",      # A-League (Австралия)
]

# Глобальный экземпляр менеджера
_manager = None


def get_football_db_connection():
    """Создает соединение с БД футбольных матчей."""
    conn = sqlite3.connect(FOOTBALL_DATABASE_FILE, check_same_thread=False, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn


def init_football_db():
    """
    Инициализирует базу данных футбольных матчей.
    Создает таблицу matches если её нет.
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON;")

        # --- Создание таблицы matches ---
        print("[FootballDB] Checking/Creating 'matches' table...")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fixture_id TEXT UNIQUE NOT NULL,
                sofascore_event_id INTEGER,
                home_team TEXT NOT NULL,
                away_team TEXT NOT NULL,
                fav TEXT NOT NULL,
                fav_team_id INTEGER NOT NULL,  -- 1=home, 0=away
                match_date TEXT NOT NULL,
                match_time TEXT NOT NULL,
                initial_odds REAL,
                status TEXT NOT NULL DEFAULT 'scheduled',
                stats_60min TEXT,  -- JSON с статистикой на 60-й минуте
                bet INTEGER,  -- Результат проверки условий на 60-й минуте
                final_score_home INTEGER,
                final_score_away INTEGER,
                fav_won INTEGER,  -- 1 если фаворит выиграл, 0 если нет
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        
        # --- Проверка и добавление поля bet ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet' not in columns:
            print("[FootballDB] Adding 'bet' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet INTEGER")
            conn.commit()
        
        # --- Проверка и добавление поля sofascore_join ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'sofascore_join' not in columns:
            print("[FootballDB] Adding 'sofascore_join' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN sofascore_join TEXT")
            conn.commit()
            print("[FootballDB] Column 'sofascore_join' added successfully.")
        else:
            print("[FootballDB] Column 'sofascore_join' already exists.")
        
                # --- Проверка и добавление поля last_odds ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'last_odds' not in columns:
            print("[FootballDB] Adding 'last_odds' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN last_odds REAL")
            conn.commit()
            print("[FootballDB] Column 'last_odds' added successfully.")
        else:
            print("[FootballDB] Column 'last_odds' already exists.")
        
        # --- Проверка и добавление поля bet_ai ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_ai' not in columns:
            print("[FootballDB] Adding 'bet_ai' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_ai TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_ai' added successfully.")
        else:
            print("[FootballDB] Column 'bet_ai' already exists.")
        
        # --- Проверка и добавление поля bet_ai_reason ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_ai_reason' not in columns:
            print("[FootballDB] Adding 'bet_ai_reason' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_ai_reason TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_ai_reason' added successfully.")
        else:
            print("[FootballDB] Column 'bet_ai_reason' already exists.")
        
        # --- Проверка и добавление поля bet_ai_full_response ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_ai_full_response' not in columns:
            print("[FootballDB] Adding 'bet_ai_full_response' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_ai_full_response TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_ai_full_response' added successfully.")
        else:
            print("[FootballDB] Column 'bet_ai_full_response' already exists.")
        
        # --- Проверка и добавление поля bet_ai_model_name ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_ai_model_name' not in columns:
            print("[FootballDB] Adding 'bet_ai_model_name' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_ai_model_name TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_ai_model_name' added successfully.")
        else:
            print("[FootballDB] Column 'bet_ai_model_name' already exists.")
        
        # --- Проверка и добавление поля bet_ai_odds ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_ai_odds' not in columns:
            print("[FootballDB] Adding 'bet_ai_odds' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_ai_odds REAL")
            conn.commit()
            print("[FootballDB] Column 'bet_ai_odds' added successfully.")
        else:
            print("[FootballDB] Column 'bet_ai_odds' already exists.")
        
        # --- Проверка и добавление поля bet_approve ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_approve' not in columns:
            print("[FootballDB] Adding 'bet_approve' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_approve INTEGER")
            conn.commit()
            print("[FootballDB] Column 'bet_approve' added successfully.")
        else:
            print("[FootballDB] Column 'bet_approve' already exists.")

        # --- Проверка и добавление поля bet_approve_reason ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_approve_reason' not in columns:
            print("[FootballDB] Adding 'bet_approve_reason' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_approve_reason TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_approve_reason' added successfully.")
        else:
            print("[FootballDB] Column 'bet_approve_reason' already exists.")
        
        # --- Проверка и добавление поля bet_alt_code ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_alt_code' not in columns:
            print("[FootballDB] Adding 'bet_alt_code' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_alt_code TEXT")
            conn.commit()
            print("[FootballDB] Column 'bet_alt_code' added successfully.")
        else:
            print("[FootballDB] Column 'bet_alt_code' already exists.")
        
        # --- Проверка и добавление поля bet_alt_odds ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_alt_odds' not in columns:
            print("[FootballDB] Adding 'bet_alt_odds' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_alt_odds REAL")
            conn.commit()
            print("[FootballDB] Column 'bet_alt_odds' added successfully.")
        else:
            print("[FootballDB] Column 'bet_alt_odds' already exists.")
        
        # --- Проверка и добавление поля bet_alt_confirm ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'bet_alt_confirm' not in columns:
            print("[FootballDB] Adding 'bet_alt_confirm' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN bet_alt_confirm INTEGER")
            conn.commit()
            print("[FootballDB] Column 'bet_alt_confirm' added successfully.")
        else:
            print("[FootballDB] Column 'bet_alt_confirm' already exists.")
        
        # --- Проверка и добавление поля live_odds ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'live_odds' not in columns:
            print("[FootballDB] Adding 'live_odds' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN live_odds REAL")
            conn.commit()
            print("[FootballDB] Column 'live_odds' added successfully.")
        else:
            print("[FootballDB] Column 'live_odds' already exists.")
        
        # --- Проверка и добавление полей для коэффициентов исходов (для расчета bet_ai_odds) ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        for odds_field in ['live_odds_1', 'live_odds_x', 'live_odds_2']:
            if odds_field not in columns:
                print(f"[FootballDB] Adding '{odds_field}' column to 'matches' table...")
                cursor.execute(f"ALTER TABLE matches ADD COLUMN {odds_field} REAL")
                conn.commit()
                print(f"[FootballDB] Column '{odds_field}' added successfully.")
            else:
                print(f"[FootballDB] Column '{odds_field}' already exists.")
        
        # --- Проверка и добавление поля sport_key ---
        cursor.execute("PRAGMA table_info(matches)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'sport_key' not in columns:
            print("[FootballDB] Adding 'sport_key' column to 'matches' table...")
            cursor.execute("ALTER TABLE matches ADD COLUMN sport_key TEXT")
            conn.commit()
            print("[FootballDB] Column 'sport_key' added successfully.")
        else:
            print("[FootballDB] Column 'sport_key' already exists.")
        
        # --- Создание индексов ---
        print("[FootballDB] Creating indexes...")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_matches_status ON matches(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_matches_date ON matches(match_date)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_matches_fixture_id ON matches(fixture_id)")
        conn.commit()
        
        # --- Создание/миграция таблицы подписок Telegram ---
        print("[FootballDB] Checking/Creating 'football_telegram_subscriptions' table...")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS football_telegram_subscriptions (
                user_id TEXT PRIMARY KEY,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                is_active BOOLEAN DEFAULT 1
            )
        """)
        conn.commit()
        # Проверяем, нет ли старой схемы с колонкой token
        cursor.execute("PRAGMA table_info(football_telegram_subscriptions)")
        cols = [row[1] for row in cursor.fetchall()]
        if "token" in cols:
            print("[FootballDB] Migrating 'football_telegram_subscriptions' to drop 'token' and add unique user_id...")
            # Переименовываем старую таблицу и переносим данные в новую схему
            cursor.execute("ALTER TABLE football_telegram_subscriptions RENAME TO football_telegram_subscriptions_old")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS football_telegram_subscriptions (
                    user_id TEXT PRIMARY KEY,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_active BOOLEAN DEFAULT 1
                )
            """)
            # Переносим агрегированные данные: берём последнюю created_at и максимальный is_active по user_id
            cursor.execute("""
                INSERT OR REPLACE INTO football_telegram_subscriptions (user_id, created_at, is_active)
                SELECT user_id,
                       MAX(created_at) as created_at,
                       MAX(COALESCE(is_active,0)) as is_active
                FROM football_telegram_subscriptions_old
                GROUP BY user_id
            """)
            cursor.execute("DROP TABLE football_telegram_subscriptions_old")
            conn.commit()
            print("[FootballDB] Migration completed.")
        print("[FootballDB] Table 'football_telegram_subscriptions' created/verified.")

        print("[FootballDB] Database initialization complete.")

    except sqlite3.Error as e:
        print(f"[FootballDB ERROR] Database initialization failed: {e}")
        raise
    finally:
        if conn:
            conn.close()


def get_manager():
    """Получает глобальный экземпляр менеджера."""
    global _manager
    if _manager is None:
        _manager = FootballManager()
    return _manager


class FootballManager:
    """
    Менеджер для работы с футбольными матчами и коэффициентами.
    """

    def __init__(self):
        if not ODDS_API_KEYS:
            # Попытка использовать старый формат для обратной совместимости
            old_key = os.getenv("ODDS_API_KEY")
            if old_key:
                ODDS_API_KEYS.append(old_key)
            else:
                raise ValueError("Не установлены переменные окружения ODDS_API_KEY_1, ODDS_API_KEY_2, ODDS_API_KEY_3, ODDS_API_KEY_4 или ODDS_API_KEY")
        
        # Инициализация ротации ключей
        self.api_keys = ODDS_API_KEYS.copy()
        self.current_key_index = 0
        self.api_key = self.api_keys[self.current_key_index]
        
        # Словарь для отслеживания лимитов каждого ключа: {key_index: {'remaining': int, 'used': int}}
        self.key_limits = {i: {'remaining': None, 'used': None} for i in range(len(self.api_keys))}
        # Внешний провайдер для текущих счетов (TheSportsDB)
        self.thesportsdb_api_key = os.getenv("THESPORTSDB_API_KEY", "123")
        
        # OpenRouter API для ИИ-прогнозов
        self.openrouter_api_key = OPENROUTER_API_KEY
        self.openrouter_api_url = "https://openrouter.ai/api/v1"
        
        # Модели для футбольных прогнозов из конфигурации
        self.ai_primary_model = get_model_for_operation('football_predict', 'primary')
        self.ai_fallback_model1 = get_model_for_operation('football_predict', 'fallback_level1')
        self.ai_fallback_model2 = get_model_for_operation('football_predict', 'fallback_level2')
        self.ai_fallback_model3 = get_model_for_operation('football_predict', 'fallback_level3')
        
        # Модели для анализа риска ставки
        self.risk_analysis_primary = get_model_for_operation('bet_risk_analysis', 'primary')
        self.risk_analysis_fallback1 = get_model_for_operation('bet_risk_analysis', 'fallback_level1')
        self.risk_analysis_fallback2 = get_model_for_operation('bet_risk_analysis', 'fallback_level2')
        self.risk_analysis_fallback3 = get_model_for_operation('bet_risk_analysis', 'fallback_level3')

        # Переменные для отслеживания лимитов API (в памяти)
        self.requests_remaining = None
        self.requests_used = None
        self.requests_last_cost = None
        
        # Словарь для хранения статуса отправки уведомлений о проигрыше фаворита
        # Ключ: fixture_id, Значение: True (уведомление уже отправлено)
        self.favorite_losing_notifications_sent = {}
        
        # Получаем список лиг для сбора (из переменной окружения или по умолчанию)
        leagues_env = os.getenv("FOOTBALL_LEAGUES")
        if leagues_env:
            # Парсим список лиг из переменной окружения (через запятую)
            self.leagues = [league.strip() for league in leagues_env.split(",") if league.strip()]
            print(f"[Football] Используются лиги из FOOTBALL_LEAGUES: {len(self.leagues)} лиг")
        else:
            self.leagues = DEFAULT_FOOTBALL_LEAGUES
            print(f"[Football] Используются лиги по умолчанию: {len(self.leagues)} лиг")
        
        # Инициализируем БД
        init_football_db()
        
        # Получаем начальные значения лимитов через запрос к /sports
        self._initialize_api_limits()
        
        print("[Football] Менеджер инициализирован")

    def build_parlay_preview(self, fixture_ids: List[str], include_all_if_empty: bool = False) -> Optional[Dict[str, Any]]:
        """
        Формирует запрос к ИИ для составления экспресса на основе сырых данных:
        - Используются только live_odds_1, live_odds_x, live_odds_2 и полные stats_60min.
        - Не используются bet_ai и bet_ai_odds.
        Returns dict with keys: {'parlay_json': any|None, 'raw': str|None}
        """
        if not self.openrouter_api_key:
            print("[Football Parlay] OpenRouter API ключ не установлен, пропускаем составление экспресса")
            return None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            try:
                print(f"[Football Parlay] build start fixture_ids={len(fixture_ids)} include_all={include_all_if_empty}")
            except Exception:
                pass
            if fixture_ids:
                qmarks = ",".join("?" for _ in fixture_ids)
                cursor.execute(f"""
                    SELECT fixture_id, home_team, away_team, status, match_date, match_time,
                           live_odds_1, live_odds_x, live_odds_2, stats_60min
                    FROM matches
                    WHERE fixture_id IN ({qmarks})
                    ORDER BY match_date, match_time
                """, tuple(fixture_ids))
            elif include_all_if_empty:
                cursor.execute("""
                    SELECT fixture_id, home_team, away_team, status, match_date, match_time,
                           live_odds_1, live_odds_x, live_odds_2, stats_60min
                    FROM matches
                    ORDER BY match_date, match_time
                """)
            else:
                conn.close()
                return {'parlay_json': None, 'raw': None}
            rows = cursor.fetchall()
            conn.close()

            matches_payload = []
            for row in rows:
                stats = None
                try:
                    stats = json.loads(row['stats_60min']) if row['stats_60min'] else None
                except Exception:
                    stats = row['stats_60min']
                # Примечание: текущий счет находится внутри stats_60min['score'] = {'home': X, 'away': Y}
                matches_payload.append({
                    'fixture_id': row['fixture_id'],
                    'home_team': row['home_team'],
                    'away_team': row['away_team'],
                    'status': row['status'],
                    'match_date': row['match_date'],
                    'match_time': row['match_time'],
                    'live_odds_1': row['live_odds_1'],
                    'live_odds_x': row['live_odds_x'],
                    'live_odds_2': row['live_odds_2'],
                    'stats_60min': stats
                })

            if not matches_payload:
                return {'parlay_json': None, 'raw': None}

            # Подготавливаем большой промпт: строго JSON-ответ
            context_json = json.dumps({'matches': matches_payload}, ensure_ascii=False)
            try:
                print(f"[Football Parlay] context matches={len(matches_payload)} size={len(context_json)}")
            except Exception:
                pass
            system_instruction = (
                "Ты - аналитик футбольных матчей и эксперт в области спортивных ставок. "
                "Тебе предоставлена статистика первых половин матчей нескольких команд. "
                "Твоя задача - составить экспресс из 2–4 событий из предоставленных матчей. "
                "Ты должен учитывать статистику первых половин матчей, текущие коэффициенты букмекеров и другие факторы, которые могут повлиять на исход матча, в том числе исторические и статистические данные. "
                "ВСЕГДА учитывай текущий счет из stats_60min['score'] (stats_60min.score.home + stats_60min.score.away = текущее количество голов) при расчете коэффициентов для тоталов и гандикапов: чем ближе текущий счет к проходу ставки, тем ниже должен быть коэффициент. "
                "Твой экспресс должен быть оптимален по коэффициенту и минимален по риску. "
                "Разрешенные рынки: 1X2, DoubleChance, Handicap, Total. "
                "Для Handicap используй стороны Home/Away и ТОЛЬКО половинные линии (…,-2.5,-2.0,-1.5,-1.0,-0.5,+0.5,+1.0,+1.5,+2.0,+2.5,…); никаких четвертных (0.25/0.75). "
                "Для Total используй Over/Under с ТОЛЬКО половинными линиями (… 2.0, 2.5, 3.0, 3.5 …). Размер линий не ограничивай. "
                "Если точного коэффициента нет, оцени приблизительно на основе темпа/статистики и live_odds_1/x/2, округли до двух знаков и проставь odds_estimated=true. "
                "Не включай взаимно коррелированные ноги одного и того же матча. "
                "Верни СТРОГО JSON (без текста вокруг) формата: "
                "{\"legs\":[{\"fixture_id\":str,\"market\":\"1X2|DoubleChance|Handicap|Total\",\"pick\":\"1|X|2|1X|X2|Home|Away|Over|Under\",\"line\":number|null,\"odds\":number|null,\"odds_estimated\":boolean|null,\"reason\":str}],"
                "\"total_odds\":number|null}."
            )
            prompt = f"{system_instruction}\n\nДанные:\n{context_json}"

            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": os.getenv("OPENROUTER_SITE_URL", "http://localhost:5000")
            }
            models_to_try = [self.ai_primary_model, self.ai_fallback_model1, self.ai_fallback_model2, self.ai_fallback_model3]

            last_raw = None
            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue
                print(f"[Football Parlay] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model}")
                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {"role": "user", "content": prompt}
                        ],
                        "max_tokens": 4000,
                        "temperature": 0.4
                    }
                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=600
                    )
                    try:
                        print(f"[Football Parlay] model={model} status={response.status_code}")
                    except Exception:
                        pass
                    if response.status_code == 200:
                        data = response.json()
                        if 'choices' in data and data['choices']:
                            raw = data['choices'][0]['message']['content']
                            last_raw = raw
                            # Пытаемся извлечь JSON из ответа (с предобработкой частых артефактов)
                            parsed = None
                            try:
                                txt = raw.strip()
                                # Удаляем markdown-фенс, если модель вернула ```json ... ```
                                if txt.startswith('```'):
                                    lines = txt.splitlines()
                                    # убираем первую и последнюю строку, если они выглядят как ```...```
                                    if lines and lines[0].startswith('```'):
                                        lines = lines[1:]
                                    if lines and lines[-1].startswith('```'):
                                        lines = lines[:-1]
                                    txt = "\n".join(lines).strip()
                                # Если total_odds отдан как выражение (например: 1.1 * 1.2 …), завернём в строку, чтобы json распарсился
                                import re as _re
                                txt_quoted = _re.sub(r'("total_odds"\s*:\s*)([^,\}\n]+)', r'\1"\2"', txt)
                                try:
                                    parsed = json.loads(txt_quoted)
                                except Exception:
                                    # Попробуем вытащить первый JSON-блок по скобкам и применить ту же подмену
                                    m = _re.search(r'\{[\s\S]*\}', txt)
                                    if m:
                                        candidate = m.group(0)
                                        candidate = _re.sub(r'("total_odds"\s*:\s*)([^,\}\n]+)', r'\1"\2"', candidate)
                                        parsed = json.loads(candidate)
                            except Exception:
                                parsed = None

                            # Если удалось распарсить, но total_odds не число — попробуем вычислить произведение коэффициентов
                            if isinstance(parsed, dict):
                                # Всегда пересчитываем total_odds из коэффициентов legs для гарантии правильности
                                try:
                                    legs = parsed.get('legs') or []
                                    prod = 1.0
                                    have_any = False
                                    for lg in legs:
                                        od = lg.get('odds')
                                        if isinstance(od, (int, float)):
                                            prod *= float(od)
                                            have_any = True
                                    parsed['total_odds'] = round(prod, 2) if have_any else None
                                except Exception:
                                    pass
                                # проверяем, что действительно есть ноги; иначе пробуем следующую модель
                                legs_list = parsed.get('legs') if isinstance(parsed.get('legs'), list) else []
                                if legs_list:
                                    try:
                                        print(f"[Football Parlay] parsed legs={len(legs_list)} total_odds={parsed.get('total_odds')}")
                                    except Exception:
                                        pass
                                    return {'parlay_json': parsed, 'raw': raw}
                                else:
                                    print("[Football Parlay] parsed but no legs found, trying next model…")
                                    continue
                            else:
                                print("[Football Parlay] could not parse JSON, trying next model…")
                                continue
                        else:
                            print(f"[Football Parlay] Неверный формат ответа от модели {model}")
                    else:
                        print(f"[Football Parlay] HTTP ошибка {response.status_code} для модели {model}")
                        if response.status_code == 429:
                            continue
                except requests.exceptions.Timeout:
                    print(f"[Football Parlay] Таймаут модели {model}")
                    continue
                except Exception as e:
                    print(f"[Football Parlay] Ошибка запроса к модели {model}: {e}")
                    continue
            print("[Football Parlay] Не удалось получить ответ ни от одной модели")
            return {'parlay_json': None, 'raw': last_raw}
        except Exception as e:
            print(f"[Football Parlay ERROR] Ошибка составления экспресса: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def update_inprogress_scores_from_thesportsdb(self) -> int:
        """
        Обновляет текущий счет для матчей в статусе 'in_progress' через API TheSportsDB.
        Возвращает количество обновленных записей.
        """
        updated = 0
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, fixture_id, home_team, away_team, match_date, final_score_home, final_score_away, fav_team_id, fav
                FROM matches
                WHERE status = 'in_progress'
            """)
            rows = cursor.fetchall()
            if not rows:
                return 0
            for row in rows:
                fixture_id = row['fixture_id']
                home = (row['home_team'] or '').strip()
                away = (row['away_team'] or '').strip()
                date_str = row['match_date']
                if not home or not away or not date_str:
                    continue
                # Формируем slug вида "Home_vs_Away" для TheSportsDB
                def to_slug(s: str) -> str:
                    return s.replace(' ', '_')
                slug = f"{to_slug(home)}_vs_{to_slug(away)}"
                url = f"https://www.thesportsdb.com/api/v1/json/{self.thesportsdb_api_key}/searchevents.php?e={slug}&d={date_str}"
                try:
                    resp = requests.get(url, timeout=10)
                    if resp.status_code != 200:
                        continue
                    data = resp.json()
                    events = data.get('event') or []
                    if not events:
                        continue
                    evt = events[0]
                    h = evt.get('intHomeScore')
                    a = evt.get('intAwayScore')
                    # Некоторые значения могут быть строками; пробуем привести к int
                    try:
                        h_val = int(h) if h is not None and str(h).isdigit() else None
                        a_val = int(a) if a is not None and str(a).isdigit() else None
                    except Exception:
                        h_val = None
                        a_val = None
                    if h_val is None or a_val is None:
                        continue
                    # Обновляем счет в БД (используем поля final_* как хранилище текущего счёта)
                    cursor.execute("""
                        UPDATE matches
                        SET final_score_home = ?, final_score_away = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    """, (h_val, a_val, row['id']))
                    conn.commit()
                    updated += 1
                    
                    # Проверяем и отправляем уведомление, если фаворит проигрывает
                    fav_team_id = row['fav_team_id'] if 'fav_team_id' in row.keys() else None
                    fav_team_name = row['fav'] if 'fav' in row.keys() else None
                    self._check_and_notify_favorite_losing(
                        row['fixture_id'],
                        row['home_team'],
                        row['away_team'],
                        fav_team_id,
                        fav_team_name,
                        h_val,
                        a_val
                    )
                except Exception as ex:
                    print(f"[Football Scores] Ошибка обновления для {fixture_id} ({home} vs {away}): {ex}")
                    continue
        except Exception as e:
            print(f"[Football Scores ERROR] {e}")
            import traceback
            print(traceback.format_exc())
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
        return updated

    def _check_and_notify_favorite_losing(
        self,
        fixture_id: str,
        home_team: str,
        away_team: str,
        fav_team_id: Optional[int],
        fav_team_name: Optional[str],
        home_score: int,
        away_score: int
    ) -> None:
        """
        Проверяет, проигрывает ли фаворит, и отправляет уведомление один раз.
        
        Args:
            fixture_id: ID матча
            home_team: Название домашней команды
            away_team: Название гостевой команды
            fav_team_id: ID фаворита (1=home, 0=away, None/-1=нет фаворита)
            fav_team_name: Название фаворита
            home_score: Счет домашней команды
            away_score: Счет гостевой команды
        """
        # Пропускаем, если нет фаворита
        if fav_team_id is None or fav_team_id == -1 or not fav_team_name or fav_team_name == 'NONE':
            return
        
        # Пропускаем, если уведомление уже было отправлено
        if self.favorite_losing_notifications_sent.get(fixture_id, False):
            return
        
        # Проверяем, проигрывает ли фаворит
        is_favorite_losing = False
        
        if fav_team_id == 1:  # Фаворит дома
            if home_score < away_score:
                is_favorite_losing = True
        elif fav_team_id == 0:  # Фаворит в гостях
            if away_score < home_score:
                is_favorite_losing = True
        
        # Если фаворит проигрывает, отправляем уведомление
        if is_favorite_losing:
            try:
                if not TELEGRAM_AVAILABLE:
                    print(f"[Football] Telegram notifier недоступен, пропускаем уведомление для {fixture_id}")
                    return
                
                # Получаем список подписчиков
                subscribers = get_football_subscribers()
                
                if not subscribers:
                    print(f"[Football] Нет подписчиков для уведомления о проигрыше фаворита {fixture_id}")
                    return
                
                # Формируем сообщение
                message = f"""⚠️ <b>Фаворит проигрывает!</b>

🏟️ <b>Матч:</b> {home_team} vs {away_team}
📊 <b>Счет:</b> {home_score} - {away_score}
⭐ <b>Фаворит:</b> {fav_team_name}
                """.strip()
                
                # Отправляем уведомление всем подписчикам
                success_count = 0
                fail_count = 0
                for recipient_id in subscribers:
                    if telegram_notifier.send_message_to_user(recipient_id, message):
                        success_count += 1
                    else:
                        fail_count += 1
                
                if success_count > 0:
                    # Помечаем, что уведомление отправлено
                    self.favorite_losing_notifications_sent[fixture_id] = True
                    print(f"[Football] Отправлено уведомление о проигрыше фаворита: {fixture_id} ({home_team} vs {away_team}, счет {home_score}-{away_score})")
                else:
                    print(f"[Football] Не удалось отправить уведомление о проигрыше фаворита {fixture_id}: все попытки неудачны")
                    
            except Exception as e:
                print(f"[Football ERROR] Ошибка отправки уведомления о проигрыше фаворита для {fixture_id}: {e}")
                import traceback
                print(traceback.format_exc())

    def _extract_api_limits_from_headers(self, response: requests.Response):
        """
        Извлекает лимиты API из заголовков ответа и обновляет переменные класса.
        Проверяет остаток запросов и переключается на следующий ключ при необходимости.
        
        Args:
            response: Объект ответа от requests
        """
        try:
            # Извлекаем заголовки (API использует lowercase заголовки)
            remaining = response.headers.get('x-requests-remaining')
            used = response.headers.get('x-requests-used')
            last_cost = response.headers.get('x-requests-last')
            
            remaining_int = None
            used_int = None
            last_cost_int = None
            
            if remaining is not None:
                try:
                    remaining_int = int(remaining)
                    self.requests_remaining = remaining_int
                except (ValueError, TypeError):
                    pass
            
            if used is not None:
                try:
                    used_int = int(used)
                    self.requests_used = used_int
                except (ValueError, TypeError):
                    pass
            
            if last_cost is not None:
                try:
                    last_cost_int = int(last_cost)
                    self.requests_last_cost = last_cost_int
                except (ValueError, TypeError):
                    pass
            
            # Обновляем лимиты для текущего ключа
            if remaining_int is not None:
                self.key_limits[self.current_key_index]['remaining'] = remaining_int
            if used_int is not None:
                self.key_limits[self.current_key_index]['used'] = used_int
            
            # Логируем текущие значения лимитов
            if self.requests_remaining is not None:
                print(f"[Football API Limits] Ключ #{self.current_key_index + 1}: Осталось запросов: {self.requests_remaining}, Использовано: {self.requests_used}, Стоимость последнего: {self.requests_last_cost}")
                
                # Проверяем, нужно ли переключиться на следующий ключ
                if self.requests_remaining <= ODDS_API_SWITCH_THRESHOLD:
                    print(f"[Football WARNING] Ключ #{self.current_key_index + 1} приближается к лимиту ({self.requests_remaining} запросов). Переключение на следующий ключ...")
                    self._switch_to_next_key()
                elif self.requests_remaining < 50:
                    print(f"[Football WARNING] Критически низкий лимит запросов для ключа #{self.current_key_index + 1}: {self.requests_remaining}")
                elif self.requests_remaining < 100:
                    print(f"[Football WARNING] Низкий лимит запросов для ключа #{self.current_key_index + 1}: {self.requests_remaining}")
                    
        except Exception as e:
            print(f"[Football ERROR] Ошибка извлечения лимитов из заголовков: {e}")
    
    def _switch_to_next_key(self):
        """
        Переключается на следующий доступный ключ API.
        """
        if len(self.api_keys) <= 1:
            print(f"[Football WARNING] Только один ключ доступен, переключение невозможно")
            return
        
        # Находим следующий ключ с достаточным лимитом
        start_index = self.current_key_index
        switched = False
        
        for attempt in range(len(self.api_keys)):
            # Переходим к следующему ключу (циклически)
            self.current_key_index = (self.current_key_index + 1) % len(self.api_keys)
            
            # Проверяем лимит этого ключа
            key_limit = self.key_limits[self.current_key_index]
            remaining = key_limit.get('remaining')
            
            # Если лимит неизвестен или достаточен - используем этот ключ
            if remaining is None or remaining > ODDS_API_SWITCH_THRESHOLD:
                self.api_key = self.api_keys[self.current_key_index]
                print(f"[Football] Переключено на ключ #{self.current_key_index + 1} (осталось запросов: {remaining if remaining is not None else 'неизвестно'})")
                switched = True
                break
        
        if not switched:
            # Если все ключи исчерпаны, используем первый доступный
            self.current_key_index = 0
            self.api_key = self.api_keys[0]
            print(f"[Football WARNING] Все ключи близки к лимиту, используем ключ #{self.current_key_index + 1}")

    def _initialize_api_limits(self):
        """
        Инициализирует начальные значения лимитов API через запрос к /sports для всех ключей.
        Вызывается при старте приложения.
        """
        try:
            print(f"[Football] Получение начальных значений лимитов API для {len(self.api_keys)} ключей...")
            
            # Инициализируем лимиты для всех ключей
            for i, key in enumerate(self.api_keys):
                try:
                    print(f"[Football] Инициализация лимитов для ключа #{i + 1}...")
                    params = {'apiKey': key}
                    url = f"{ODDS_API_URL}/sports"
                    response = requests.get(url, params=params, timeout=30)
                    response.raise_for_status()
                    
                    # Временно переключаемся на этот ключ для извлечения лимитов
                    old_index = self.current_key_index
                    old_key = self.api_key
                    self.current_key_index = i
                    self.api_key = key
                    
                    # Извлекаем лимиты из заголовков
                    self._extract_api_limits_from_headers(response)
                    
                    # Возвращаемся к исходному ключу
                    self.current_key_index = old_index
                    self.api_key = old_key
                    
                    print(f"[Football] Ключ #{i + 1}: осталось={self.key_limits[i].get('remaining', 'неизвестно')}, использовано={self.key_limits[i].get('used', 'неизвестно')}")
                    
                except Exception as e:
                    print(f"[Football ERROR] Ошибка инициализации лимитов для ключа #{i + 1}: {e}")
            
            # Устанавливаем текущий ключ на тот, у которого больше всего запросов
            best_key_index = 0
            best_remaining = self.key_limits[0].get('remaining', 0) or 0
            for i in range(1, len(self.api_keys)):
                remaining = self.key_limits[i].get('remaining', 0) or 0
                if remaining > best_remaining:
                    best_remaining = remaining
                    best_key_index = i
            
            self.current_key_index = best_key_index
            self.api_key = self.api_keys[best_key_index]
            print(f"[Football] Выбран ключ #{best_key_index + 1} с наибольшим остатком запросов: {best_remaining}")
            print(f"[Football] Начальные лимиты API установлены: осталось={self.requests_remaining}, использовано={self.requests_used}")
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка инициализации лимитов API: {e}")
            # Не падаем, если не удалось получить лимиты - продолжим без них

    def _make_api_request(self, endpoint: str, params: dict) -> Optional[list]:
        """
        Выполняет запрос к The Odds API.
        
        Args:
            endpoint: Путь эндпоинта (например, "/sports/soccer_epl/odds")
            params: Параметры запроса
            
        Returns:
            Список данных или None в случае ошибки
        """
        try:
            url = f"{ODDS_API_URL}{endpoint}"
            print(f"[Football] Запрос к API: {endpoint}, params: {params}")
            
            # Добавляем API ключ в параметры
            params['apiKey'] = self.api_key
            
            response = requests.get(url, params=params, timeout=30)
            
            # Проверяем статус ответа
            if response.status_code == 429:
                # Too Many Requests - переключаемся на следующий ключ
                print(f"[Football WARNING] Получен 429 (Too Many Requests) для ключа #{self.current_key_index + 1}. Переключение на следующий ключ...")
                self._switch_to_next_key()
                # Повторяем запрос с новым ключом
                params['apiKey'] = self.api_key
                response = requests.get(url, params=params, timeout=30)
            
            response.raise_for_status()
            
            # Извлекаем и обновляем лимиты из заголовков ответа
            self._extract_api_limits_from_headers(response)
            
            data = response.json()
            
            print(f"[Football] Успешный ответ от API, получено {len(data) if isinstance(data, list) else 1} записей")
            return data
            
        except requests.exceptions.RequestException as e:
            print(f"[Football ERROR] Ошибка запроса к API: {e}")
            # Если это ошибка 429 и есть другие ключи, пробуем переключиться
            if hasattr(e, 'response') and e.response is not None and e.response.status_code == 429:
                if len(self.api_keys) > 1:
                    print(f"[Football] Пробуем переключиться на следующий ключ после ошибки 429...")
                    self._switch_to_next_key()
            return None
        except json.JSONDecodeError as e:
            print(f"[Football ERROR] Ошибка парсинга JSON: {e}")
            return None

    def get_available_soccer_leagues(self) -> List[Dict[str, Any]]:
        """
        Получает список доступных футбольных лиг из API.
        
        Returns:
            Список словарей с информацией о лигах: [{'key': 'soccer_epl', 'title': 'EPL', ...}, ...]
        """
        try:
            data = self._make_api_request("/sports", {})
            
            if not data:
                print("[Football] Не удалось получить список лиг")
                return []
            
            # Фильтруем только футбольные лиги (без outrights)
            soccer_leagues = [
                league for league in data
                if league.get('group') == 'Soccer' and not league.get('has_outrights', False)
            ]
            
            print(f"[Football] Найдено {len(soccer_leagues)} доступных футбольных лиг")
            return soccer_leagues
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка получения списка лиг: {e}")
            import traceback
            print(traceback.format_exc())
            return []

    def _normalize_team_name(self, name: str) -> str:
        """
        Нормализует название команды для сравнения.
        Убирает пробелы, приводит к нижнему регистру, убирает специальные символы и префиксы.
        Нормализует специальные символы (датские, норвежские, немецкие буквы и т.д.).
        
        Args:
            name: Исходное название команды
            
        Returns:
            Нормализованное название
        """
        if not name:
            return ""
        
        # Список префиксов для удаления (распространенные префиксы футбольных клубов)
        prefixes = [
            'sk ', 'fc ', 'sc ', 'cf ', 'ac ', 'as ', 'rc ', 'fk ', 'if ', 'bk ',
            '1. ', '1 ', '2. ', '3. ', 'cd ', 'ud ', 'cf ', 'sd ', 'fc. ', 'sc. ',
            'royale ', 'royal ', 'r. ', 'r ', 'h. ', 'h ', 'v. ', 'v ', 'vs ', 'vs. ',
            'the ', 'of ', 'de ', 'la ', 'le ', 'los ', 'las ', 'el ', 'der ', 'die ', 'das ',
            'afc ', 'cfc ', 'dfc ', 'sfc ', 'pfc ', 'kfc ', 'bfc ', 'vfc ', 'tsv ', 'fsv ',
            'vv ', 'vv. ', 'vvv ', 'vvv-', 'vvv. ', 'vvv ', 'vvv-', 'vvv. '
        ]
        
        normalized = name.lower().strip()
        
        # Удаляем префиксы
        for prefix in prefixes:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):].strip()
        
        # Нормализуем специальные символы (датские, норвежские, немецкие, испанские и т.д.)
        # Это поможет сопоставить "Copenhagen" с "København", "München" с "Munich", и т.д.
        char_replacements = {
            'ø': 'o', 'Ø': 'o',  # Датская/норвежская буква
            'æ': 'ae', 'Æ': 'ae',  # Датская/норвежская буква
            'å': 'aa', 'Å': 'aa',  # Датская/норвежская буква
            'ö': 'o', 'Ö': 'o',  # Немецкая/шведская буква
            'ü': 'u', 'Ü': 'u',  # Немецкая буква
            'ä': 'a', 'Ä': 'a',  # Немецкая/шведская буква
            'ß': 'ss',  # Немецкая буква
            'ñ': 'n', 'Ñ': 'n',  # Испанская буква
            'ç': 'c', 'Ç': 'c',  # Французская/португальская буква
            'é': 'e', 'É': 'e',  # Французская буква
            'è': 'e', 'È': 'e',  # Французская буква
            'ê': 'e', 'Ê': 'e',  # Французская буква
            'ë': 'e', 'Ë': 'e',  # Французская буква
            'à': 'a', 'À': 'a',  # Французская буква
            'á': 'a', 'Á': 'a',  # Испанская буква
            'â': 'a', 'Â': 'a',  # Французская буква
            'ã': 'a', 'Ã': 'a',  # Португальская буква
            'í': 'i', 'Í': 'i',  # Испанская буква
            'î': 'i', 'Î': 'i',  # Французская буква
            'ï': 'i', 'Ï': 'i',  # Французская буква
            'ó': 'o', 'Ó': 'o',  # Испанская буква
            'ô': 'o', 'Ô': 'o',  # Французская буква
            'õ': 'o', 'Õ': 'o',  # Португальская буква
            'ú': 'u', 'Ú': 'u',  # Испанская буква
            'û': 'u', 'Û': 'u',  # Французская буква
            'ý': 'y', 'Ý': 'y',  # Чешская буква
        }
        for old_char, new_char in char_replacements.items():
            normalized = normalized.replace(old_char, new_char)
        
        # Убираем пробелы, приводим к нижнему регистру, убираем дефисы и подчеркивания
        normalized = normalized.replace(" ", "").replace("-", "").replace("_", "")
        
        # Убираем другие специальные символы (точки, запятые и т.д.)
        normalized = ''.join(c for c in normalized if c.isalnum())
        
        return normalized

    def _fetch_sofascore_events(self, date: str, max_retries: int = 5) -> Optional[List[Dict]]:
        """
        Получает список запланированных событий из SofaScore для указанной даты.
        Использует простой requests с ретраями и экспоненциальным бэкоффом.
        
        Args:
            date: Дата в формате YYYY-MM-DD
            max_retries: Максимальное количество попыток
            
        Returns:
            Список событий или None в случае ошибки
        """
        import random
        
        url = f"{SOFASCORE_API_URL}/sport/football/scheduled-events/{date}"
        attempt = 0
        
        while attempt < max_retries:
            attempt += 1
            
            # Формируем заголовки со случайным User-Agent
            headers = SOFASCORE_DEFAULT_HEADERS.copy()
            headers["User-Agent"] = random.choice(SOFASCORE_USER_AGENTS)
            
            try:
                print(f"[Football SofaScore] Запрос событий на дату {date} (попытка {attempt}/{max_retries})")
                response = requests.get(url, headers=headers, timeout=15.0)
                code = response.status_code
                
                if code == 200:
                    # Успешный ответ
                    try:
                        data = response.json()
                        events = data.get('events', [])
                        print(f"[Football SofaScore] Получено {len(events)} событий на дату {date}")
                        return events
                    except json.JSONDecodeError as e:
                        print(f"[Football SofaScore ERROR] Ошибка парсинга JSON: {e}")
                        return None
                
                elif code == 403:
                    # 403 Forbidden - пробуем с задержкой
                    retry_after = response.headers.get("Retry-After")
                    wait = 5 + random.uniform(0.5, 3.0)
                    print(f"[Football SofaScore] 403 Forbidden для даты {date}. Retry-After: {retry_after}. Ждём {wait:.1f}s и пробуем снова...")
                    if attempt < max_retries:
                        time.sleep(wait)
                        continue
                    else:
                        print(f"[Football SofaScore ERROR] Не удалось получить данные за {max_retries} попыток (403 Forbidden)")
                        return None
                
                elif 500 <= code < 600:
                    # Серверная ошибка - экспоненциальный бэкофф
                    wait = min(2 ** attempt + random.random(), 60)
                    print(f"[Football SofaScore] Серверная ошибка {code} для даты {date}. Ждём {wait:.1f}s...")
                    if attempt < max_retries:
                        time.sleep(wait)
                        continue
                    else:
                        print(f"[Football SofaScore ERROR] Серверная ошибка {code} после {max_retries} попыток")
                        return None
                
                else:
                    # Другие коды - выводим и прекращаем
                    print(f"[Football SofaScore ERROR] HTTP {code} для даты {date}. Response: {response.text[:200]}")
                    return None
                    
            except requests.RequestException as e:
                # Ошибка сети - экспоненциальный бэкофф
                wait = min(2 ** attempt + random.random(), 30)
                print(f"[Football SofaScore] Ошибка сети для даты {date}: {e}. Ждём {wait:.1f}s и повторяем...")
                if attempt < max_retries:
                    time.sleep(wait)
                    continue
                else:
                    print(f"[Football SofaScore ERROR] Не удалось получить данные из-за сетевой ошибки после {max_retries} попыток: {e}")
                    return None
            except Exception as e:
                print(f"[Football SofaScore ERROR] Неожиданная ошибка для даты {date}: {e}")
                import traceback
                traceback.print_exc()
                return None
        
        print(f"[Football SofaScore ERROR] Не удалось получить данные за {max_retries} попыток")
        return None

    def _match_sofascore_event(self, match: Dict, sofascore_events: List[Dict]) -> Optional[int]:
        """
        Сопоставляет матч из The Odds API с событием из SofaScore по названиям команд.
        Время не проверяется, так как данные фильтруются по дате, и две команды не могут играть два матча в один день.
        
        Args:
            match: Словарь с данными матча из БД (или The Odds API)
                  Должен содержать: home_team, away_team
            sofascore_events: Список событий из SofaScore (уже отфильтрованные по дате)
            
        Returns:
            sofascore_event_id если найден, иначе None
        """
        try:
            home_team_odds = match.get('home_team', '')
            away_team_odds = match.get('away_team', '')
            
            if not all([home_team_odds, away_team_odds]):
                return None
            
            # Нормализуем названия команд из матча
            home_normalized = self._normalize_team_name(home_team_odds)
            away_normalized = self._normalize_team_name(away_team_odds)
            
            # Ищем совпадение в событиях SofaScore
            for event in sofascore_events:
                try:
                    event_id = event.get('id')
                    if not event_id:
                        continue
                    
                    home_team_obj = event.get('homeTeam', {})
                    away_team_obj = event.get('awayTeam', {})
                    
                    if not home_team_obj or not away_team_obj:
                        continue
                    
                    # Получаем все возможные варианты названий команд из SofaScore
                    home_team_variants = []
                    away_team_variants = []
                    
                    # Основное название
                    if home_team_obj.get('name'):
                        home_team_variants.append(home_team_obj['name'])
                    if away_team_obj.get('name'):
                        away_team_variants.append(away_team_obj['name'])
                    
                    # Короткое название
                    if home_team_obj.get('shortName'):
                        home_team_variants.append(home_team_obj['shortName'])
                    if away_team_obj.get('shortName'):
                        away_team_variants.append(away_team_obj['shortName'])
                    
                    # Переводы (русский, английский и другие)
                    home_translations = home_team_obj.get('fieldTranslations', {}).get('nameTranslation', {})
                    if home_translations:
                        for lang, translation in home_translations.items():
                            if translation:
                                home_team_variants.append(translation)
                    
                    away_translations = away_team_obj.get('fieldTranslations', {}).get('nameTranslation', {})
                    if away_translations:
                        for lang, translation in away_translations.items():
                            if translation:
                                away_team_variants.append(translation)
                    
                    # Нормализуем все варианты
                    home_sf_normalized_set = {self._normalize_team_name(v) for v in home_team_variants if v}
                    away_sf_normalized_set = {self._normalize_team_name(v) for v in away_team_variants if v}
                    
                    # Проверяем совпадение названий команд (оба варианта: прямой и обратный)
                    # Проверяем точное совпадение И частичное (если одно название содержит другое)
                    teams_match = False
                    for home_sf_norm in home_sf_normalized_set:
                        for away_sf_norm in away_sf_normalized_set:
                            # Точное совпадение (прямое или обратное)
                            exact_match = (
                                (home_normalized == home_sf_norm and away_normalized == away_sf_norm) or
                                (home_normalized == away_sf_norm and away_normalized == home_sf_norm)
                            )
                            
                            # Частичное совпадение: одно название является частью другого
                            # Используем минимальную длину 3 символа, чтобы избежать случайных совпадений
                            home_partial_match = (
                                (len(home_normalized) >= 3 and len(home_sf_norm) >= 3) and
                                (home_normalized in home_sf_norm or home_sf_norm in home_normalized)
                            )
                            away_partial_match = (
                                (len(away_normalized) >= 3 and len(away_sf_norm) >= 3) and
                                (away_normalized in away_sf_norm or away_sf_norm in away_normalized)
                            )
                            
                            # Обратное частичное совпадение
                            home_away_partial_match = (
                                (len(home_normalized) >= 3 and len(away_sf_norm) >= 3) and
                                (home_normalized in away_sf_norm or away_sf_norm in home_normalized)
                            )
                            away_home_partial_match = (
                                (len(away_normalized) >= 3 and len(home_sf_norm) >= 3) and
                                (away_normalized in home_sf_norm or home_sf_norm in away_normalized)
                            )
                            
                            # Совпадение, если обе команды совпадают (точно или частично) в одном порядке
                            if exact_match or (home_partial_match and away_partial_match):
                                teams_match = True
                                break
                            
                            # Или обратный порядок
                            if (home_normalized == away_sf_norm and away_normalized == home_sf_norm) or \
                               (home_away_partial_match and away_home_partial_match):
                                teams_match = True
                                break
                            
                        if teams_match:
                            break
                    
                    if not teams_match:
                        continue

                                        # Найдено совпадение по названиям команд (время не проверяем, так как данные уже отфильтрованы по дате)
                    print(f"[Football SofaScore] Найдено совпадение: {home_team_odds} vs {away_team_odds} -> event_id={event_id}")
                    return event_id
                    
                except Exception as e:
                    print(f"[Football SofaScore] Ошибка при обработке события SofaScore: {e}")
                    continue
            
            # Если не нашли совпадение, выводим детальную информацию для отладки
            if home_team_odds and away_team_odds:
                print(f"[Football SofaScore DEBUG] Не найдено совпадение для {home_team_odds} vs {away_team_odds}")
                print(f"[Football SofaScore DEBUG] Нормализованные: {home_normalized} vs {away_normalized}")
                print(f"[Football SofaScore DEBUG] Проверено событий SofaScore: {len(sofascore_events)}")
                # Выводим первые 3 события для примера
                for idx, event in enumerate(sofascore_events[:3]):
                    event_home = event.get('homeTeam', {}).get('name', 'N/A')
                    event_away = event.get('awayTeam', {}).get('name', 'N/A')
                    print(f"[Football SofaScore DEBUG]   Событие {idx+1}: {event_home} vs {event_away}")

            return None
            
        except Exception as e:
            print(f"[Football SofaScore ERROR] Ошибка сопоставления матча: {e}")
            return None

    def _match_sofascore_event_by_team_and_time(self, match: Dict, sofascore_events: List[Dict], time_tolerance_minutes: int = 5) -> Optional[Dict]:
        """
        Сопоставляет матч по одной команде (home или away) + времени.
        Используется как второй проход для матчей, которые не были найдены по двум командам.
        
        Args:
            match: Словарь с данными матча из БД (должен содержать: home_team, away_team, match_date, match_time)
            sofascore_events: Список событий из SofaScore
            time_tolerance_minutes: Допуск по времени в минутах (по умолчанию 5, плюс автоматическая поправка на часовой пояс)
            
        Returns:
            Словарь с ключами 'event_id', 'slug', 'startTimestamp' если найден, иначе None
        """
        try:
            home_team_odds = match.get('home_team', '')
            away_team_odds = match.get('away_team', '')
            match_date = match.get('match_date', '')
            match_time = match.get('match_time', '')
            
            if not all([home_team_odds, away_team_odds, match_date, match_time]):
                return None
            
            # Нормализуем названия команд
            home_normalized = self._normalize_team_name(home_team_odds)
            away_normalized = self._normalize_team_name(away_team_odds)
            
            # Парсим время матча из БД (в UTC, но без tzinfo)
            try:
                match_datetime_naive = datetime.strptime(f"{match_date} {match_time}", "%Y-%m-%d %H:%M")
                # Добавляем UTC часовой пояс, так как время в БД сохранено в UTC
                match_datetime = match_datetime_naive.replace(tzinfo=timezone.utc)
            except Exception as e:
                print(f"[Football SofaScore] Ошибка парсинга времени матча: {e}")
                return None
            
            # Ищем совпадение в событиях SofaScore
            for event in sofascore_events:
                try:
                    event_id = event.get('id')
                    if not event_id:
                        continue
                    
                    home_team_obj = event.get('homeTeam', {})
                    away_team_obj = event.get('awayTeam', {})
                    
                    if not home_team_obj or not away_team_obj:
                        continue
                    
                    # Проверяем время начала матча
                    start_timestamp = event.get('startTimestamp')
                    if not start_timestamp:
                        continue
                    
                    try:
                        # startTimestamp от SofaScore - это Unix timestamp в UTC
                        event_datetime = datetime.fromtimestamp(start_timestamp, tz=timezone.utc)
                        time_diff_seconds = abs((match_datetime - event_datetime).total_seconds())
                        time_diff_minutes = time_diff_seconds / 60
                        
                        # Проверяем, что время совпадает в пределах ±5 минут + поправка на часовой пояс
                        # Если разница близка к целому количеству часов (в пределах ±5 минут),
                        # то это разница в часовых поясах, и мы ее учитываем
                        hours_diff = round(time_diff_minutes / 60)
                        minutes_remainder = abs(time_diff_minutes - hours_diff * 60)
                        
                        # Если остаток меньше 5 минут, значит это разница в часовых поясах
                        if minutes_remainder <= 5:
                            # Время совпадает с учетом часового пояса
                            pass  # Продолжаем проверку команд
                        elif time_diff_minutes <= 5:
                            # Время совпадает без учета часового пояса (тот же часовой пояс)
                            pass  # Продолжаем проверку команд
                        else:
                            # Время не совпадает
                            continue
                    except Exception as e:
                        continue
                    
                    # Для отладки: проверяем, есть ли хотя бы частичное совпадение по командам
                    # Это поможет понять, почему не находится совпадение
                    
                    # Получаем все возможные варианты названий команд из SofaScore
                    home_team_variants = []
                    away_team_variants = []
                    
                    if home_team_obj.get('name'):
                        home_team_variants.append(home_team_obj['name'])
                    if home_team_obj.get('shortName'):
                        home_team_variants.append(home_team_obj['shortName'])
                    home_translations = home_team_obj.get('fieldTranslations', {}).get('nameTranslation', {})
                    if home_translations:
                        for lang, translation in home_translations.items():
                            if translation:
                                home_team_variants.append(translation)
                    
                    if away_team_obj.get('name'):
                        away_team_variants.append(away_team_obj['name'])
                    if away_team_obj.get('shortName'):
                        away_team_variants.append(away_team_obj['shortName'])
                    away_translations = away_team_obj.get('fieldTranslations', {}).get('nameTranslation', {})
                    if away_translations:
                        for lang, translation in away_translations.items():
                            if translation:
                                away_team_variants.append(translation)
                    
                    # Нормализуем все варианты
                    home_sf_normalized_set = {self._normalize_team_name(v) for v in home_team_variants if v}
                    away_sf_normalized_set = {self._normalize_team_name(v) for v in away_team_variants if v}
                    
                    # Проверяем совпадение хотя бы одной команды (home или away) с учетом частичного совпадения
                    home_match = False
                    away_match = False
                    
                    # Проверяем home команду
                    for home_sf_norm in home_sf_normalized_set:
                        if (home_normalized == home_sf_norm or 
                            (len(home_normalized) >= 3 and len(home_sf_norm) >= 3 and 
                             (home_normalized in home_sf_norm or home_sf_norm in home_normalized))):
                            home_match = True
                            break
                    
                    # Проверяем away команду
                    for away_sf_norm in away_sf_normalized_set:
                        if (away_normalized == away_sf_norm or 
                            (len(away_normalized) >= 3 and len(away_sf_norm) >= 3 and 
                             (away_normalized in away_sf_norm or away_sf_norm in away_normalized))):
                            away_match = True
                            break
                    
                    # Проверяем обратный порядок (home vs away или away vs home)
                    if not home_match:
                        for away_sf_norm in away_sf_normalized_set:
                            if (home_normalized == away_sf_norm or 
                                (len(home_normalized) >= 3 and len(away_sf_norm) >= 3 and 
                                 (home_normalized in away_sf_norm or away_sf_norm in home_normalized))):
                                home_match = True
                                break
                    
                    if not away_match:
                        for home_sf_norm in home_sf_normalized_set:
                            if (away_normalized == home_sf_norm or 
                                (len(away_normalized) >= 3 and len(home_sf_norm) >= 3 and 
                                 (away_normalized in home_sf_norm or home_sf_norm in away_normalized))):
                                away_match = True
                                break
                    
                    # Если хотя бы одна команда совпадает и время совпадает, считаем это совпадением
                    if home_match or away_match:
                        print(f"[Football SofaScore] Найдено совпадение по команде+времени: {home_team_odds} vs {away_team_odds} -> event_id={event_id} (home_match={home_match}, away_match={away_match})")
                        # Возвращаем словарь с event_id и данными для сохранения
                        return {
                            'event_id': event_id,
                            'slug': event.get('slug', ''),
                            'startTimestamp': event.get('startTimestamp')
                        }
                    
                except Exception as e:
                    continue
            
            # Если не нашли совпадение, выводим детальную информацию для отладки
            if home_team_odds and away_team_odds:
                print(f"[Football SofaScore DEBUG 2nd pass] Не найдено совпадение по команде+времени для {home_team_odds} vs {away_team_odds}")
                print(f"[Football SofaScore DEBUG 2nd pass] Нормализованные: {home_normalized} vs {away_normalized}")
                print(f"[Football SofaScore DEBUG 2nd pass] Время матча из БД: {match_datetime}")
                print(f"[Football SofaScore DEBUG 2nd pass] Проверено событий SofaScore: {len(sofascore_events)}")
                
                # Ищем события с похожим временем (в пределах 60 минут)
                similar_time_events = []
                for event in sofascore_events:
                    event_time = event.get('startTimestamp')
                    if event_time:
                        # startTimestamp от SofaScore - это Unix timestamp в UTC
                        event_dt = datetime.fromtimestamp(event_time, tz=timezone.utc)
                        time_diff = abs((match_datetime - event_dt).total_seconds()) / 60
                        if time_diff <= 60:  # В пределах часа
                            event_home = event.get('homeTeam', {}).get('name', 'N/A')
                            event_away = event.get('awayTeam', {}).get('name', 'N/A')
                            similar_time_events.append((event_home, event_away, event_dt, time_diff))
                
                if similar_time_events:
                    print(f"[Football SofaScore DEBUG 2nd pass] Найдено {len(similar_time_events)} событий с похожим временем (в пределах 60 мин):")
                    for idx, (eh, ea, edt, tdiff) in enumerate(similar_time_events[:5]):  # Показываем первые 5
                        print(f"[Football SofaScore DEBUG 2nd pass]   {idx+1}. {eh} vs {ea}, время: {edt}, разница: {tdiff:.1f} мин")
                else:
                    print(f"[Football SofaScore DEBUG 2nd pass] Нет событий с похожим временем (в пределах 60 мин)")
                    # Выводим первые 3 события для примера
                    for idx, event in enumerate(sofascore_events[:3]):
                        event_home = event.get('homeTeam', {}).get('name', 'N/A')
                        event_away = event.get('awayTeam', {}).get('name', 'N/A')
                        event_time = event.get('startTimestamp')
                        if event_time:
                            # startTimestamp от SofaScore - это Unix timestamp в UTC
                            event_dt = datetime.fromtimestamp(event_time, tz=timezone.utc)
                            time_diff = abs((match_datetime - event_dt).total_seconds()) / 60
                            print(f"[Football SofaScore DEBUG 2nd pass]   Событие {idx+1}: {event_home} vs {event_away}, время: {event_dt}, разница: {time_diff:.1f} мин")
                        else:
                            print(f"[Football SofaScore DEBUG 2nd pass]   Событие {idx+1}: {event_home} vs {event_away}, время: N/A")
            
            return None
            
        except Exception as e:
            print(f"[Football SofaScore ERROR] Ошибка сопоставления по команде+времени: {e}")
            return None

    def update_sofascore_ids(self) -> Dict[str, int]:
        """
        Обновляет sofascore_event_id для матчей, у которых он отсутствует.
        Запрашивает события из SofaScore для дат матчей без sofascore_event_id.
        
        Returns:
            Словарь со статистикой: {'updated': int, 'failed': int}
        """
        stats = {
            'updated': 0,
            'failed': 0,
            'dates_processed': 0
        }
        
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            
            # Получаем все матчи без sofascore_event_id, сгруппированные по дате
            cursor.execute("""
                SELECT DISTINCT match_date 
                FROM matches 
                WHERE sofascore_event_id IS NULL 
                AND status IN ('scheduled', 'in_progress')
                ORDER BY match_date
            """)
            
            dates_to_process = [row[0] for row in cursor.fetchall()]
            
            if not dates_to_process:
                print("[Football SofaScore] Нет матчей без sofascore_event_id для обновления")
                return stats
            
            print(f"[Football SofaScore] Найдено {len(dates_to_process)} дат для обработки")
            
            # Обрабатываем каждую дату
            for date_str in dates_to_process:
                try:
                    # Получаем события из SofaScore для этой даты
                    events = self._fetch_sofascore_events(date_str)
                    if not events:
                        stats['failed'] += 1
                        continue
                    
                    stats['dates_processed'] += 1
                    
                    # Получаем все матчи на эту дату без sofascore_event_id
                    # Исключаем большие поля: bet_ai_full_response, bet_ai_reason, stats_60min
                    cursor.execute("""
                        SELECT id, fixture_id, home_team, away_team, match_date, match_time, sofascore_event_id, status
                        FROM matches 
                        WHERE match_date = ? 
                        AND sofascore_event_id IS NULL
                        AND status IN ('scheduled', 'in_progress')
                                        """, (date_str,))

                    matches = cursor.fetchall()
                    print(f"[Football SofaScore] Обрабатываем {len(matches)} матчей на дату {date_str}")

                    # Первый проход: сопоставляем по двум командам
                    unmatched_matches = []
                    for match_row in matches:
                        match_dict = dict(match_row)
                        event_id = self._match_sofascore_event(match_dict, events)

                        if event_id:
                            # Обновляем sofascore_event_id в БД
                            cursor.execute("""
                                UPDATE matches
                                SET sofascore_event_id = ?, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                            """, (event_id, match_dict['id']))

                            stats['updated'] += 1
                            print(f"[Football SofaScore] Обновлен sofascore_event_id={event_id} для матча {match_dict['home_team']} vs {match_dict['away_team']}")
                        else:
                            # Сохраняем для второго прохода
                            unmatched_matches.append(match_dict)

                    conn.commit()

                    # Второй проход: сопоставляем по одной команде + времени
                    if unmatched_matches:
                        print(f"[Football SofaScore] Второй проход: ищем {len(unmatched_matches)} ненайденных матчей по команде+времени")
                        for match_dict in unmatched_matches:
                            event_data = self._match_sofascore_event_by_team_and_time(match_dict, events)

                            if event_data:
                                event_id = event_data['event_id']
                                # Формируем JSON для сохранения в sofascore_join
                                sofascore_join_data = {
                                    'slug': event_data.get('slug', ''),
                                    'startTimestamp': event_data.get('startTimestamp')
                                }
                                sofascore_join_json = json.dumps(sofascore_join_data, ensure_ascii=False)
                                
                                # Обновляем sofascore_event_id и sofascore_join в БД
                                cursor.execute("""
                                    UPDATE matches
                                    SET sofascore_event_id = ?, sofascore_join = ?, updated_at = CURRENT_TIMESTAMP
                                    WHERE id = ?
                                """, (event_id, sofascore_join_json, match_dict['id']))

                                stats['updated'] += 1
                                stats['failed'] -= 1  # Уменьшаем счетчик failed, так как теперь нашли
                                print(f"[Football SofaScore] Обновлен sofascore_event_id={event_id} для матча {match_dict['home_team']} vs {match_dict['away_team']} (второй проход)")
                            else:
                                stats['failed'] += 1
                                print(f"[Football SofaScore] Не найдено совпадение для {match_dict['home_team']} vs {match_dict['away_team']} ({match_dict['match_date']} {match_dict['match_time']})")

                    conn.commit()
                    
                    # Задержка между запросами к SofaScore (минимум 2-3 секунды для избежания блокировки)
                    # SofaScore может заблокировать IP при >5 запросов/сек или при слишком частых запросах
                    # Используем 2.5 секунды для безопасности
                    time.sleep(2.5)
                    
                except Exception as e:
                    print(f"[Football SofaScore ERROR] Ошибка обработки даты {date_str}: {e}")
                    stats['failed'] += 1
                    continue
            
            print(f"[Football SofaScore] Обновление завершено: обновлено={stats['updated']}, не найдено={stats['failed']}, дат обработано={stats['dates_processed']}")
            
        except Exception as e:
            print(f"[Football SofaScore ERROR] Критическая ошибка при обновлении sofascore_event_id: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
        
        return stats

    def sync_matches(self, leagues: Optional[List[str]] = None) -> Dict[str, int]:
        """
        Синхронизирует матчи из API с БД.
        Собирает ВСЕ матчи из указанных лиг, независимо от даты.
        Обновляет существующие матчи, удаляет матчи с коэффициентами > 1.30.
        
        Args:
            leagues: Список ключей лиг для сбора (например, ['soccer_epl', 'soccer_spain_la_liga']).
                     Если None, используется список из self.leagues.
        
        Returns:
            Словарь со статистикой: {'added': int, 'updated': int, 'deleted': int, ...}
        """
        # Используем переданный список лиг или список по умолчанию
        leagues_to_process = leagues if leagues is not None else self.leagues
        
        print(f"[Football] Начинаем синхронизацию матчей из {len(leagues_to_process)} лиг")
        
        stats = {
            'added': 0,
            'updated': 0,
            'deleted': 0,
            'skipped_no_fav': 0,
            'skipped_past': 0,
            'leagues_processed': 0,
            'leagues_failed': 0,
            'stale_closed': 0
        }
        
        now = datetime.now()
        fixture_ids_from_api = set()  # Для отслеживания матчей из API

        # Обрабатываем каждую лигу
        for league_key in leagues_to_process:
            try:
                print(f"[Football] Обрабатываем лигу: {league_key}")
                
                params = {
                    "regions": "eu",
                    "markets": "h2h",
                    "oddsFormat": "decimal"
                }
                
                                # Запрашиваем матчи для конкретной лиги
                data = self._make_api_request(f"/sports/{league_key}/odds", params)

                if not data:
                    print(f"[Football] Нет матчей для лиги {league_key} или ошибка запроса")
                    stats['leagues_failed'] += 1
                    continue
                
                print(f"[Football] Получено {len(data)} матчей из лиги {league_key}")
                stats['leagues_processed'] += 1
                
                for match_data in data:
                    fixture_id = match_data.get('id')
                    if not fixture_id:
                        continue
                    
                    fixture_ids_from_api.add(fixture_id)
                    
                    # Проверяем дату матча (пропускаем только матчи в прошлом)      
                    commence_time = match_data.get('commence_time')
                    if not commence_time:
                        continue

                    # Парсим время начала матча
                    try:
                        match_dt = datetime.fromisoformat(commence_time.replace('Z', '+00:00'))
                        match_dt = match_dt.replace(tzinfo=None)
                    except Exception as e:
                        print(f"[Football] Ошибка парсинга времени матча: {e}")     
                        continue

                    # Пропускаем матчи в прошлом
                    if match_dt < now:
                        stats['skipped_past'] += 1
                        continue

                    # Извлекаем коэффициенты 1, X, 2 для всех матчей
                    odds_1_x_2 = self._extract_odds_1_x_2(match_data)
                    
                    # Определяем фаворита
                    fav_info = self._determine_favorite(match_data)
                    
                    # Проверяем, существует ли матч в БД
                    match_exists = self._match_exists(fixture_id)
                    
                    # Определяем, есть ли фаворит с кэфом <= 1.30
                    has_favorite = fav_info is not None and fav_info['odds'] <= 1.50
                    
                    if has_favorite:
                        # Матч с фаворитом - заполняем все поля
                        if match_exists:
                            # Проверяем статус перед обновлением - не обновляем завершенные матчи
                            match_status = self._get_match_status(fixture_id)
                            if match_status == 'finished':
                                print(f"[Football] Пропущен матч с фаворитом {match_data.get('home_team')} vs {match_data.get('away_team')} - матч завершен")
                                continue
                            # Обновляем существующий матч
                            success = self._update_match(fixture_id, fav_info, match_data, odds_1_x_2)
                            if success:
                                stats['updated'] += 1
                                print(f"[Football] Обновлен матч с фаворитом {match_data.get('home_team')} vs {match_data.get('away_team')}, кэф: {fav_info['odds']}")
                        else:
                            # Добавляем новый матч
                            success = self._save_match(match_data, fav_info, odds_1_x_2)
                            if success:
                                stats['added'] += 1
                                print(f"[Football] Добавлен матч с фаворитом {match_data.get('home_team')} vs {match_data.get('away_team')}, кэф: {fav_info['odds']}")
                    else:
                        # Матч без фаворита или с кэфом > 1.30 - заполняем только базовые поля
                        if match_exists:
                            # Проверяем статус перед обновлением - не обновляем завершенные матчи
                            match_status = self._get_match_status(fixture_id)
                            if match_status == 'finished':
                                print(f"[Football] Пропущен матч без фаворита {match_data.get('home_team')} vs {match_data.get('away_team')} - матч завершен")
                                continue
                            # Обновляем существующий матч (без fav)
                            success = self._update_match_without_fav(fixture_id, match_data, odds_1_x_2)
                            if success:
                                stats['updated'] += 1
                                print(f"[Football] Обновлен матч без фаворита {match_data.get('home_team')} vs {match_data.get('away_team')}")
                        else:
                            # Добавляем новый матч (без fav)
                            success = self._save_match_without_fav(match_data, odds_1_x_2)
                            if success:
                                stats['added'] += 1
                                print(f"[Football] Добавлен матч без фаворита {match_data.get('home_team')} vs {match_data.get('away_team')}")
                
            except Exception as e:
                print(f"[Football ERROR] Ошибка при обработке лиги {league_key}: {e}")
                stats['leagues_failed'] += 1
                continue

        # Удаляем матчи из БД, которых больше нет в API (опционально, если нужно)
        # Пока не реализовано, так как API может не возвращать все матчи

        print(f"[Football] Синхронизация завершена: лиг обработано={stats['leagues_processed']}, лиг с ошибками={stats['leagues_failed']}, добавлено={stats['added']}, обновлено={stats['updated']}, удалено={stats['deleted']}, пропущено (прошлое)={stats['skipped_past']}")
        
        # Обновляем sofascore_event_id для матчей без него
        print("[Football] Начинаем обновление sofascore_event_id...")
        sofascore_stats = self.update_sofascore_ids()
        stats['sofascore_updated'] = sofascore_stats['updated']
        stats['sofascore_failed'] = sofascore_stats['failed']
        stats['sofascore_dates_processed'] = sofascore_stats['dates_processed']

        # Принудительно закрываем старые матчи, которые так и не завершились
        stale_closed = self._close_stale_matches()
        stats['stale_closed'] = stale_closed
        if stale_closed:
            print(f"[Football] Принудительно закрыто {stale_closed} матчей со статусом 'finished' (старше 20 часов)")
        else:
            print("[Football] Просроченных матчей для принудительного закрытия не найдено")
        
        return stats

    def collect_tomorrow_matches(self) -> int:
        """
        Алиас для sync_matches для обратной совместимости.
        Теперь использует sync_matches.
        
        Returns:
            Количество добавленных матчей
        """
        stats = self.sync_matches()
        return stats['added']

    def _extract_odds_1_x_2(self, match_data: Dict) -> Optional[Dict[str, float]]:
        """
        Извлекает медианные коэффициенты для исходов 1, X, 2 из данных матча.
        
        Args:
            match_data: Данные матча от API (уже содержат bookmakers)
            
        Returns:
            Словарь с коэффициентами: {'odds_1': float, 'odds_x': float, 'odds_2': float} или None
        """
        try:
            home_team = match_data.get('home_team')
            away_team = match_data.get('away_team')
            bookmakers = match_data.get('bookmakers', [])
            
            if not home_team or not away_team or not bookmakers:
                return None
            
            # Собираем коэффициенты для каждой команды и ничьей
            home_odds = []
            away_odds = []
            draw_odds = []
            
            for bookmaker in bookmakers:
                markets = bookmaker.get('markets', [])
                for market in markets:
                    if market.get('key') != 'h2h':
                        continue
                    
                    outcomes = market.get('outcomes', [])
                    for outcome in outcomes:
                        name = outcome.get('name')
                        price = outcome.get('price')
                        
                        if not price or not name:
                            continue
                        
                        if name == home_team:
                            home_odds.append(float(price))
                        elif name == away_team:
                            away_odds.append(float(price))
                        elif name.lower() == 'draw':
                            draw_odds.append(float(price))
            
            if not home_odds or not away_odds or not draw_odds:
                return None
            
            # Вычисляем медианные коэффициенты
            def get_median(odds_list):
                n = len(odds_list)
                if n == 0:
                    return None
                sorted_odds = sorted(odds_list)
                if n % 2 == 0:
                    return (sorted_odds[n//2 - 1] + sorted_odds[n//2]) / 2.0
                else:
                    return sorted_odds[n//2]
            
            odds_1 = get_median(home_odds)
            odds_x = get_median(draw_odds)
            odds_2 = get_median(away_odds)
            
            if odds_1 is None or odds_x is None or odds_2 is None:
                return None
            
            return {
                'odds_1': odds_1,
                'odds_x': odds_x,
                'odds_2': odds_2
            }
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка извлечения коэффициентов 1, X, 2: {e}")
            return None

    def _determine_favorite(self, match_data: Dict) -> Optional[Dict]:
        """
        Определяет фаворита по коэффициентам из The Odds API.
        
        Args:
            match_data: Данные матча от API (уже содержат bookmakers)
            
        Returns:
            Словарь с информацией о фаворите: {'team', 'is_home', 'odds'} или None
        """
        try:
            home_team = match_data.get('home_team')
            away_team = match_data.get('away_team')
            
            if not home_team or not away_team:
                print("[Football] Нет названий команд в данных матча")
                return None
            
            # Получаем коэффициенты из bookmakers
            bookmakers = match_data.get('bookmakers', [])
            if not bookmakers:
                print("[Football] Нет букмекеров в данных")
                return None
            
            # Собираем все коэффициенты для каждой команды по всем букмекерам
            home_odds = []
            away_odds = []
            
            for bookmaker in bookmakers:
                markets = bookmaker.get('markets', [])
                for market in markets:
                    if market.get('key') != 'h2h':
                        continue
                    
                    outcomes = market.get('outcomes', [])
                    for outcome in outcomes:
                        name = outcome.get('name')
                        price = outcome.get('price')
                        
                        if not price or not name:
                            continue
                        
                        # Пропускаем Draw
                        if name.lower() == 'draw':
                            continue
                        
                        # Определяем команду по имени
                        if name == home_team:
                            home_odds.append(float(price))
                        elif name == away_team:
                            away_odds.append(float(price))
            
            if not home_odds or not away_odds:
                print(f"[Football] Не удалось получить коэффициенты для команд")
                return None
            
            # Сортируем коэффициенты для расчета медианы
            home_odds_sorted = sorted(home_odds)
            away_odds_sorted = sorted(away_odds)
            
            # Берем медианный коэффициент для каждой команды (устойчив к выбросам)
            def get_median(odds_list):
                n = len(odds_list)
                if n == 0:
                    return None
                if n % 2 == 0:
                    return (odds_list[n//2 - 1] + odds_list[n//2]) / 2.0
                else:
                    return odds_list[n//2]
            
            median_home_odd = get_median(home_odds_sorted)
            median_away_odd = get_median(away_odds_sorted)
            
            if median_home_odd is None or median_away_odd is None:
                print(f"[Football] Не удалось рассчитать медианные коэффициенты")
                return None
            
            # Определяем фаворита (меньший медианный коэффициент)
            if median_home_odd <= median_away_odd:
                fav_team = home_team
                fav_is_home = True
                fav_odd = median_home_odd
            else:
                fav_team = away_team
                fav_is_home = False
                fav_odd = median_away_odd
            
            print(f"[Football] Фаворит: {fav_team} (кэф: {fav_odd})")
            
            return {
                'team': fav_team,
                'is_home': fav_is_home,
                'odds': fav_odd
            }
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка определения фаворита: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def _save_match(self, match_data: Dict, fav_info: Dict, odds_1_x_2: Optional[Dict[str, float]] = None) -> bool:
        """
        Сохраняет матч в БД с фаворитом.
        
        Args:
            match_data: Данные матча от The Odds API
            fav_info: Информация о фаворите
            odds_1_x_2: Словарь с коэффициентами {'odds_1': float, 'odds_x': float, 'odds_2': float}
            
        Returns:
            True если успешно, False если ошибка
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            
            # The Odds API использует "id" вместо "fixture_id"
            event_id = match_data.get('id')
            
            # Проверяем, не существует ли уже матч
            cursor.execute("SELECT id FROM matches WHERE fixture_id = ?", (event_id,))
            if cursor.fetchone():
                print(f"[Football] Матч {event_id} уже существует, пропускаем")
                return False
            
            # Извлекаем данные
            home_team = match_data.get('home_team')
            away_team = match_data.get('away_team')
            sport_key = match_data.get('sport_key')

            if not home_team or not away_team:
                print(f"[Football] Нет названий команд для матча {event_id}")
                return False

            # Дата и время матча (в UTC от The Odds API)
            commence_time = match_data.get('commence_time')
            if commence_time:
                # Парсим UTC время от The Odds API
                dt = datetime.fromisoformat(commence_time.replace('Z', '+00:00'))
                # Сохраняем в UTC (без tzinfo для совместимости с текстовыми полями БД)
                # ВАЖНО: время в БД хранится в UTC, при чтении нужно добавлять timezone.utc
                dt = dt.replace(tzinfo=None)
                match_date = dt.strftime('%Y-%m-%d')
                match_time = dt.strftime('%H:%M')
            else:
                print(f"[Football] Нет даты для матча {event_id}")        
                return False

            # Сохраняем
            # При первом сохранении initial_odds и last_odds одинаковые   
            fav_odds = fav_info['odds']
            
            # Извлекаем коэффициенты 1, X, 2
            live_odds_1 = odds_1_x_2.get('odds_1') if odds_1_x_2 else None
            live_odds_x = odds_1_x_2.get('odds_x') if odds_1_x_2 else None
            live_odds_2 = odds_1_x_2.get('odds_2') if odds_1_x_2 else None
            
            cursor.execute("""
                INSERT INTO matches
                (fixture_id, home_team, away_team, fav, fav_team_id,      
                 match_date, match_time, initial_odds, last_odds, status, sport_key,
                 live_odds_1, live_odds_x, live_odds_2) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                event_id,  # fixture_id = event_id из The Odds API        
                home_team,
                away_team,
                fav_info['team'],
                1 if fav_info['is_home'] else 0,  # fav_team_id: 1=home, 0=away
                match_date,
                match_time,
                fav_odds,  # initial_odds - первая котировка
                fav_odds,  # last_odds - при первом сохранении такая же   
                'scheduled',
                sport_key,  # sport_key для использования в запросах live odds
                live_odds_1,
                live_odds_x,
                live_odds_2
            ))

            conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка сохранения матча: {e}")
            return False
        except Exception as e:
            print(f"[Football ERROR] Неожиданная ошибка: {e}")
            import traceback
            print(traceback.format_exc())
            return False
        finally:
            if conn:
                conn.close()

    def _match_exists(self, fixture_id: str) -> bool:
        """
        Проверяет, существует ли матч в БД.

        Args:
            fixture_id: ID матча из API

        Returns:
            True если матч существует, False если нет
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM matches WHERE fixture_id = ?", (fixture_id,))
            return cursor.fetchone() is not None
        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка проверки существования матча: {e}")
            return False
        finally:
            if conn:
                conn.close()

    def _close_stale_matches(self, older_than_hours: int = 20) -> int:
        """
        Принудительно закрывает матчи, которые должны были начаться давно, но до сих пор не имеют статуса finished.

        Args:
            older_than_hours: Количество часов, прошедших с предполагаемого начала матча

        Returns:
            Количество матчей, статус которых был обновлён
        """
        conn = None
        closed = 0

        try:
            cutoff = datetime.now(timezone.utc) - timedelta(hours=older_than_hours)

            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, fixture_id, match_date, match_time
                FROM matches
                WHERE status != 'finished'
                  AND match_date IS NOT NULL
                  AND match_time IS NOT NULL
            """)

            rows = cursor.fetchall()
            if not rows:
                return 0

            for row in rows:
                try:
                    match_datetime_str = f"{row['match_date']} {row['match_time']}"
                    match_datetime = datetime.strptime(match_datetime_str, "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
                except (ValueError, TypeError):
                    continue

                if match_datetime <= cutoff:
                    cursor.execute(
                        """
                        UPDATE matches
                        SET status = 'finished',
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (row['id'],)
                    )
                    closed += 1

            if closed:
                conn.commit()

            return closed

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка принудительного закрытия матчей: {e}")
            return 0
        finally:
            if conn:
                conn.close()

    def _get_match_status(self, fixture_id: str) -> Optional[str]:
        """
        Получает статус матча из БД.

        Args:
            fixture_id: ID матча из API

        Returns:
            Статус матча ('scheduled', 'in_progress', 'finished') или None если матч не найден
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT status FROM matches WHERE fixture_id = ?", (fixture_id,))
            result = cursor.fetchone()
            return result[0] if result else None
        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка получения статуса матча: {e}")
            return None
        finally:
            if conn:
                conn.close()

    def _get_match_bet_value(self, fixture_id: str) -> Optional[int]:
        """
        Получает значение bet для матча.

        Args:
            fixture_id: ID матча из API

        Returns:
            Значение bet или None, если матч не найден или bet не установлен
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT bet FROM matches WHERE fixture_id = ?", (fixture_id,))
            row = cursor.fetchone()
            if row and row[0] is not None:
                return row[0]
            return None
        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка получения bet для матча: {e}")
            return None
        finally:
            if conn:
                conn.close()

    def _update_match(self, fixture_id: str, fav_info: Dict, match_data: Dict, odds_1_x_2: Optional[Dict[str, float]] = None) -> bool:
        """
        Обновляет коэффициент существующего матча с фаворитом.

        Args:
            fixture_id: ID матча из API
            fav_info: Информация о фаворите
            match_data: Данные матча от API
            odds_1_x_2: Словарь с коэффициентами {'odds_1': float, 'odds_x': float, 'odds_2': float}

        Returns:
            True если успешно, False если ошибка
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()

            # Обновляем только коэффициент (last_odds), фаворита, sport_key, коэффициенты 1/X/2 и время обновления
            # initial_odds не трогаем - там хранится первая котировка
            sport_key = match_data.get('sport_key')
            
            # Извлекаем коэффициенты 1, X, 2
            live_odds_1 = odds_1_x_2.get('odds_1') if odds_1_x_2 else None
            live_odds_x = odds_1_x_2.get('odds_x') if odds_1_x_2 else None
            live_odds_2 = odds_1_x_2.get('odds_2') if odds_1_x_2 else None
            
            cursor.execute("""
                UPDATE matches
                SET fav = ?, fav_team_id = ?, last_odds = ?, sport_key = ?,
                    live_odds_1 = ?, live_odds_x = ?, live_odds_2 = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE fixture_id = ?
            """, (
                fav_info['team'],
                1 if fav_info['is_home'] else 0,
                fav_info['odds'],
                sport_key,
                live_odds_1,
                live_odds_x,
                live_odds_2,
                fixture_id
            ))

            conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка обновления матча: {e}")
            return False
        except Exception as e:
            print(f"[Football ERROR] Неожиданная ошибка при обновлении: {e}")
            import traceback
            print(traceback.format_exc())
            return False
        finally:
            if conn:
                conn.close()

    def _save_match_without_fav(self, match_data: Dict, odds_1_x_2: Optional[Dict[str, float]] = None) -> bool:
        """
        Сохраняет матч в БД без фаворита (только базовые поля и коэффициенты 1, X, 2).
        
        Args:
            match_data: Данные матча от The Odds API
            odds_1_x_2: Словарь с коэффициентами {'odds_1': float, 'odds_x': float, 'odds_2': float}
            
        Returns:
            True если успешно, False если ошибка
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()
            
            # The Odds API использует "id" вместо "fixture_id"
            event_id = match_data.get('id')
            
            # Проверяем, не существует ли уже матч
            cursor.execute("SELECT id FROM matches WHERE fixture_id = ?", (event_id,))
            if cursor.fetchone():
                print(f"[Football] Матч {event_id} уже существует, пропускаем")
                return False
            
            # Извлекаем данные
            home_team = match_data.get('home_team')
            away_team = match_data.get('away_team')
            sport_key = match_data.get('sport_key')

            if not home_team or not away_team:
                print(f"[Football] Нет названий команд для матча {event_id}")
                return False

            # Дата и время матча (в UTC от The Odds API)
            commence_time = match_data.get('commence_time')
            if commence_time:
                # Парсим UTC время от The Odds API
                dt = datetime.fromisoformat(commence_time.replace('Z', '+00:00'))
                dt = dt.replace(tzinfo=None)
                match_date = dt.strftime('%Y-%m-%d')
                match_time = dt.strftime('%H:%M')
            else:
                print(f"[Football] Нет даты для матча {event_id}")        
                return False

            # Извлекаем коэффициенты 1, X, 2
            live_odds_1 = odds_1_x_2.get('odds_1') if odds_1_x_2 else None
            live_odds_x = odds_1_x_2.get('odds_x') if odds_1_x_2 else None
            live_odds_2 = odds_1_x_2.get('odds_2') if odds_1_x_2 else None
            
            # Сохраняем только базовые поля (fav = 'NONE', fav_team_id = -1, initial_odds, last_odds остаются NULL)
            cursor.execute("""
                INSERT INTO matches
                (fixture_id, home_team, away_team, fav, fav_team_id, match_date, match_time, status, sport_key,
                 live_odds_1, live_odds_x, live_odds_2) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                event_id,
                home_team,
                away_team,
                'NONE',  # Специальное значение вместо NULL
                -1,  # Специальное значение вместо NULL
                match_date,
                match_time,
                'scheduled',
                sport_key,
                live_odds_1,
                live_odds_x,
                live_odds_2
            ))

            conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка сохранения матча без фаворита: {e}")
            return False
        except Exception as e:
            print(f"[Football ERROR] Неожиданная ошибка: {e}")
            import traceback
            print(traceback.format_exc())
            return False
        finally:
            if conn:
                conn.close()

    def _update_match_without_fav(self, fixture_id: str, match_data: Dict, odds_1_x_2: Optional[Dict[str, float]] = None) -> bool:
        """
        Обновляет матч без фаворита (только базовые поля и коэффициенты 1, X, 2).

        Args:
            fixture_id: ID матча из API
            match_data: Данные матча от API
            odds_1_x_2: Словарь с коэффициентами {'odds_1': float, 'odds_x': float, 'odds_2': float}

        Returns:
            True если успешно, False если ошибка
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()

            sport_key = match_data.get('sport_key')
            
            # Извлекаем коэффициенты 1, X, 2
            live_odds_1 = odds_1_x_2.get('odds_1') if odds_1_x_2 else None
            live_odds_x = odds_1_x_2.get('odds_x') if odds_1_x_2 else None
            live_odds_2 = odds_1_x_2.get('odds_2') if odds_1_x_2 else None
            
            # Обновляем только базовые поля и коэффициенты (fav, initial_odds, last_odds не трогаем)
            cursor.execute("""
                UPDATE matches
                SET sport_key = ?, live_odds_1 = ?, live_odds_x = ?, live_odds_2 = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE fixture_id = ?
            """, (
                sport_key,
                live_odds_1,
                live_odds_x,
                live_odds_2,
                fixture_id
            ))

            conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка обновления матча без фаворита: {e}")
            return False
        except Exception as e:
            print(f"[Football ERROR] Неожиданная ошибка при обновлении: {e}")
            import traceback
            print(traceback.format_exc())
            return False
        finally:
            if conn:
                conn.close()

    def _delete_match(self, fixture_id: str) -> bool:
        """
        Удаляет матч из БД.

        Args:
            fixture_id: ID матча из API

        Returns:
            True если успешно, False если ошибка
        """
        conn = None
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()

            cursor.execute("DELETE FROM matches WHERE fixture_id = ?", (fixture_id,))
            conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"[Football ERROR] Ошибка удаления матча: {e}")
            return False
        except Exception as e:
            print(f"[Football ERROR] Неожиданная ошибка при удалении: {e}")
            import traceback
            print(traceback.format_exc())
            return False
        finally:
            if conn:
                conn.close()

    def check_matches_and_collect(self):
        """
        Проверяет активные матчи и собирает финальный результат.
        Вызывается каждые 5 минут.
        Обработка 60-й минуты вынесена в отдельный метод check_matches_60min_and_status (3-минутный интервал).
        """
        print("[Football] Проверка матчей и сбор финального результата")

        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()

            # ===== ЧАСТЬ 1.5: Обновление live_odds для уже обработанных матчей без live_odds =====
            # Исключаем большие поля: bet_ai_full_response, bet_ai_reason, stats_60min
            cursor.execute("""
                SELECT id, fixture_id, match_date, match_time, sport_key
                FROM matches
                WHERE status = 'in_progress'
                AND bet IS NOT NULL
                AND live_odds IS NULL
                ORDER BY match_date, match_time
            """)
            
            matches_for_live_odds = cursor.fetchall()
            print(f"[Football] Найдено {len(matches_for_live_odds)} матчей с bet, но без live_odds для обновления")
            
            for match in matches_for_live_odds:
                try:
                    fixture_id = match['fixture_id']
                    match_datetime_str = f"{match['match_date']} {match['match_time']}"
                    match_datetime_naive = datetime.strptime(match_datetime_str, "%Y-%m-%d %H:%M")
                    match_datetime = match_datetime_naive.replace(tzinfo=timezone.utc)
                    now = datetime.now(timezone.utc)
                    minutes_diff = (now - match_datetime).total_seconds() / 60.0
                    
                    # Обновляем live_odds только если прошло >= 50 минут
                    if minutes_diff >= 50:
                        print(f"[Football] Обновляем live_odds для матча {fixture_id} (прошло {minutes_diff:.1f} минут)...")
                        sport_key = match['sport_key'] if 'sport_key' in match.keys() else None
                        live_odds_value = self._get_live_odds(fixture_id, sport_key)
                        if live_odds_value:
                            print(f"[Football] Получены live odds для {fixture_id}: {live_odds_value}")
                            cursor.execute("""
                                UPDATE matches
                                SET live_odds = ?, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                            """, (live_odds_value, match['id']))
                            conn.commit()
                        else:
                            print(f"[Football] Не удалось получить live odds для {fixture_id}")
                except Exception as e:
                    print(f"[Football ERROR] Ошибка обновления live_odds для {match['fixture_id']}: {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            # ===== ЧАСТЬ 1.7: Второй проход риск-менеджмента (ЗАКОММЕНТИРОВАН, НЕ ИСПОЛЬЗУЕТСЯ) =====
            # try:
            #     self._perform_bet_approval_checks(cursor, conn)
            # except Exception as e:
            #     print(f"[Football ERROR] Ошибка выполнения второго прохода риск-менеджмента: {e}")
            #     import traceback
            #     print(traceback.format_exc())

            # ===== ЧАСТЬ 2: Сбор финального результата (для всех матчей in_progress, независимо от bet) =====
            # Исключаем большие поля: bet_ai_full_response, bet_ai_reason, stats_60min
            cursor.execute("""
                SELECT id, fixture_id, sofascore_event_id, match_date, match_time, status
                FROM matches
                WHERE status = 'in_progress'
                ORDER BY match_date, match_time
            """)

            matches_for_final = cursor.fetchall()
            print(f"[Football] Найдено {len(matches_for_final)} матчей in_progress для проверки финального результата")

            # Проверяем каждый матч на завершение
            for match in matches_for_final:
                match_id = match['id']
                fixture_id = match['fixture_id']
                sofascore_event_id = match['sofascore_event_id'] if 'sofascore_event_id' in match.keys() and match['sofascore_event_id'] else None
                match_datetime_str = f"{match['match_date']} {match['match_time']}"

                try:
                    # Парсим дату и время из БД (они в UTC, но без tzinfo)
                    match_datetime_naive = datetime.strptime(match_datetime_str, "%Y-%m-%d %H:%M")
                    # Добавляем UTC часовой пояс, так как время в БД сохранено в UTC
                    match_datetime = match_datetime_naive.replace(tzinfo=timezone.utc)

                    # Используем UTC время для сравнения (независимо от часового пояса сервера)
                    now = datetime.now(timezone.utc)

                    # Вычисляем разницу во времени
                    time_diff = now - match_datetime
                    minutes_diff = time_diff.total_seconds() / 60

                    # Проверяем статус матча из SofaScore API (предпочтительный способ)
                    # Вызываем только если прошло минимум 100 минут (90 минут игры + ~15 минут перерыва)
                    # Матч должен быть близок к завершению или уже завершен
                    should_check_final = False
                    
                    if sofascore_event_id and minutes_diff >= 100:
                        event_status = self._fetch_sofascore_event_status(sofascore_event_id)
                        
                        if event_status == 'finished':
                            should_check_final = True
                        elif event_status:
                            # Если матч не завершен по статусу, но прошло много времени - используем запасной вариант
                            if minutes_diff >= 200:
                                should_check_final = True
                        else:
                            # Если не удалось получить статус из API, используем проверку по времени
                            if minutes_diff >= 200:
                                should_check_final = True
                    elif not sofascore_event_id:
                        # Если нет sofascore_event_id, используем только проверку по времени
                        if minutes_diff >= 200:
                            should_check_final = True
                    elif minutes_diff < 100:
                        # Матч еще слишком рано (меньше 100 минут) - не проверяем статус из API
                        pass

                    if should_check_final:
                        try:
                            self._collect_final_result(match)
                        except Exception as e:
                            print(f"[Football ERROR] Ошибка сбора финального результата для {fixture_id}: {e}")
                            import traceback
                            print(traceback.format_exc())

                except Exception as e:
                    print(f"[Football ERROR] Ошибка проверки финального результата для {fixture_id}: {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            conn.close()
            print(f"[Football] Обработка матчей завершена. Проверено на финальный результат: {len(matches_for_final)}")
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка проверки матчей: {e}")
            import traceback
            print(traceback.format_exc())
        finally:
            if conn:
                conn.close()

    def check_matches_60min_and_status(self):
        """
        Проверяет активные матчи только для смены статуса и сбора статистики на 60-й минуте (без проверки финального счета).
        Используется для более частого (например, каждые 3 минуты) детектора 60-й минуты.
        """
        print("[Football] (3-мин) Проверка статуса и 60-й минуты")
        try:
            conn = get_football_db_connection()
            cursor = conn.cursor()

            # Матчи с фаворитом, еще не обработанные (bet IS NULL)
            # Исключаем большие поля: bet_ai_full_response, bet_ai_reason, stats_60min
            # Но включаем sofascore_event_id и sport_key, так как они нужны для _collect_60min_stats
            cursor.execute("""
                SELECT id, fixture_id, sofascore_event_id, sport_key, match_date, match_time, status
                FROM matches
                WHERE status IN ('scheduled', 'in_progress')
                  AND bet IS NULL
                  AND fav != 'NONE'
                ORDER BY match_date, match_time
            """)
            matches_with_fav = cursor.fetchall()

            # Матчи без фаворита, еще не обработанные (bet IS NULL)
            # Исключаем большие поля: bet_ai_full_response, bet_ai_reason, stats_60min
            # Но включаем sofascore_event_id и sport_key, так как они нужны для _collect_60min_stats_without_fav
            cursor.execute("""
            SELECT id, fixture_id, sofascore_event_id, sport_key, match_date, match_time, status
                FROM matches
                WHERE status IN ('scheduled', 'in_progress')
                  AND bet IS NULL
                  AND fav = 'NONE'
            ORDER BY match_date, match_time
            """)
            matches_without_fav = cursor.fetchall()

            # Обработка матчей с фаворитом
            for match in matches_with_fav:
                match_id = match['id']
                fixture_id = match['fixture_id']
                match_datetime_str = f"{match['match_date']} {match['match_time']}"
                try:
                    match_datetime_naive = datetime.strptime(match_datetime_str, "%Y-%m-%d %H:%M")
                    match_datetime = match_datetime_naive.replace(tzinfo=timezone.utc)
                    now = datetime.now(timezone.utc)
                    minutes_diff = (now - match_datetime).total_seconds() / 60

                    if minutes_diff < 0:
                        continue

                    if match['status'] == 'scheduled':
                        cursor.execute(
                            "UPDATE matches SET status = 'in_progress', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                            (match_id,)
                        )
                        conn.commit()

                    if minutes_diff >= 50:
                        try:
                            self._collect_60min_stats(match)
                        except Exception as e:
                            print(f"[Football ERROR] Ошибка сбора статистики 60min для {fixture_id}: {e}")
                            import traceback
                            print(traceback.format_exc())
                            cursor.execute(
                                "UPDATE matches SET bet = 0, bet_approve = NULL, bet_approve_reason = NULL, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                                (match_id,)
                            )
                            conn.commit()
                except Exception as e:
                    print(f"[Football ERROR] Ошибка обработки матча {fixture_id} (fav): {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            # Обработка матчей без фаворита
            for match in matches_without_fav:
                match_id = match['id']
                fixture_id = match['fixture_id']
                match_datetime_str = f"{match['match_date']} {match['match_time']}"
                try:
                    match_datetime_naive = datetime.strptime(match_datetime_str, "%Y-%m-%d %H:%M")
                    match_datetime = match_datetime_naive.replace(tzinfo=timezone.utc)
                    now = datetime.now(timezone.utc)
                    minutes_diff = (now - match_datetime).total_seconds() / 60

                    if minutes_diff < 0:
                        continue

                    if match['status'] == 'scheduled':
                        cursor.execute(
                            "UPDATE matches SET status = 'in_progress', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                            (match_id,)
                        )
                        conn.commit()

                    if minutes_diff >= 50:
                        try:
                            self._collect_60min_stats_without_fav(match)
                        except Exception as e:
                            print(f"[Football ERROR] Ошибка сбора статистики 60min (без фаворита) для {fixture_id}: {e}")
                            import traceback
                            print(traceback.format_exc())
                            cursor.execute(
                                "UPDATE matches SET bet = 0, bet_approve = NULL, bet_approve_reason = NULL, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                                (match_id,)
                            )
                            conn.commit()
                except Exception as e:
                    print(f"[Football ERROR] Ошибка обработки матча {fixture_id} (no fav): {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            # Проверяем матчи с stats_60min, но без bet_alt_code (для запроса альтернативной ставки)
            # Только для матчей в процессе, не для завершенных!
            # Включаем stats_60min, но исключаем bet_ai_full_response
            cursor.execute("""
                SELECT id, fixture_id, match_date, match_time, stats_60min, bet_alt_code
                FROM matches
                WHERE stats_60min IS NOT NULL
                  AND (bet_alt_code IS NULL OR bet_alt_code = '')
                  AND status = 'in_progress'
                ORDER BY match_date, match_time
            """)
            matches_for_alt_bet = cursor.fetchall()
            
            if matches_for_alt_bet:
                print(f"[Football] Найдено {len(matches_for_alt_bet)} матчей с stats_60min, но без bet_alt_code")
                import time
                for idx, match in enumerate(matches_for_alt_bet, 1):
                    fixture_id = match['fixture_id']
                    try:
                        import json
                        stats = json.loads(match['stats_60min']) if isinstance(match['stats_60min'], str) else match['stats_60min']
                        
                        start_time = time.time()
                        print(f"[Football] [{idx}/{len(matches_for_alt_bet)}] Запрашиваем альтернативную ставку для fixture {fixture_id} (время начала: {time.strftime('%H:%M:%S')})")
                        alt_result = self._get_alternative_bet(match, stats)
                        elapsed = time.time() - start_time
                        print(f"[Football] [{idx}/{len(matches_for_alt_bet)}] Запрос для fixture {fixture_id} завершен за {elapsed:.2f} сек")
                        if alt_result:
                            bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason = alt_result
                            cursor.execute("""
                                UPDATE matches
                                SET bet_alt_code = ?,
                                    bet_alt_odds = ?,
                                    bet_alt_confirm = ?,
                                    bet_ai_reason = ?,
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                            """, (bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason if bet_alt_reason else None, match['id']))
                            conn.commit()
                            print(f"[Football] Альтернативная ставка сохранена для fixture {fixture_id}: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                        else:
                            print(f"[Football] _get_alternative_bet вернул None для fixture {fixture_id}")
                    except Exception as e:
                        print(f"[Football ERROR] Ошибка получения альтернативной ставки для fixture {fixture_id}: {e}")
                        import traceback
                        traceback.print_exc()
                        continue

            conn.close()
        except Exception as e:
            print(f"[Football ERROR] Ошибка 3-мин проверки: {e}")
            import traceback
            print(traceback.format_exc())
        finally:
            try:
                if conn:
                    conn.close()
            except:
                pass

    def _parse_confirm_value(self, confirm) -> int:
        """
        Умно парсит значение confirm из различных форматов.
        
        Args:
            confirm: Значение confirm (может быть int, bool, str, None)
        
        Returns:
            1 если confirm = True/1/"true"/"yes"/"да" и т.д., 0 в остальных случаях
        """
        if confirm is None:
            return 0
        
        # Если это число
        if isinstance(confirm, (int, float)):
            return 1 if confirm == 1 else 0
        
        # Если это булево значение
        if isinstance(confirm, bool):
            return 1 if confirm is True else 0
        
        # Если это строка
        if isinstance(confirm, str):
            confirm_lower = confirm.lower().strip()
            # Положительные значения
            if confirm_lower in ['1', 'true', 'yes', 'да', 'yes.', 'да.', 'true.', '1.']:
                return 1
            # Отрицательные значения (для ясности, хотя по умолчанию будет 0)
            if confirm_lower in ['0', 'false', 'no', 'нет', 'false.', 'no.', 'нет.', '0.']:
                return 0
        
        # Если не распознано, возвращаем 0 (безопасный выбор)
        return 0
    
    def _encode_alternative_bet(self, market: str, pick: str, line: Optional[float] = None) -> str:
        """
        Преобразует JSON ответ модели в короткую кодировку ставки.
        
        Args:
            market: Рынок (1X2, DoubleChance, Handicap, Total)
            pick: Выбор (1/X/2/1X/X2/Home/Away/Over/Under)
            line: Линия (для Handicap/Total)
        
        Returns:
            Кодировка: Ф1-1.5, Ф2+2.5, Т2.5Б, Т0.5М, 1, X, 2, 1X, X2, 12
        """
        if market == "Handicap":
            # Гандикап: Ф1-1.5 (фора хозяев -1.5), Ф2+2.5 (фора гостей +2.5)
            if pick == "Home":
                sign = "-" if line and line < 0 else "+"
                line_str = f"{abs(line):.1f}" if line else "0.0"
                return f"Ф1{sign}{line_str}"
            elif pick == "Away":
                sign = "+" if line and line > 0 else "-"
                line_str = f"{abs(line):.1f}" if line else "0.0"
                return f"Ф2{sign}{line_str}"
        elif market == "Total":
            # Тотал: Т2.5Б (больше 2.5), Т0.5М (меньше 0.5)
            if pick == "Over":
                line_str = f"{line:.1f}" if line else "2.5"
                return f"Т{line_str}Б"
            elif pick == "Under":
                line_str = f"{line:.1f}" if line else "2.5"
                return f"Т{line_str}М"
        elif market == "1X2":
            # 1X2: 1, X, 2
            return pick
        elif market == "DoubleChance":
            # Двойной шанс: 1X, X2, 12
            return pick
        
        # Если не распознано, возвращаем исходный pick
        return pick

    def _get_alternative_bet(self, match: sqlite3.Row, stats: Dict) -> Optional[Tuple[str, float, int, str]]:
        """
        Получает альтернативную ставку от ИИ для одного матча.
        
        Args:
            match: Запись матча из БД
            stats: Статистика матча
        
        Returns:
            Tuple (bet_alt_code, bet_alt_odds, bet_alt_confirm, reason) или None
        """
        if not self.openrouter_api_key:
            print("[Football Alt Bet] OpenRouter API ключ не установлен, пропускаем")
            return None
        
        try:
            fixture_id = match['fixture_id']
            home_team = match['home_team']
            away_team = match['away_team']
            
            # Получаем коэффициенты 1X2
            live_odds_1 = match['live_odds_1'] if 'live_odds_1' in match.keys() else None
            live_odds_x = match['live_odds_x'] if 'live_odds_x' in match.keys() else None
            live_odds_2 = match['live_odds_2'] if 'live_odds_2' in match.keys() else None
            
            # Формируем промпт для одного матча
            # Примечание: текущий счет находится внутри stats_60min['score'] = {'home': X, 'away': Y}
            match_data = {
                'fixture_id': fixture_id,
                'home_team': home_team,
                'away_team': away_team,
                'live_odds_1': live_odds_1,
                'live_odds_x': live_odds_x,
                'live_odds_2': live_odds_2,
                'stats_60min': stats
            }
            
            context_json = json.dumps({'match': match_data}, ensure_ascii=False)
            
            system_instruction = (
                "Ты - аналитик футбольных матчей и эксперт в области спортивных ставок. "
                "Тебе предоставлена статистика первой половины матча. "
                "Твоя задача - выбрать ОДНУ оптимальную ставку из следующих рынков: 1X2, DoubleChance, Handicap, Total. "
                "Ты должен учитывать статистику первой половины матча, текущий счет, текущие коэффициенты букмекеров и другие факторы (в том числе исторические и статистические данные). "
                "Для Handicap используй стороны Home/Away и ТОЛЬКО половинные линии (…,-2.5,-2.0,-1.5,-1.0,-0.5,0,+0.5,+1.0,+1.5,+2.0,+2.5,…); никаких четвертных (0.25/0.75). "
                "Для Total используй Over/Under с ТОЛЬКО половинными линиями (0, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5 …). Размер линий не ограничивай. "
                "Если точного коэффициента нет, оцени приблизительно на основе темпа/статистики и live_odds_1/x/2, округли до двух знаков и проставь odds_estimated=true. "
                "Верни СТРОГО JSON (без текста вокруг) формата: "
                "{\"market\":\"1X2|DoubleChance|Handicap|Total\",\"pick\":\"1|X|2|1X|X2|Home|Away|Over|Under\",\"line\":number|null,\"odds\":number,\"odds_estimated\":boolean,\"reason\":str}."
            )
            
            prompt = f"{system_instruction}\n\nДанные:\n{context_json}"
            
            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": os.getenv("OPENROUTER_SITE_URL", "http://localhost:5000")
            }
            
            models_to_try = [self.ai_primary_model, self.ai_fallback_model1, self.ai_fallback_model2, self.ai_fallback_model3]
            
            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue
                
                print(f"[Football Alt Bet] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model} для fixture {fixture_id}")
                
                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {"role": "user", "content": prompt}
                        ],
                        "max_tokens": 2000,
                        "temperature": 0.4
                    }
                    
                    print(f"[Football Alt Bet] Отправка запроса к OpenRouter API (модель: {model})")
                    print(f"[Football Alt Bet] URL: {self.openrouter_api_url}/chat/completions")
                    
                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=300
                    )
                    
                    print(f"[Football Alt Bet] Ответ от модели {model}: статус {response.status_code}")
                    
                    if response.status_code == 200:
                        data = response.json()
                        if 'choices' in data and data['choices']:
                            raw = data['choices'][0]['message']['content']
                            print(f"[Football Alt Bet] Получен ответ длиной {len(raw)} символов от модели {model}")
                            
                            # Пытаемся извлечь JSON из ответа
                            parsed = None
                            try:
                                txt = raw.strip()
                                # Удаляем markdown-фенс, если модель вернула ```json ... ```
                                if txt.startswith('```'):
                                    lines = txt.splitlines()
                                    if lines and lines[0].startswith('```'):
                                        lines = lines[1:]
                                    if lines and lines[-1].startswith('```'):
                                        lines = lines[:-1]
                                    txt = "\n".join(lines).strip()
                                
                                parsed = json.loads(txt)
                            except Exception:
                                # Попробуем вытащить первый JSON-блок
                                import re as _re
                                m = _re.search(r'\{[\s\S]*\}', txt)
                                if m:
                                    parsed = json.loads(m.group(0))
                            
                            if isinstance(parsed, dict):
                                market = parsed.get('market')
                                pick = parsed.get('pick')
                                line = parsed.get('line')
                                odds = parsed.get('odds')
                                reason = parsed.get('reason', '')
                                
                                if market and pick and odds:
                                    # Преобразуем в кодировку
                                    bet_alt_code = self._encode_alternative_bet(market, pick, line)
                                    bet_alt_odds = float(odds) if isinstance(odds, (int, float)) else None
                                    # Получаем reason или пустую строку
                                    bet_alt_reason = str(reason).strip() if reason else ''
                                    
                                    # Вычисляем bet_alt_confirm по алгоритму:
                                    # Если bet_alt_odds <= bet_ai_odds и bet_alt_odds > 1.10, то bet_alt_confirm=1, иначе 0
                                    bet_ai_odds = match['bet_ai_odds'] if 'bet_ai_odds' in match.keys() and match['bet_ai_odds'] is not None else None
                                    if bet_alt_odds is not None:
                                        if bet_ai_odds is not None and bet_alt_odds <= bet_ai_odds and bet_alt_odds > 1.10:
                                            bet_alt_confirm = 1
                                        else:
                                            bet_alt_confirm = 0
                                    else:
                                        bet_alt_confirm = 0
                                    
                                    if bet_alt_code and bet_alt_odds is not None:
                                        print(f"[Football Alt Bet] Получена альтернативная ставка от модели {model}: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                                        return (bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason)
                                    else:
                                        print(f"[Football Alt Bet] Не удалось преобразовать в кодировку: market={market}, pick={pick}, line={line}")
                                        continue
                                else:
                                    print(f"[Football Alt Bet] Неполный ответ модели {model}: market={market}, pick={pick}, odds={odds}")
                                    continue
                            else:
                                print(f"[Football Alt Bet] Не удалось распарсить JSON от модели {model}, пробуем следующую")
                                continue
                        else:
                            print(f"[Football Alt Bet] Неверный формат ответа от модели {model}")
                    else:
                        print(f"[Football Alt Bet] HTTP ошибка {response.status_code} для модели {model}")
                        if response.status_code == 429:
                            print(f"[Football Alt Bet] Превышен лимит запросов для модели {model}, пробуем следующую")
                            continue
                        try:
                            error_data = response.json()
                            print(f"[Football Alt Bet] Ошибка API: {response.status_code} - {error_data}")
                        except:
                            print(f"[Football Alt Bet] Ошибка API: {response.status_code} - {response.text[:200]}")
                except requests.exceptions.Timeout:
                    print(f"[Football Alt Bet] Таймаут модели {model}")
                    continue
                except Exception as e:
                    print(f"[Football Alt Bet] Ошибка запроса к модели {model}: {e}")
                    continue
            
            print(f"[Football Alt Bet] Не удалось получить альтернативную ставку ни от одной модели для fixture {fixture_id}")
            return None
            
        except Exception as e:
            print(f"[Football Alt Bet ERROR] Ошибка получения альтернативной ставки: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    # ===== СТАРЫЙ АППРУВ (ЗАКОММЕНТИРОВАН, НЕ ИСПОЛЬЗУЕТСЯ) =====
    # def _perform_bet_approval_checks(self, cursor: sqlite3.Cursor, conn: sqlite3.Connection):
    #     """
    #     Выполняет второй проход риск-менеджмента (подтверждение ставок).
    #
    #     Анализирует все матчи, у которых есть рекомендация bet >= 1, но ещё не выставлен bet_approve.
    #
    #     Args:
    #         cursor: Активный курсор БД
    #         conn: Активное соединение с БД
    #     """
    #     cursor.execute(
    #         """
    #         SELECT * FROM matches
    #         WHERE bet IS NOT NULL
    #           AND bet >= 1
    #           AND bet_ai IS NOT NULL
    #           AND bet_ai_odds IS NOT NULL
    #           AND stats_60min IS NOT NULL
    #           AND bet_approve IS NULL
    #         ORDER BY match_date, match_time
    #         """
    #     )
    #
    #     matches_to_check = cursor.fetchall()
    #     if not matches_to_check:
    #         print("[Football Risk Approve] Нет матчей для подтверждения ставок")
    #         return
    #
    #     print(f"[Football Risk Approve] Запускаем подтверждение ставок для {len(matches_to_check)} матчей")
    #
    #     for match in matches_to_check:
    #         fixture_id = match['fixture_id']
    #         bet_ai = match['bet_ai']
    #         bet_ai_odds = match['bet_ai_odds']
    #         stats_json = match['stats_60min']
    #
    #         if not bet_ai or bet_ai_odds is None or not stats_json:
    #             print(f"[Football Risk Approve] Недостаточно данных для матча {fixture_id}, устанавливаем bet_approve = 0")
    #             cursor.execute(
    #                 """
    #                 UPDATE matches
    #                 SET bet_approve = 0,
    #                     bet_approve_reason = ?,
    #                     updated_at = CURRENT_TIMESTAMP
    #                 WHERE id = ?
    #                 """,
    #                 ("Недостаточно данных для подтверждения", match['id'])
    #             )
    #             conn.commit()
    #             continue
    #
    #         print(f"[Football Risk Approve] Анализируем матч {fixture_id} (прогноз {bet_ai}, кэф {bet_ai_odds})")
    #         analysis_response = self.analyze_bet_risk(fixture_id, bet_ai, float(bet_ai_odds), stats_json)
    #
    #         bet_approve_value = self._parse_bet_approve_decision(analysis_response)
    #         if bet_approve_value is None:
    #             print(f"[Football Risk Approve] Не удалось распознать резюме для матча {fixture_id}, устанавливаем 0")
    #             bet_approve_value = 0
    #
    #         cursor.execute(
    #             """
    #             UPDATE matches
    #             SET bet_approve = ?,
    #                 bet_approve_reason = ?,
    #                 updated_at = CURRENT_TIMESTAMP
    #             WHERE id = ?
    #             """,
    #             (bet_approve_value, analysis_response, match['id'])
    #         )
    #         conn.commit()
    #
    #         status_text = "ОДОБРЕНО" if bet_approve_value == 1 else "ОТКЛОНЕНО"
    #         print(f"[Football Risk Approve] Матч {fixture_id}: результат {status_text}")

    def _fetch_sofascore_statistics(self, sofascore_event_id: int) -> Optional[Dict]:
        """
        Получает статистику матча с SofaScore API.

        Args:
            sofascore_event_id: ID события в SofaScore

        Returns:
            Словарь со статистикой или None в случае ошибки
        """
        import random
        
        url = f"{SOFASCORE_API_URL}/event/{sofascore_event_id}/statistics"
        max_retries = 5
        attempt = 0
        
        while attempt < max_retries:
            try:
                # Выбираем случайный User-Agent
                headers = SOFASCORE_DEFAULT_HEADERS.copy()
                headers['User-Agent'] = random.choice(SOFASCORE_USER_AGENTS)
                
                # Случайная задержка перед запросом (1-3 секунды)
                if attempt > 0:
                    delay = random.uniform(2.0, 4.0) * (2 ** attempt)  # Экспоненциальный backoff
                    time.sleep(delay)
                
                response = requests.get(url, headers=headers, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    return data
                elif response.status_code == 403:
                    print(f"[Football SofaScore] 403 Forbidden при запросе статистики для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    if attempt < max_retries:
                        time.sleep(random.uniform(5.0, 10.0))
                    continue
                elif response.status_code >= 500:
                    print(f"[Football SofaScore] Ошибка сервера {response.status_code} при запросе статистики для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    continue
                else:
                    print(f"[Football SofaScore] Ошибка {response.status_code} при запросе статистики для event_id={sofascore_event_id}")
                    return None
                    
            except requests.exceptions.RequestException as e:
                print(f"[Football SofaScore] Сетевая ошибка при запросе статистики для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}: {e}")
                attempt += 1
                if attempt >= max_retries:
                    return None
                time.sleep(random.uniform(2.0, 4.0) * (2 ** attempt))
        
        print(f"[Football SofaScore] Не удалось получить статистику для event_id={sofascore_event_id} после {max_retries} попыток")
        return None

    def _fetch_sofascore_event_status(self, sofascore_event_id: int) -> Optional[str]:
        """
        Получает статус матча из SofaScore API.

        Args:
            sofascore_event_id: ID события в SofaScore

        Returns:
            Статус матча ('finished', 'live', 'notstarted', 'postponed' и т.д.) или None в случае ошибки
        """
        import random

        url = f"{SOFASCORE_API_URL}/event/{sofascore_event_id}"
        max_retries = 3
        attempt = 0

        while attempt < max_retries:
            try:
                # Выбираем случайный User-Agent
                headers = SOFASCORE_DEFAULT_HEADERS.copy()
                headers['User-Agent'] = random.choice(SOFASCORE_USER_AGENTS)

                # Случайная задержка перед запросом
                if attempt > 0:
                    delay = random.uniform(2.0, 4.0) * (2 ** attempt)
                    time.sleep(delay)

                response = requests.get(url, headers=headers, timeout=30)

                if response.status_code == 200:
                    data = response.json()
                    
                    # Извлекаем статус из различных возможных полей
                    # Обычно статус находится в event.status или event.statusText
                    event = data.get('event', {})
                    
                    # Варианты полей со статусом
                    status = event.get('status') or event.get('statusText') or event.get('statusDescription')
                    
                    if status:
                        # Нормализуем статус
                        status_lower = str(status).lower()
                        if 'finished' in status_lower or 'ft' in status_lower:
                            return 'finished'
                        elif 'live' in status_lower or 'inprogress' in status_lower:
                            return 'live'
                        elif 'notstarted' in status_lower or 'not started' in status_lower:
                            return 'notstarted'
                        elif 'postponed' in status_lower or 'cancelled' in status_lower:
                            return 'postponed'
                        else:
                            return status_lower
                    
                    # Если статус не найден, проверяем другие поля
                    # Иногда статус может быть в корне объекта
                    status = data.get('status') or data.get('statusText')
                    if status:
                        return str(status).lower()
                    
                    return None
                    
                elif response.status_code == 403:
                    print(f"[Football SofaScore] 403 Forbidden при запросе статуса для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    if attempt < max_retries:
                        time.sleep(random.uniform(5.0, 10.0))
                    continue
                elif response.status_code >= 500:
                    print(f"[Football SofaScore] Ошибка сервера {response.status_code} при запросе статуса для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    continue
                else:
                    print(f"[Football SofaScore] Ошибка {response.status_code} при запросе статуса для event_id={sofascore_event_id}")
                    return None

            except requests.exceptions.RequestException as e:
                print(f"[Football SofaScore] Сетевая ошибка при запросе статуса для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}: {e}")
                attempt += 1
                if attempt >= max_retries:
                    return None
                time.sleep(random.uniform(2.0, 4.0) * (2 ** attempt))

        print(f"[Football SofaScore] Не удалось получить статус для event_id={sofascore_event_id} после {max_retries} попыток")
        return None

    def _fetch_sofascore_event(self, sofascore_event_id: int) -> Optional[Dict]:
        """
        Получает полные данные о событии из SofaScore API.

        Args:
            sofascore_event_id: ID события в SofaScore

        Returns:
            Словарь с данными события или None в случае ошибки
        """
        import random

        url = f"{SOFASCORE_API_URL}/event/{sofascore_event_id}"
        max_retries = 3
        attempt = 0

        while attempt < max_retries:
            try:
                # Выбираем случайный User-Agent
                headers = SOFASCORE_DEFAULT_HEADERS.copy()
                headers['User-Agent'] = random.choice(SOFASCORE_USER_AGENTS)

                # Случайная задержка перед запросом
                if attempt > 0:
                    delay = random.uniform(2.0, 4.0) * (2 ** attempt)
                    time.sleep(delay)

                response = requests.get(url, headers=headers, timeout=30)

                if response.status_code == 200:
                    data = response.json()
                    return data

                elif response.status_code == 403:
                    print(f"[Football SofaScore] 403 Forbidden при запросе события для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    if attempt < max_retries:
                        time.sleep(random.uniform(5.0, 10.0))
                    continue
                elif response.status_code >= 500:
                    print(f"[Football SofaScore] Ошибка сервера {response.status_code} при запросе события для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}")
                    attempt += 1
                    continue
                else:
                    print(f"[Football SofaScore] Ошибка {response.status_code} при запросе события для event_id={sofascore_event_id}")
                    return None

            except requests.exceptions.RequestException as e:
                print(f"[Football SofaScore] Сетевая ошибка при запросе события для event_id={sofascore_event_id}, попытка {attempt + 1}/{max_retries}: {e}")
                attempt += 1
                if attempt >= max_retries:
                    return None
                time.sleep(random.uniform(2.0, 4.0) * (2 ** attempt))

        print(f"[Football SofaScore] Не удалось получить данные события для event_id={sofascore_event_id} после {max_retries} попыток")
        return None

    def _get_live_odds(self, fixture_id: str, sport_key: Optional[str] = None) -> Optional[float]:
        """
        Получает актуальные live коэффициенты фаворита на победу с The Odds API.
        
        Использует эндпойнт /v4/sports/{sport}/events/{eventId}/odds для получения коэффициентов конкретного события.
        Требует конкретный sport_key (например, 'soccer_uefa_champs_league'), который должен быть сохранен в БД.

        Args:
            fixture_id: ID матча в The Odds API (eventId)
            sport_key: Ключ вида спорта (например, 'soccer_epl'). Если не указан, будет получен из БД.

        Returns:
            Коэффициент фаворита на победу или None в случае ошибки/отсутствия live odds
        """
        try:
            # Если sport_key не передан, пытаемся получить из БД
            if not sport_key:
                conn = get_football_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT sport_key FROM matches WHERE fixture_id = ?", (fixture_id,))
                row = cursor.fetchone()
                conn.close()
                
                if row and row['sport_key']:
                    sport_key = row['sport_key']
                else:
                    # Если sport_key не найден в БД - это ошибка, не делаем запрос
                    print(f"[Football ERROR] sport_key не найден в БД для fixture {fixture_id}, пропускаем запрос live odds")
                    print(f"[Football] Запустите синхронизацию матчей (/api/football/sync) для обновления sport_key")
                    return None
            
            # Параметры запроса
            params = {
                "regions": "eu",
                "markets": "h2h",
                "oddsFormat": "decimal"
            }
            
            # Используем эндпойнт /sports/{sport}/events/{eventId}/odds
            # Требуется конкретный sport_key (например, 'soccer_uefa_champs_league')
            endpoint = f"/sports/{sport_key}/events/{fixture_id}/odds"
            data = self._make_api_request(endpoint, params)
            
            if not data or not isinstance(data, dict):
                print(f"[Football] Не удалось получить live odds для fixture {fixture_id} (ответ не является объектом)")
                return None
            
            # Проверяем, что ID матча совпадает
            if data.get('id') != fixture_id:
                print(f"[Football] Несоответствие ID: запрошен {fixture_id}, получен {data.get('id')}")
                return None
            
            # Извлекаем медианные коэффициенты для 1, X, 2 для сохранения в БД
            home_team = data.get('home_team')
            away_team = data.get('away_team')
            bookmakers = data.get('bookmakers', [])
            
            live_odds_1 = None
            live_odds_x = None
            live_odds_2 = None
            
            if home_team and away_team and bookmakers:
                # Собираем коэффициенты для каждой команды и ничьей
                home_odds = []
                away_odds = []
                draw_odds = []
                
                for bookmaker in bookmakers:
                    markets = bookmaker.get('markets', [])
                    for market in markets:
                        if market.get('key') != 'h2h':
                            continue
                        
                        outcomes = market.get('outcomes', [])
                        for outcome in outcomes:
                            name = outcome.get('name')
                            price = outcome.get('price')
                            
                            if not price or not name:
                                continue
                            
                            if name == home_team:
                                home_odds.append(float(price))
                            elif name == away_team:
                                away_odds.append(float(price))
                            elif name.lower() == 'draw':
                                draw_odds.append(float(price))
                
                # Вычисляем медианные коэффициенты
                def get_median(odds_list):
                    n = len(odds_list)
                    if n == 0:
                        return None
                    sorted_odds = sorted(odds_list)
                    if n % 2 == 0:
                        return (sorted_odds[n//2 - 1] + sorted_odds[n//2]) / 2.0
                    else:
                        return sorted_odds[n//2]
                
                live_odds_1 = get_median(home_odds)
                live_odds_x = get_median(draw_odds)
                live_odds_2 = get_median(away_odds)
            
            # Сохраняем коэффициенты в БД
            if live_odds_1 is not None or live_odds_x is not None or live_odds_2 is not None:
                conn = get_football_db_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE matches
                    SET live_odds_1 = ?, live_odds_x = ?, live_odds_2 = ?
                    WHERE fixture_id = ?
                """, (live_odds_1, live_odds_x, live_odds_2, fixture_id))
                conn.commit()
                conn.close()
            
            # Находим фаворита по медианному коэффициенту
            fav_info = self._determine_favorite(data)
            if fav_info:
                return fav_info['odds']
            
            print(f"[Football] Не удалось определить фаворита для fixture {fixture_id}")
            return None
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка получения live odds для fixture {fixture_id}: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def _get_ai_prediction_odds(self, fixture_id: str, bet_ai: str) -> Optional[float]:
        """
        Получает коэффициент на прогнозированный исход ИИ (1, 1X, X, X2, 2) из БД.
        
        Для одиночных исходов (1, X, 2) берет коэффициент из БД (live_odds_1, live_odds_x, live_odds_2).
        Для двойных шансов (1X, X2) вычисляет по формуле: 1 / (1/odd1 + 1/oddX)
        
        Коэффициенты должны быть сохранены в БД при запросе live_odds (_get_live_odds).
        
        Args:
            fixture_id: ID матча в The Odds API
            bet_ai: Прогноз ИИ ('1', '1X', 'X', 'X2', '2')

        Returns:
            Коэффициент на прогнозированный исход или None в случае ошибки
        """
        try:
            if not bet_ai:
                return None
            
            # Получаем коэффициенты из БД
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT live_odds_1, live_odds_x, live_odds_2
                FROM matches
                WHERE fixture_id = ?
            """, (fixture_id,))
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                print(f"[Football ERROR] Матч не найден в БД для fixture {fixture_id}")
                return None
            
            odd1 = row['live_odds_1']
            oddX = row['live_odds_x']
            odd2 = row['live_odds_2']
            
            if odd1 is None or oddX is None or odd2 is None:
                print(f"[Football] Коэффициенты для расчета bet_ai_odds не найдены в БД для fixture {fixture_id}")
                print(f"[Football] Возможно, live_odds еще не были запрошены. Коэффициенты: 1={odd1}, X={oddX}, 2={odd2}")
                return None
            
            # Возвращаем коэффициент в зависимости от прогноза ИИ
            bet_ai_upper = bet_ai.upper()
            
            if bet_ai_upper == '1':
                return float(odd1)
            elif bet_ai_upper == 'X':
                return float(oddX)
            elif bet_ai_upper == '2':
                return float(odd2)
            elif bet_ai_upper == '1X':
                # Двойной шанс: победа хозяев или ничья
                return 1.0 / (1.0/float(odd1) + 1.0/float(oddX))
            elif bet_ai_upper == 'X2':
                # Двойной шанс: ничья или победа гостей
                return 1.0 / (1.0/float(oddX) + 1.0/float(odd2))
            else:
                print(f"[Football] Неизвестный прогноз ИИ: {bet_ai}")
                return None
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка получения коэффициента для прогноза ИИ (bet_ai={bet_ai}): {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def _collect_60min_stats(self, match: sqlite3.Row):
        """
        Собирает статистику на 60-й минуте с SofaScore.

        Args:
            match: Запись матча из БД
        """
        try:
            fixture_id = match['fixture_id']
            sofascore_event_id = match['sofascore_event_id'] if 'sofascore_event_id' in match.keys() else None

            if not sofascore_event_id:
                print(f"[Football] Нет sofascore_event_id для матча {fixture_id}, пропускаем")
                return

            # Сначала получаем основное событие для актуального счета
            event_data = self._fetch_sofascore_event(sofascore_event_id)
            actual_score = None
            if event_data and 'event' in event_data:
                event = event_data['event']
                home_score_obj = event.get('homeScore', {})
                away_score_obj = event.get('awayScore', {})
                
                if isinstance(home_score_obj, dict) and isinstance(away_score_obj, dict):
                    # Приоритет: current (текущий счет) > normaltime > display
                    score_home = home_score_obj.get('current') or home_score_obj.get('normaltime') or home_score_obj.get('display')
                    score_away = away_score_obj.get('current') or away_score_obj.get('normaltime') or away_score_obj.get('display')
                    
                    if score_home is not None and score_away is not None:
                        try:
                            actual_score = {
                                'home': int(score_home),
                                'away': int(score_away)
                            }
                            print(f"[Football] Актуальный счет для fixture {fixture_id}: {actual_score['home']}-{actual_score['away']}")
                        except (ValueError, TypeError):
                            print(f"[Football] Ошибка преобразования счета в числа: home={score_home}, away={score_away}")

            # Задержка между запросами к SofaScore (2-5 секунд) для избежания бана
            delay_between_requests = random.uniform(2.0, 5.0)
            print(f"[Football] Задержка {delay_between_requests:.1f} сек перед запросом статистики для матча {fixture_id}")
            time.sleep(delay_between_requests)

            # Получаем статистику с SofaScore
            stats_data = self._fetch_sofascore_statistics(sofascore_event_id)

            if not stats_data:
                print(f"[Football] Не удалось получить статистику с SofaScore для event_id={sofascore_event_id}")
                return

            # Парсим статистику из SofaScore
            stats = self._parse_sofascore_statistics(stats_data, match)
            
            # Перезаписываем счет актуальным из основного события, если он был получен
            if actual_score:
                stats['score'] = actual_score
                print(f"[Football] Счет заменен на актуальный из основного события: {actual_score}")

            # ВСЕГДА запрашиваем live_odds, независимо от условий
            print(f"[Football] Запрашиваем live odds для матча {fixture_id}...")
            sport_key = match['sport_key'] if 'sport_key' in match.keys() else None
            live_odds_value = self._get_live_odds(fixture_id, sport_key)
            if live_odds_value:
                print(f"[Football] Получены live odds для {fixture_id}: {live_odds_value}")
            else:
                print(f"[Football] Не удалось получить live odds для {fixture_id}")

                        # Проверяем условия и записываем bet
            bet_value, _, ai_decision, ai_reason = self._calculate_bet(match, stats, fixture_id)

            # Сохраняем в БД (всегда сохраняем live_odds, даже если условия не выполнены)
            conn = get_football_db_connection()
            cursor = conn.cursor()

            stats_json = json.dumps(stats)
            # Обновляем счет в БД, если он был получен из SofaScore
            if actual_score:
                cursor.execute("""
                    UPDATE matches
                    SET stats_60min = ?,
                        bet = ?,
                        live_odds = ?,
                        final_score_home = ?,
                        final_score_away = ?,
                        bet_approve = NULL,
                        bet_approve_reason = NULL,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (stats_json, bet_value, live_odds_value, actual_score['home'], actual_score['away'], match['id']))
            else:
                cursor.execute("""
                    UPDATE matches
                    SET stats_60min = ?,
                        bet = ?,
                        live_odds = ?,
                        bet_approve = NULL,
                        bet_approve_reason = NULL,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (stats_json, bet_value, live_odds_value, match['id']))

            conn.commit()
            conn.close()

            print(f"[Football] Статистика на 60-й минуте сохранена для fixture {fixture_id}, bet: {bet_value}")
            
            # Получаем прогноз от ИИ
            print(f"[Football] Запрашиваем ИИ-прогноз для fixture {fixture_id}...")
            bet_ai, bet_ai_reason, bet_ai_model_name = self._get_ai_prediction(match, stats) 

            # Сохраняем результат ИИ в БД, даже если bet_ai не распознан, но есть полный ответ
            if bet_ai_reason:
                # Получаем коэффициент на прогнозированный исход из БД
                bet_ai_odds = None
                if bet_ai:
                    print(f"[Football] Получаем коэффициент для прогноза ИИ '{bet_ai}' для fixture {fixture_id}...")
                    bet_ai_odds = self._get_ai_prediction_odds(fixture_id, bet_ai)
                    if bet_ai_odds:
                        print(f"[Football] Получен коэффициент {bet_ai_odds} для прогноза ИИ '{bet_ai}'")
                    else:
                        print(f"[Football] Не удалось получить коэффициент для прогноза ИИ '{bet_ai}' (возможно, live_odds еще не были запрошены)")
                
                # Сохраняем результат ИИ в БД
                conn = get_football_db_connection()
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE matches
                    SET bet_ai = ?,
                        bet_ai_reason = ?,
                        bet_ai_full_response = ?,
                        bet_ai_model_name = ?,
                        bet_ai_odds = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (bet_ai, bet_ai_reason, bet_ai_reason, bet_ai_model_name, bet_ai_odds, match['id']))

                conn.commit()
                conn.close()

                if bet_ai:
                    print(f"[Football] ИИ-прогноз сохранен для fixture {fixture_id}: {bet_ai}, коэффициент: {bet_ai_odds}")
                else:
                    print(f"[Football] ИИ-прогноз не распознан, но ответ сохранен для fixture {fixture_id}")
                
                # Проверяем условие для отправки уведомления: bet_ai IS NOT NULL И bet_ai_odds > 1.50 И K60 > K1
                # Читаем данные из БД после сохранения bet_ai
                try:
                    conn = get_football_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT bet_ai, bet_ai_odds, live_odds, last_odds FROM matches WHERE id = ?", (match['id'],))
                    db_row = cursor.fetchone()
                    conn.close()
                    
                    # Проверяем условия: bet_ai_odds > 1.50 И K60 > K1
                    if (db_row and db_row['bet_ai'] and db_row['bet_ai_odds'] and db_row['bet_ai_odds'] > 1.50):
                        # Проверяем условие K60 > K1
                        live_odds = db_row['live_odds'] if db_row['live_odds'] is not None else None
                        last_odds = db_row['last_odds'] if db_row['last_odds'] is not None else None
                        
                        # K60 > K1 означает, что live_odds > last_odds (коэффициент вырос)
                        k60_greater_than_k1 = False
                        if live_odds is not None and last_odds is not None:
                            k60_greater_than_k1 = live_odds > last_odds
                        elif live_odds is not None and last_odds is None:
                            # Если K1 нет, но K60 есть - считаем что условие выполнено
                            k60_greater_than_k1 = True
                        
                        if k60_greater_than_k1:
                            # Читаем данные матча из БД для уведомления (исключаем bet_ai_full_response)
                            conn = get_football_db_connection()
                            cursor = conn.cursor()
                            cursor.execute("""
                                SELECT id, fixture_id, home_team, away_team, fav, bet_ai, bet_ai_odds, 
                                       bet_ai_reason, bet_alt_code, bet_alt_odds, bet_alt_confirm
                                FROM matches WHERE id = ?
                            """, (match['id'],))
                            match_for_notification = cursor.fetchone()
                            conn.close()
                            
                            if match_for_notification:
                                try:
                                    self._send_match_notification(match_for_notification, stats)
                                except Exception as notify_error:
                                    print(f"[Football ERROR] Ошибка отправки уведомления для фаворита: {notify_error}")
                        # Читаем данные матча из БД для уведомления (исключаем bet_ai_full_response)
                        conn = get_football_db_connection()
                        cursor = conn.cursor()
                        cursor.execute("""
                            SELECT id, fixture_id, home_team, away_team, fav, bet_ai, bet_ai_odds, 
                                   bet_ai_reason, bet_alt_code, bet_alt_odds, bet_alt_confirm
                            FROM matches WHERE id = ?
                        """, (match['id'],))
                        match_for_notification = cursor.fetchone()
                        conn.close()
                        
                        if match_for_notification:
                            try:
                                self._send_match_notification(match_for_notification, stats)
                            except Exception as notify_error:
                                print(f"[Football ERROR] Ошибка отправки уведомления для фаворита: {notify_error}")
                except Exception as fav_check_error:
                    print(f"[Football ERROR] Ошибка проверки условий для уведомления (фаворит): {fav_check_error}")
            
            # Получаем альтернативную ставку ОДИН РАЗ для каждого матча (если есть stats_60min и нет bet_alt_code)
            if stats:
                conn_alt = None
                try:
                    # Проверяем, есть ли уже bet_alt_code
                    conn_alt = get_football_db_connection()
                    cursor = conn_alt.cursor()
                    cursor.execute("SELECT bet_alt_code FROM matches WHERE id = ?", (match['id'],))
                    db_row = cursor.fetchone()
                    conn_alt.close()
                    conn_alt = None
                    
                    if db_row and not db_row['bet_alt_code']:
                        print(f"[Football] Запрашиваем альтернативную ставку для fixture {fixture_id} (есть stats_60min, нет bet_alt_code)")
                        # Получаем актуальные данные матча из БД для альтернативной ставки (исключаем bet_ai_full_response)
                        conn_alt = get_football_db_connection()
                        cursor = conn_alt.cursor()
                        cursor.execute("""
                            SELECT id, fixture_id, home_team, away_team, match_date, match_time, 
                                   live_odds_1, live_odds_x, live_odds_2, bet_ai_odds
                            FROM matches WHERE id = ?
                        """, (match['id'],))
                        match_updated = cursor.fetchone()
                        conn_alt.close()
                        conn_alt = None
                        
                        if match_updated:
                            alt_result = self._get_alternative_bet(match_updated, stats)
                            if alt_result:
                                bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason = alt_result
                                print(f"[Football] Получена альтернативная ставка: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                                # Сохраняем альтернативную ставку в БД (сохраняем reason в bet_ai_reason для не-фаворитов)
                                conn_alt = get_football_db_connection()
                                cursor = conn_alt.cursor()
                                cursor.execute("""
                                    UPDATE matches
                                    SET bet_alt_code = ?,
                                        bet_alt_odds = ?,
                                        bet_alt_confirm = ?,
                                        bet_ai_reason = ?,
                                        updated_at = CURRENT_TIMESTAMP
                                    WHERE id = ?
                                """, (bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason if bet_alt_reason else None, match['id']))
                                conn_alt.commit()
                                conn_alt.close()
                                conn_alt = None
                                print(f"[Football] Альтернативная ставка сохранена для fixture {fixture_id}: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                            else:
                                print(f"[Football] _get_alternative_bet вернул None для fixture {fixture_id}")
                except Exception as alt_error:
                    print(f"[Football Alt Bet ERROR] Ошибка получения альтернативной ставки для fixture {fixture_id}: {alt_error}")
                    import traceback
                    traceback.print_exc()
                finally:
                    if conn_alt:
                        try:
                            conn_alt.close()
                        except:
                            pass
            

        except Exception as e:
            print(f"[Football ERROR] Ошибка сбора статистики 60min: {e}")
            import traceback
            print(traceback.format_exc())
            
            # Даже при ошибке пытаемся отправить уведомление с информацией об ошибке
            try:
                # Пытаемся получить минимальные данные для уведомления
                # Если stats не была получена, создаем минимальную структуру
                try:
                    error_stats = stats
                except NameError:
                    # Если stats не определена, создаем минимальную структуру
                    error_stats = {'score': {'home': 0, 'away': 0}}
                
                error_reason = f"Ошибка обработки: {str(e)[:200]}"
                self._send_match_notification(match, error_stats, None, None, error_reason, None)
            except Exception as notify_error:
                print(f"[Football ERROR] Не удалось отправить уведомление об ошибке: {notify_error}")

    def _collect_60min_stats_without_fav(self, match: sqlite3.Row):
        """
        Собирает статистику на 60-й минуте для матчей без фаворита.
        Не запрашивает live_odds, только статистику и прогноз ИИ.
        
        Args:
            match: Запись матча из БД
        """
        try:
            fixture_id = match['fixture_id']
            sofascore_event_id = match['sofascore_event_id'] if 'sofascore_event_id' in match.keys() else None

            if not sofascore_event_id:
                print(f"[Football] Нет sofascore_event_id для матча {fixture_id}, пропускаем")
                return

            # Сначала получаем основное событие для актуального счета
            event_data = self._fetch_sofascore_event(sofascore_event_id)
            actual_score = None
            if event_data and 'event' in event_data:
                event = event_data['event']
                home_score_obj = event.get('homeScore', {})
                away_score_obj = event.get('awayScore', {})
                
                if isinstance(home_score_obj, dict) and isinstance(away_score_obj, dict):
                    # Приоритет: current (текущий счет) > normaltime > display
                    score_home = home_score_obj.get('current') or home_score_obj.get('normaltime') or home_score_obj.get('display')
                    score_away = away_score_obj.get('current') or away_score_obj.get('normaltime') or away_score_obj.get('display')
                    
                    if score_home is not None and score_away is not None:
                        try:
                            actual_score = {
                                'home': int(score_home),
                                'away': int(score_away)
                            }
                            print(f"[Football] Актуальный счет для fixture {fixture_id}: {actual_score['home']}-{actual_score['away']}")
                        except (ValueError, TypeError):
                            print(f"[Football] Ошибка преобразования счета в числа: home={score_home}, away={score_away}")

            # Задержка между запросами к SofaScore (2-5 секунд) для избежания бана
            delay_between_requests = random.uniform(2.0, 5.0)
            print(f"[Football] Задержка {delay_between_requests:.1f} сек перед запросом статистики для матча без фаворита {fixture_id}")
            time.sleep(delay_between_requests)

            # Получаем статистику с SofaScore
            stats_data = self._fetch_sofascore_statistics(sofascore_event_id)

            if not stats_data:
                print(f"[Football] Не удалось получить статистику с SofaScore для event_id={sofascore_event_id}")
                return

            # Парсим статистику из SofaScore
            stats = self._parse_sofascore_statistics(stats_data, match)
            
            # Перезаписываем счет актуальным из основного события, если он был получен
            if actual_score:
                stats['score'] = actual_score
                print(f"[Football] Счет заменен на актуальный из основного события: {actual_score}")

            # ===== ОТЛАДКА: Запрашиваем live odds для матчей без фаворита =====
            # TODO: Убрать этот блок после отладки или при достижении лимитов API
            # Цель: обновить live_odds_1, live_odds_x, live_odds_2 в таблице для отображения
            # ВАЖНО: Это расходует запросы к The Odds API. При достижении лимитов - закомментировать
            live_odds_value = None
            try:
                print(f"[Football DEBUG] Запрашиваем live odds для матча без фаворита {fixture_id}...")
                sport_key = match['sport_key'] if 'sport_key' in match.keys() else None
                live_odds_value = self._get_live_odds(fixture_id, sport_key)
                if live_odds_value:
                    print(f"[Football DEBUG] Получены live odds для матча без фаворита {fixture_id}: {live_odds_value}")
                else:
                    print(f"[Football DEBUG] Не удалось получить live odds для матча без фаворита {fixture_id}")
            except Exception as e:
                print(f"[Football DEBUG ERROR] Ошибка получения live odds для матча без фаворита {fixture_id}: {e}")
                # Не прерываем выполнение, продолжаем без live odds
            # ===== КОНЕЦ ОТЛАДКИ =====

            # Сохраняем статистику в БД (bet пока не устанавливаем, он будет установлен после получения рекомендации ИИ)
            conn = get_football_db_connection()
            cursor = conn.cursor()

            stats_json = json.dumps(stats)
            # Обновляем счет в БД, если он был получен из SofaScore
            if actual_score:
                cursor.execute("""
                    UPDATE matches
                    SET stats_60min = ?,
                        final_score_home = ?,
                        final_score_away = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (stats_json, actual_score['home'], actual_score['away'], match['id']))
            else:
                cursor.execute("""
                    UPDATE matches
                    SET stats_60min = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (stats_json, match['id']))

            conn.commit()
            conn.close()

            print(f"[Football] Статистика на 60-й минуте сохранена для матча без фаворита {fixture_id}")
            
            # Получаем прогноз от ИИ (без упоминания фаворита)
            print(f"[Football] Запрашиваем ИИ-прогноз для матча без фаворита {fixture_id}...")
            bet_ai, bet_ai_reason, bet_recommendation, bet_ai_model_name = self._get_ai_prediction_without_fav(match, stats) 

            # Устанавливаем bet на основе рекомендации
            bet_value = 1 if bet_recommendation else 0

            if bet_ai or bet_ai_reason:
                # Получаем коэффициент на прогнозированный исход из БД
                bet_ai_odds = None
                if bet_ai:
                    print(f"[Football] Получаем коэффициент для прогноза ИИ '{bet_ai}' для fixture {fixture_id}...")
                    bet_ai_odds = self._get_ai_prediction_odds(fixture_id, bet_ai)
                    if bet_ai_odds:
                        print(f"[Football] Получен коэффициент {bet_ai_odds} для прогноза ИИ '{bet_ai}'")
                    else:
                        print(f"[Football] Не удалось получить коэффициент для прогноза ИИ '{bet_ai}'")
                
                # Сохраняем результат ИИ в БД
                conn = get_football_db_connection()
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE matches
                    SET bet_ai = ?,
                        bet_ai_reason = ?,
                        bet_ai_full_response = ?,
                        bet_ai_model_name = ?,
                        bet_ai_odds = ?,
                        bet = ?,
                        bet_approve = NULL,
                        bet_approve_reason = NULL,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (bet_ai, bet_ai_reason, bet_ai_reason, bet_ai_model_name, bet_ai_odds, bet_value, match['id']))

                conn.commit()
                conn.close()

                recommendation_text = "СТАВИМ" if bet_recommendation else "ИГНОРИРУЕМ"
                if bet_ai:
                    print(f"[Football] ИИ-прогноз сохранен для матча без фаворита {fixture_id}: {bet_ai}, коэффициент: {bet_ai_odds}, рекомендация: {recommendation_text}, bet: {bet_value}")
                else:
                    print(f"[Football] ИИ-прогноз не распознан, но ответ сохранен для матча без фаворита {fixture_id}, bet: {bet_value}")
                
                # Получаем альтернативную ставку ОДИН РАЗ для каждого матча (если есть stats_60min и нет bet_alt_code)
                if stats:
                    conn_alt = None
                    try:
                        # Проверяем, есть ли уже bet_alt_code
                        conn_alt = get_football_db_connection()
                        cursor = conn_alt.cursor()
                        cursor.execute("SELECT bet_alt_code FROM matches WHERE id = ?", (match['id'],))
                        db_row = cursor.fetchone()
                        conn_alt.close()
                        conn_alt = None
                        
                        if db_row and not db_row['bet_alt_code']:
                            print(f"[Football] Запрашиваем альтернативную ставку для матча без фаворита {fixture_id} (есть stats_60min, нет bet_alt_code)")
                            # Получаем актуальные данные матча из БД для альтернативной ставки (исключаем bet_ai_full_response)
                            conn_alt = get_football_db_connection()
                            cursor = conn_alt.cursor()
                            cursor.execute("""
                                SELECT id, fixture_id, home_team, away_team, match_date, match_time, 
                                       live_odds_1, live_odds_x, live_odds_2, bet_ai_odds
                                FROM matches WHERE id = ?
                            """, (match['id'],))
                            match_updated = cursor.fetchone()
                            conn_alt.close()
                            conn_alt = None
                            
                            if match_updated:
                                alt_result = self._get_alternative_bet(match_updated, stats)
                                if alt_result:
                                    bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason = alt_result
                                    print(f"[Football] Получена альтернативная ставка: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                                    # Сохраняем альтернативную ставку в БД (сохраняем reason в bet_ai_reason для не-фаворитов)
                                    conn_alt = get_football_db_connection()
                                    cursor = conn_alt.cursor()
                                    cursor.execute("""
                                        UPDATE matches
                                        SET bet_alt_code = ?,
                                            bet_alt_odds = ?,
                                            bet_alt_confirm = ?,
                                            bet_ai_reason = ?,
                                            updated_at = CURRENT_TIMESTAMP
                                        WHERE id = ?
                                    """, (bet_alt_code, bet_alt_odds, bet_alt_confirm, bet_alt_reason if bet_alt_reason else None, match['id']))
                                    conn_alt.commit()
                                    conn_alt.close()
                                    conn_alt = None
                                    print(f"[Football] Альтернативная ставка сохранена для матча без фаворита {fixture_id}: {bet_alt_code} (коэф. {bet_alt_odds}, confirm={bet_alt_confirm})")
                                    
                                    # Проверяем условие для отправки уведомления: bet_alt_code IS NOT NULL И bet_alt_odds > 1.75 И bet_alt_confirm = 1
                                    if bet_alt_code and bet_alt_odds and bet_alt_odds > 1.75 and bet_alt_confirm == 1:
                                        # Читаем данные матча из БД для уведомления (исключаем bet_ai_full_response)
                                        conn_alt = get_football_db_connection()
                                        cursor = conn_alt.cursor()
                                        cursor.execute("""
                                            SELECT id, fixture_id, home_team, away_team, fav, bet_ai, bet_ai_odds, 
                                                   bet_ai_reason, bet_alt_code, bet_alt_odds, bet_alt_confirm
                                            FROM matches WHERE id = ?
                                        """, (match['id'],))
                                        match_for_notification = cursor.fetchone()
                                        conn_alt.close()
                                        conn_alt = None
                                        
                                        if match_for_notification:
                                            try:
                                                self._send_match_notification(match_for_notification, stats)
                                            except Exception as notify_error:
                                                print(f"[Football ERROR] Ошибка отправки уведомления для матча без фаворита: {notify_error}")
                                else:
                                    print(f"[Football] _get_alternative_bet вернул None для матча без фаворита {fixture_id}")
                    except Exception as alt_error:
                        print(f"[Football Alt Bet ERROR] Ошибка получения альтернативной ставки для матча без фаворита {fixture_id}: {alt_error}")
                        import traceback
                        traceback.print_exc()
                    finally:
                        if conn_alt:
                            try:
                                conn_alt.close()
                            except:
                                pass
                
            else:
                # Если прогноз не получен, все равно обновляем bet = 0
                conn_else = None
                try:
                    conn_else = get_football_db_connection()
                    cursor = conn_else.cursor()
                    cursor.execute("""
                        UPDATE matches
                        SET bet = ?,
                            bet_approve = NULL,
                            bet_approve_reason = NULL,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    """, (bet_value, match['id']))
                    conn_else.commit()
                    print(f"[Football] ИИ-прогноз не получен для матча без фаворита {fixture_id}, установлен bet: {bet_value}")
                except Exception as e:
                    print(f"[Football ERROR] Ошибка обновления bet для матча без фаворита {fixture_id}: {e}")
                finally:
                    if conn_else:
                        try:
                            conn_else.close()
                        except:
                            pass

        except Exception as e:
            print(f"[Football ERROR] Ошибка сбора статистики 60min для матча без фаворита: {e}")
            import traceback
            print(traceback.format_exc())

    def _get_ai_prediction_without_fav(self, match: sqlite3.Row, stats: Dict) -> Tuple[Optional[str], Optional[str], Optional[bool], Optional[str]]:
        """
        Получает прогноз от ИИ для матчей без фаворита (без упоминания фаворита в промпте).
        
        Args:
            match: Запись матча из БД
            stats: Статистика на 60-й минуте (из stats_60min)
        
        Returns:
            Кортеж (bet_ai, bet_ai_reason, bet_recommendation, model_name):
            - bet_ai: Прогноз ('1', '1X', 'X', 'X2', '2') или None
            - bet_ai_reason: Полный ответ от ИИ или None
            - bet_recommendation: True если СТАВИМ, False если ИГНОРИРУЕМ, None если не распознано
            - model_name: Имя модели, давшей ответ, или None
        """
        if not self.openrouter_api_key:
            print("[Football] OpenRouter API ключ не установлен, пропускаем ИИ-прогноз")
            return None, None, None, None
        
        try:
            # Формируем промпт без упоминания фаворита
            home_team = match['home_team']
            away_team = match['away_team']
            
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)
            
            # Получаем коэффициенты из БД (они должны быть уже сохранены при запросе live_odds)
            fixture_id = match['fixture_id']
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT live_odds_1, live_odds_x, live_odds_2
                FROM matches
                WHERE fixture_id = ?
            """, (fixture_id,))
            row = cursor.fetchone()
            conn.close()
            
            live_odds_1 = row['live_odds_1'] if row else None
            live_odds_x = row['live_odds_x'] if row else None
            live_odds_2 = row['live_odds_2'] if row else None
            
            # Форматируем статистику как JSON для передачи ИИ
            stats_json = json.dumps(stats, ensure_ascii=False, indent=2)
            
            # Формируем строку с коэффициентами
            odds_info = ""
            if live_odds_1 is not None or live_odds_x is not None or live_odds_2 is not None:
                odds_info = f"""
- Текущие коэффициенты на исходы:
  * Победа {home_team}: {live_odds_1 if live_odds_1 is not None else 'N/A'}
  * Ничья: {live_odds_x if live_odds_x is not None else 'N/A'}
  * Победа {away_team}: {live_odds_2 if live_odds_2 is not None else 'N/A'}
"""
            
            prompt = f"""Ты - футбольный аналитик. Сейчас перерыв после первого тайма. Изучи предоставленную статистику матча после первого тайма, хорошо подумай и сделай прогноз на итоговый результат матча в основное время.

ВАЖНО: Все твои ответы должны быть полностью на русском языке.

Информация о матче:
- Команды: {home_team} vs {away_team}
- Текущий счет после первого тайма: {home_score} - {away_score}
Детальная статистика первого тайма:
{stats_json}

Твой ответ должен состоять в виде строки в формате: "Результат (1, 1X, X, X2, 2) Рекомендация (ИГНОРИРУЕМ или СТАВИМ)".

1. Результат СТРОГО в виде одного из вариантов: 1 или 1X или X или X2 или 2
Где:
- 1 = победа домашней команды ({home_team})
- 1X = ничья или победа домашней команды ({home_team})
- X = ничья
- X2 = ничья или победа гостевой команды ({away_team})
- 2 = победа гостевой команды ({away_team})

2. Рекомендация, стоит ли ставить на этот исход (СТАВИМ или ИГНОРИРУЕМ) при текущих коэффициентах букмекеров.
{odds_info}
Отвечай СТАВИМ только если прогноз имеет хорошее соотношение цены и вероятности на основе коэффициентов и статистики.

Примеры ответа:
1X СТАВИМ
1 СТАВИМ
X ИГНОРИРУЕМ
X2 СТАВИМ
2 ИГНОРИРУЕМ
1X ИГНОРИРУЕМ
1 ИГНОРИРУЕМ
X СТАВИМ
X2 ИГНОРИРУЕМ
2 СТАВИМ
"""
            
            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": os.getenv("OPENROUTER_SITE_URL", "http://localhost:5000")
            }
            
            # Список моделей для попыток (основная + три fallback)
            models_to_try = [self.ai_primary_model, self.ai_fallback_model1, self.ai_fallback_model2, self.ai_fallback_model3]
            
            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue
                    
                print(f"[Football AI] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model}")
                
                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "max_tokens": 2000,
                        "temperature": 0.3  # Низкая температура для более детерминированного ответа
                    }
                    
                    print(f"[Football AI] Отправка запроса к OpenRouter API (модель: {model})")
                    
                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=60
                    )
                    
                    if response.status_code == 200:
                        try:
                            data = response.json()
                            if 'choices' in data and len(data['choices']) > 0:
                                ai_response = data['choices'][0]['message']['content']
                                print(f"[Football AI] Получен ответ длиной {len(ai_response)} символов от модели {model}")
                                
                                # Парсим ответ - ищем один из вариантов: 1, 1X, X, X2, 2
                                bet_ai = self._parse_ai_prediction(ai_response)
                                # Парсим рекомендацию - ищем СТАВИМ/ИГНОРИРУЕМ
                                bet_recommendation = self._parse_ai_recommendation(ai_response)
                                
                                if bet_ai:
                                    recommendation_text = "СТАВИМ" if bet_recommendation else "ИГНОРИРУЕМ"
                                    print(f"[Football AI] Успешно распознан прогноз: {bet_ai}, рекомендация: {recommendation_text}")
                                    return bet_ai, ai_response, bet_recommendation, model
                                else:
                                    print(f"[Football AI] Не удалось распознать прогноз в ответе, пробуем следующую модель")
                                    if model_idx < len(models_to_try) - 1:
                                        continue
                                    else:
                                        # Последняя модель - возвращаем ответ даже если не распознан
                                        print(f"[Football AI] Все модели испробованы, возвращаем ответ без распознанного прогноза")
                                        return None, ai_response, None, model
                        except json.JSONDecodeError as e:
                            print(f"[Football AI ERROR] Ошибка парсинга JSON ответа: {e}")
                            continue
                    else:
                        print(f"[Football AI ERROR] Ошибка API: {response.status_code} - {response.text}")
                        continue
                        
                except requests.exceptions.RequestException as e:
                    print(f"[Football AI ERROR] Ошибка запроса к OpenRouter: {e}")
                    continue
            
            print(f"[Football AI] Не удалось получить прогноз от всех моделей")
            return None, None, None, None
            
        except Exception as e:
            print(f"[Football AI ERROR] Ошибка получения ИИ-прогноза: {e}")
            import traceback
            print(traceback.format_exc())
            return None, None, None, None

    def analyze_bet_risk(self, fixture_id: str, bet_ai: str, bet_ai_odds: float, stats_json: str) -> Optional[str]:
        """
        Анализирует риск ставки на основе прогноза ИИ, коэффициента и статистики.
        
        Args:
            fixture_id: ID матча
            bet_ai: Прогноз ИИ ('1', '1X', 'X', 'X2', '2')
            bet_ai_odds: Коэффициент на прогнозированный исход
            stats_json: JSON строка со статистикой матча (stats_60min)
        
        Returns:
            Ответ от ИИ с анализом риска или None в случае ошибки
        """
        if not self.openrouter_api_key:
            print("[Football] OpenRouter API ключ не установлен, пропускаем анализ риска")
            return None
        
        try:
            # Парсим статистику
            stats = json.loads(stats_json) if isinstance(stats_json, str) else stats_json
            
            # Получаем информацию о матче из БД (исключаем bet_ai_full_response)
            conn = get_football_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, fixture_id, home_team, away_team, match_date, match_time, status
                FROM matches WHERE fixture_id = ?
            """, (fixture_id,))
            match_row = cursor.fetchone()
            conn.close()
            
            if not match_row:
                print(f"[Football] Матч {fixture_id} не найден в БД")
                return None
            
            match = dict(match_row)
            home_team = match.get('home_team', '')
            away_team = match.get('away_team', '')
            
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)
            
            # Форматируем статистику для промпта
            stats_formatted = json.dumps(stats, ensure_ascii=False, indent=2)
            
            # Определяем название исхода
            outcome_names = {
                '1': f'победа домашней команды ({home_team})',
                '1X': f'ничья или победа домашней команды ({home_team})',
                'X': 'ничья',
                'X2': f'ничья или победа гостевой команды ({away_team})',
                '2': f'победа гостевой команды ({away_team})'
            }
            outcome_name = outcome_names.get(bet_ai.upper(), bet_ai)
            
            prompt = f"""Ты - эксперт по анализу рисков ставок на футбол. Твоя задача - проанализировать предложенную ставку и дать рекомендацию: стоит ли рисковать или нет.

Информация о матче:
- Команды: {home_team} vs {away_team}
- Текущий счет после первого тайма: {home_score} - {away_score}

Прогноз ИИ:
- Исход: {outcome_name} ({bet_ai})
- Коэффициент: {bet_ai_odds}

Детальная статистика первого тайма:
{stats_formatted}

Проанализируй статистику, текущий счет, прогноз ИИ и коэффициент. Дай обоснованную рекомендацию: СТОИТ ЛИ РИСКНУТЬ или НЕ СТОИТ РИСКОВАТЬ, и подробно объясни свое решение.
Предложи альтернативную ставку на гандикап или больше-меньше.

В конце анализа добавь отдельную строку СТРОГО в формате "Резюме: ОДОБРИТЬ" если считаешь, что стоит рискнуть, или "Резюме: ОТКЛОНИТЬ" если рисковать не стоит."""
            
            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": os.getenv("OPENROUTER_SITE_URL", "http://localhost:5000")
            }
            
            # Список моделей для попыток (основная + три fallback)
            models_to_try = [self.risk_analysis_primary, self.risk_analysis_fallback1, self.risk_analysis_fallback2, self.risk_analysis_fallback3]
            
            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue
                    
                print(f"[Football Risk Analysis] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model}")
                
                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "max_tokens": 2000,
                        "temperature": 0.7  # Средняя температура для более развернутого ответа
                    }
                    
                    print(f"[Football Risk Analysis] Отправка запроса к OpenRouter API (модель: {model})")
                    print(f"[Football Risk Analysis] URL: {self.openrouter_api_url}/chat/completions")
                    print(f"[Football Risk Analysis] Payload: model={model}, max_tokens={payload['max_tokens']}, temperature={payload['temperature']}")
                    
                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=60
                    )
                    
                    print(f"[Football Risk Analysis] Получен ответ от OpenRouter API (модель: {model}): статус {response.status_code}")
                    
                    if response.status_code == 200:
                        try:
                            data = response.json()
                            if 'choices' in data and len(data['choices']) > 0:
                                ai_response = data['choices'][0]['message']['content']
                                print(f"[Football Risk Analysis] Получен ответ длиной {len(ai_response)} символов от модели {model}")
                                return ai_response
                            else:
                                print(f"[Football Risk Analysis] Неожиданный формат ответа от модели {model}")
                        except Exception as e:
                            print(f"[Football Risk Analysis] Ошибка парсинга ответа от модели {model}: {e}")
                    else:
                        error_status = response.status_code
                        print(f"[Football Risk Analysis] Ошибка API для модели {model}: статус {error_status}")
                        if error_status == 429:
                            print(f"[Football Risk Analysis] Превышен лимит запросов для модели {model}, пробуем следующую")
                            continue
                        elif error_status == 401:
                            print(f"[Football Risk Analysis] Ошибка авторизации для модели {model}")
                            break
                        else:
                            # Для других ошибок тоже пробуем следующую модель
                            continue
                
                except requests.exceptions.Timeout:
                    print(f"[Football Risk Analysis] Таймаут при запросе к модели {model}")
                    continue
                except Exception as e:
                    print(f"[Football Risk Analysis] Ошибка при запросе к модели {model}: {e}")
                    continue
            
            print(f"[Football Risk Analysis] Не удалось получить ответ ни от одной модели")
            return None
            
        except Exception as e:
            print(f"[Football Risk Analysis ERROR] Ошибка анализа риска: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def _collect_final_result(self, match: sqlite3.Row):
        """
        Собирает финальный результат матча из SofaScore API.

        Args:
            match: Запись матча из БД
        """
        try:
            fixture_id = match['fixture_id']
            sofascore_event_id = match['sofascore_event_id'] if 'sofascore_event_id' in match.keys() and match['sofascore_event_id'] else None

            if not sofascore_event_id:
                print(f"[Football] У матча {fixture_id} нет sofascore_event_id, пропускаем сбор финального результата")
                return

            print(f"[Football] Получаем финальный результат из SofaScore для event_id {sofascore_event_id}")

            # Получаем полные данные о событии
            event_data = self._fetch_sofascore_event(sofascore_event_id)

            if not event_data:
                print(f"[Football] Не удалось получить данные из SofaScore для event_id {sofascore_event_id}")
                return

            # Извлекаем счет из данных SofaScore
            # Структура данных из /api/v1/event/{event_id}:
            # event.homeScore.current - счет домашней команды
            # event.awayScore.current - счет гостевой команды
            # Также доступны: display, normaltime, period1, period2
            score_home = None
            score_away = None

            event = event_data.get('event', {})
            
            # Основной способ: event.homeScore.current и event.awayScore.current
            home_score_obj = event.get('homeScore', {})
            away_score_obj = event.get('awayScore', {})
            
            if isinstance(home_score_obj, dict):
                # Приоритет: normaltime (обычное время) > current > display
                score_home = home_score_obj.get('normaltime') or home_score_obj.get('current') or home_score_obj.get('display')
            
            if isinstance(away_score_obj, dict):
                # Приоритет: normaltime (обычное время) > current > display
                score_away = away_score_obj.get('normaltime') or away_score_obj.get('current') or away_score_obj.get('display')

            if score_home is None or score_away is None:
                print(f"[Football] Не удалось извлечь счет из данных SofaScore для event_id {sofascore_event_id}")
                print(f"[Football] Доступные поля в event: {list(event.keys()) if event else 'N/A'}")
                print(f"[Football] Доступные поля в корне: {list(event_data.keys())}")
                # Попробуем вывести всю структуру для отладки
                import json
                print(f"[Football] Полная структура данных (первые 2000 символов): {json.dumps(event_data, indent=2, ensure_ascii=False)[:2000]}")
                return

            # Преобразуем счет в целые числа
            try:
                score_home = int(score_home) if score_home is not None else None
                score_away = int(score_away) if score_away is not None else None
            except (ValueError, TypeError):
                print(f"[Football] Ошибка преобразования счета в числа: home={score_home}, away={score_away}")
                return

            # Определяем, выиграл ли фаворит
            # fav_team_id: 1 = home, 0 = away
            fav_team_id = match['fav_team_id']
            fav_won = None

            if score_home > score_away:
                # Домашняя команда выиграла
                fav_won = 1 if fav_team_id == 1 else 0
            elif score_away > score_home:
                # Гостевая команда выиграла
                fav_won = 1 if fav_team_id == 0 else 0
            else:
                # Ничья
                fav_won = 0

            # Сохраняем
            conn = get_football_db_connection()
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE matches
                SET final_score_home = ?, final_score_away = ?,
                    fav_won = ?, status = 'finished', updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (score_home, score_away, fav_won, match['id']))

            conn.commit()
            conn.close()

            print(f"[Football] Финальный результат сохранен для fixture {fixture_id}: {score_home}-{score_away}, фаворит выиграл: {fav_won == 1}")

        except Exception as e:
            print(f"[Football ERROR] Ошибка сбора финального результата: {e}")
            import traceback
            print(traceback.format_exc())

    def _parse_statistics(self, stats_data: Dict) -> Dict:
        """
        Парсит статистику из API-Football.
        
        Args:
            stats_data: Сырые данные статистики
            
        Returns:
            Словарь с отпарсенной статистикой
        """
        stats = {}

        try:
            # API-Football возвращает статистику для каждой команды       
            for team_stats in stats_data.get('statistics', []):
                team = team_stats.get('team', {}).get('name', '')

                # Парсим метрики
                for stat in team_stats.get('statistics', []):
                    stat_type = stat.get('type', '')
                    stat_value = stat.get('value')

                    if stat_type == 'Ball Possession':
                        stats[team.lower()] = {'possession': stat_value}
                    elif stat_type == 'Shots on Goal':
                        if team.lower() not in stats:
                            stats[team.lower()] = {}
                        stats[team.lower()]['shots_on_target'] = stat_value
                    elif stat_type == 'expected_goals':
                        if team.lower() not in stats:
                            stats[team.lower()] = {}
                        stats[team.lower()]['xG'] = stat_value

        except Exception as e:
            print(f"[Football ERROR] Ошибка парсинга статистики: {e}")

        return stats

    def _parse_sofascore_statistics(self, stats_data: Dict, match: sqlite3.Row) -> Dict:
        """
        Парсит статистику из SofaScore API.

        Args:
            stats_data: Сырые данные статистики от SofaScore
            match: Запись матча из БД

        Returns:
            Словарь с отпарсенной статистикой: {'score': {...}, 'possession': {...}, 'shots_on_target': {...}, 'xG': {...}}
        """
        stats = {}
        
        try:
            # Сохраняем весь ответ API в raw_data для полноты информации
            stats['raw_data'] = stats_data

            # Получаем текущий счет (для удобства выносим отдельно)
            home_score = stats_data.get('homeScore', {}).get('current', 0)
            away_score = stats_data.get('awayScore', {}).get('current', 0)
            stats['score'] = {
                'home': home_score,
                'away': away_score
            }

            # Сохраняем все остальные поля из API
            # Сохраняем периоды со всей статистикой
            if 'periods' in stats_data:
                stats['periods'] = stats_data['periods']
            
            # Сохраняем статистику напрямую (если есть)
            if 'statistics' in stats_data:
                stats['statistics'] = stats_data['statistics']
            
            # Сохраняем все остальные поля из API
            for key in stats_data:
                if key not in ['homeScore', 'awayScore', 'periods', 'statistics']:
                    stats[key] = stats_data[key]

            # Получаем статистику по группам (periods или statistics) для парсинга
            periods = stats_data.get('periods', [])
            statistics = stats_data.get('statistics', [])
            
            # Парсим часто используемые поля для удобства доступа
            # Извлекаем часто используемые поля из periods
            if periods:
                for period in periods:
                    if period.get('period') == 'all' or period.get('period') == 'REGULAR':
                        groups = period.get('groups', [])
                        parsed_stats = {}
                        for group in groups:
                            group_name = group.get('groupName', '')
                            stat_items = group.get('statisticsItems', [])
                            for item in stat_items:
                                item_name = item.get('name', '')
                                # Сохраняем все статистики из группы
                                if group_name not in parsed_stats:
                                    parsed_stats[group_name] = []
                                parsed_stats[group_name].append({
                                    'name': item_name,
                                    'home': item.get('home'),
                                    'away': item.get('away'),
                                    'total': item.get('total')
                                })
                        if parsed_stats:
                            stats['parsed_period_all'] = parsed_stats

            # Извлекаем часто используемые поля из statistics
            if statistics:
                parsed_stats = {}
                for stat_group in statistics:
                    if isinstance(stat_group, dict):
                        group_name = stat_group.get('groupName', '')
                        stat_items = stat_group.get('statisticsItems', [])
                        parsed_items = []
                        for item in stat_items:
                            parsed_items.append({
                                'name': item.get('name', ''),
                                'home': item.get('home'),
                                'away': item.get('away'),
                                'total': item.get('total')
                            })
                        if parsed_items:
                            parsed_stats[group_name] = parsed_items
                if parsed_stats:
                    stats['parsed_statistics'] = parsed_stats

            print(f"[Football] Распарсена полная статистика SofaScore: score={stats.get('score')}, сохранено {len(stats)} полей")
            
        except Exception as e:
            print(f"[Football ERROR] Ошибка парсинга статистики SofaScore: {e}")
            import traceback
            print(traceback.format_exc())
            # В случае ошибки всё равно сохраняем сырые данные
            stats = {'raw_data': stats_data}
            if 'homeScore' in stats_data and 'awayScore' in stats_data:
                stats['score'] = {
                    'home': stats_data.get('homeScore', {}).get('current', 0),
                    'away': stats_data.get('awayScore', {}).get('current', 0)
                }
        
        return stats

    def _extract_stat_value(self, stats: Dict, stat_group_name: str, stat_item_name: str) -> Dict[str, float]:
        """
        Извлекает значение статистики из новой структуры данных.
        
        Args:
            stats: Словарь со статистикой
            stat_group_name: Название группы статистики (например, 'Ball possession', 'Shots on target')
            stat_item_name: Название конкретной статистики (например, 'Ball possession', 'Shots on target')
        
        Returns:
            Словарь {'home': value, 'away': value} или пустой словарь если не найдено
        """
        result = {'home': 0, 'away': 0}
        
        # Пытаемся найти в parsed_period_all
        if 'parsed_period_all' in stats:
            parsed = stats['parsed_period_all']
            if stat_group_name in parsed:
                for item in parsed[stat_group_name]:
                    if item.get('name') == stat_item_name:
                        result['home'] = item.get('home', 0) or 0
                        result['away'] = item.get('away', 0) or 0
                        return result
        
        # Пытаемся найти в parsed_statistics
        if 'parsed_statistics' in stats:
            parsed = stats['parsed_statistics']
            if stat_group_name in parsed:
                for item in parsed[stat_group_name]:
                    if item.get('name') == stat_item_name:
                        result['home'] = item.get('home', 0) or 0
                        result['away'] = item.get('away', 0) or 0
                        return result
        
        # Пытаемся найти в raw_data через periods
        if 'raw_data' in stats:
            raw_data = stats['raw_data']
            periods = raw_data.get('periods', [])
            for period in periods:
                if period.get('period') == 'all' or period.get('period') == 'REGULAR':
                    groups = period.get('groups', [])
                    for group in groups:
                        if group.get('groupName') == stat_group_name:
                            stat_items = group.get('statisticsItems', [])
                            for item in stat_items:
                                if item.get('name') == stat_item_name:
                                    result['home'] = item.get('home', 0) or 0
                                    result['away'] = item.get('away', 0) or 0
                                    return result
        
        # Пытаемся найти в raw_data через statistics
        if 'raw_data' in stats:
            raw_data = stats['raw_data']
            statistics = raw_data.get('statistics', [])
            for stat_group in statistics:
                if isinstance(stat_group, dict) and stat_group.get('groupName') == stat_group_name:
                    stat_items = stat_group.get('statisticsItems', [])
                    for item in stat_items:
                        if item.get('name') == stat_item_name:
                            result['home'] = item.get('home', 0) or 0
                            result['away'] = item.get('away', 0) or 0
                            return result
        
        return result

    def _calculate_bet(self, match: sqlite3.Row, stats: Dict, fixture_id: str) -> Tuple[Optional[float], Optional[float]]:
        """
        Рассчитывает значение bet на основе решения ИИ.

        Вместо эвристик (владение, xG и т.д.) используется ИИ, который анализирует
        всю доступную статистику и решает ДА/НЕТ.

        Args:
            match: Запись матча
            stats: Статистика на 60-й минуте (от SofaScore, с raw_data)
            fixture_id: ID матча в The Odds API

                Returns:
            Кортеж (bet_value, live_odds, ai_decision, ai_reason):
            - bet_value: Коэффициент live odds если ИИ ответил ДА, 0 если НЕТ, 1 если лимит API исчерпан
            - live_odds: Реальное значение live odds из API (может быть None если не удалось получить)
            - ai_decision: Решение ИИ (True = ДА, False = НЕТ, None = ошибка)
            - ai_reason: Полный ответ от ИИ или None
        """
        try:
            fav_team = match['fav']
            
            # Проверяем last_odds (K1) - если коэффициент > 1.50, не отправляем запрос модели
            last_odds = match['last_odds'] if 'last_odds' in match.keys() and match['last_odds'] is not None else None
            if last_odds is None or last_odds > 1.50:
                print(f"[Football] Коэффициент фаворита {fav_team} (last_odds={last_odds}) > 1.50, пропускаем запрос модели для {fixture_id}")
                return (0, None, None, None)
            
            # Проверяем текущий счет - если фаворит выигрывает, не делаем ставку
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)
            
            # Определяем, кто фаворит (home или away)
            home_team = match['home_team']
            away_team = match['away_team']
            fav_is_home = (fav_team == home_team)
            
            # Если фаворит выигрывает, не делаем ставку
            if fav_is_home and home_score > away_score:
                print(f"[Football] Фаворит {fav_team} выигрывает ({home_score}-{away_score}), пропускаем ставку для {fixture_id}")
                return (0, None, None, None)
            elif not fav_is_home and away_score > home_score:
                print(f"[Football] Фаворит {fav_team} выигрывает ({home_score}-{away_score}), пропускаем ставку для {fixture_id}")
                return (0, None, None, None)

            # Получаем решение от ИИ
            print(f"[Football] Запрашиваем решение ИИ для матча {fixture_id}...")
            is_yes, ai_reason = self._get_bet_ai_decision(match, stats)

            if is_yes is None:
                # Не удалось получить ответ от ИИ - не делаем ставку
                print(f"[Football] Не удалось получить решение ИИ для матча {fixture_id}, устанавливаем bet=0")
                return (0, None, None, ai_reason)

            if not is_yes:
                # ИИ ответил НЕТ - не делаем ставку
                print(f"[Football] ИИ ответил НЕТ для матча {fixture_id}: {ai_reason[:200] if ai_reason else 'N/A'}...")
                return (0, None, False, ai_reason)

            # ИИ ответил ДА - запрашиваем live odds
            print(f"[Football] ИИ ответил ДА для матча {fixture_id}. Запрашиваем live odds...")
            sport_key = match['sport_key'] if 'sport_key' in match.keys() else None
            live_odds = self._get_live_odds(fixture_id, sport_key)

            if live_odds is None:
                # Если не удалось получить live odds (лимит исчерпан или матч не найден), сохраняем 1 в bet
                print(f"[Football] Не удалось получить live odds для {fixture_id}, сохраняем bet=1, live_odds=NULL")
                return (1, None, True, ai_reason)

            print(f"[Football] Получены live odds для фаворита {fav_team}: {live_odds}")
            return (live_odds, live_odds, True, ai_reason)

        except Exception as e:
            print(f"[Football ERROR] Ошибка расчета bet: {e}")
            import traceback
            print(traceback.format_exc())
            return (0, None, None, None)

    def _send_match_notification(self, match: sqlite3.Row, stats: Dict) -> bool:
        """
        Отправляет уведомление в Telegram подписчикам о матче.

        Args:
            match: Запись матча из БД (должна содержать все необходимые поля)
            stats: Статистика на 60-й минуте (для получения счета)

        Returns:
            bool: True если уведомление отправлено успешно
        """
        if not TELEGRAM_AVAILABLE:
            return False

        try:
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)

            home_team = match['home_team']
            away_team = match['away_team']
            fav_team = match['fav']
            is_match_without_fav = (fav_team == 'NONE' or not fav_team)

            # Получаем обоснование из БД
            bet_ai_reason = match['bet_ai_reason'] if 'bet_ai_reason' in match.keys() else None
            if bet_ai_reason:
                ai_reason_full = str(bet_ai_reason).strip()
                if not ai_reason_full:
                    ai_reason_full = "Нет данных"
                else:
                    # Убираем markdown из текста
                    import re
                    # Удаляем markdown жирный текст **text** -> text
                    ai_reason_full = re.sub(r'\*\*(.+?)\*\*', r'\1', ai_reason_full)
                    # Удаляем markdown курсив *text* -> text
                    ai_reason_full = re.sub(r'(?<!\*)\*([^*]+?)\*(?!\*)', r'\1', ai_reason_full)
                    # Удаляем markdown код `text` -> text
                    ai_reason_full = re.sub(r'`(.+?)`', r'\1', ai_reason_full)
                    # Удаляем markdown заголовки # text -> text
                    ai_reason_full = re.sub(r'^#+\s+', '', ai_reason_full, flags=re.MULTILINE)
                    # Удаляем markdown списки - и *
                    ai_reason_full = re.sub(r'^[\-\*]\s+', '', ai_reason_full, flags=re.MULTILINE)
                    # Удаляем markdown ссылки [text](url) -> text
                    ai_reason_full = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', ai_reason_full)
                    # Заменяем множественные переносы строк на одинарные
                    ai_reason_full = re.sub(r'\n{3,}', '\n\n', ai_reason_full)
                    ai_reason_full = ai_reason_full.strip()
            else:
                ai_reason_full = "Нет данных"
            
            # Формируем сообщение (разное для матчей с фаворитом и без)
            if is_match_without_fav:
                # Для не-фаворитов: Матч, Счет, Ставка (bet_alt_code), Кэф (bet_alt_odds), Рекомендация
                bet_alt_code = match['bet_alt_code'] if 'bet_alt_code' in match.keys() else None
                bet_alt_odds = match['bet_alt_odds'] if 'bet_alt_odds' in match.keys() else None
                
                if not bet_alt_code:
                    print(f"[Football Notify] skip: bet_alt_code is NULL for fixture {match['fixture_id']}")
                    return False
                
                message = f"""
⚽ <b>Футбольная аналитика - уведомление</b>

🏟️ <b>Матч:</b> {home_team} vs {away_team}
📊 <b>Счет:</b> {home_score} - {away_score}
🎯 <b>Ставка:</b> {bet_alt_code}
💰 <b>Кэф:</b> {bet_alt_odds if bet_alt_odds else 'N/A'}
📝 <b>Рекомендация:</b> {ai_reason_full}
                """.strip()
            else:
                # Для фаворитов: Матч, Счет, Фаворит, K60, Ставка (bet_ai), Обоснование
                bet_ai = match['bet_ai'] if 'bet_ai' in match.keys() else None
                bet_ai_odds = match['bet_ai_odds'] if 'bet_ai_odds' in match.keys() else None
                live_odds = match['live_odds'] if 'live_odds' in match.keys() else None
                
                if not bet_ai:
                    print(f"[Football Notify] skip: bet_ai is NULL for fixture {match['fixture_id']}")
                    return False
                
                message = f"""
⚽ <b>Футбольная аналитика - уведомление</b>

🏟️ <b>Матч:</b> {home_team} vs {away_team}
📊 <b>Счет:</b> {home_score} - {away_score}
⭐ <b>Фаворит:</b> {fav_team}
💰 <b>K60:</b> {live_odds if live_odds else 'N/A'}
🎯 <b>Ставка:</b> {bet_ai}
📝 <b>Рекомендация:</b> {ai_reason_full}
                """.strip()

            # Получаем список подписчиков
            subscribers = get_football_subscribers()
            
            # Отправляем уведомление только подписчикам
            recipients = set(subscribers)
            
            if not recipients:
                print(f"[Football Notify] no subscribers (0) -> nothing to send for fixture {match['fixture_id']}")
                return False

            # Отправляем уведомление всем получателям
            success_count = 0
            fail_count = 0
            for recipient_id in recipients:
                if telegram_notifier.send_message_to_user(recipient_id, message):
                    success_count += 1
                else:
                    fail_count += 1
            
            if success_count > 0:
                print(f"[Football Notify] sent={success_count} failed={fail_count} total={len(recipients)} fixture={match['fixture_id']}")
                return True
            else:
                print(f"[Football Notify] delivered=0 failed={fail_count} fixture={match['fixture_id']}")
                return False

        except Exception as e:
            print(f"[Football ERROR] Ошибка отправки уведомления: {e}")
            import traceback
            print(traceback.format_exc())
            return False

    def _get_ai_prediction(self, match: sqlite3.Row, stats: Dict) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """
        Получает прогноз от ИИ на основе статистики матча.
        
        Args:
            match: Запись матча из БД
            stats: Статистика на 60-й минуте (из stats_60min)
        
        Returns:
            Кортеж (bet_ai, bet_ai_reason, model_name):
            - bet_ai: Прогноз ('1', '1X', 'X', 'X2', '2') или None
            - bet_ai_reason: Полный ответ от ИИ или None
            - model_name: Имя модели, давшей ответ, или None
        """
        if not self.openrouter_api_key:
            print("[Football] OpenRouter API ключ не установлен, пропускаем ИИ-прогноз")
            return None, None, None
        
        try:
            # Формируем промпт
            home_team = match['home_team']
            away_team = match['away_team']
            fav = match['fav']
            initial_odds = match['initial_odds'] if 'initial_odds' in match.keys() and match['initial_odds'] is not None else '-'
            last_odds = match['last_odds'] if 'last_odds' in match.keys() and match['last_odds'] is not None else '-'
            
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)
            
            # Форматируем статистику как JSON для передачи ИИ
            stats_json = json.dumps(stats, ensure_ascii=False, indent=2)
            
            prompt = f"""Ты - спортивный аналитик. Изучи предоставленную статистику матча после первого тайма, хорошо подумай и сделай прогноз на итоговый результат матча в основное время.

ВАЖНО: Все твои ответы должны быть полностью на русском языке.

Информация о матче:
- Команды: {home_team} vs {away_team}
- Фаворит: {fav} (текущий коэффициент ставки на победу фаворита {last_odds})
- Текущий счет после первого тайма: {home_score} - {away_score}

Детальная статистика первого тайма:
{stats_json}

Ответ верни ТОЛЬКО в виде одного из вариантов: 1 или 1X или X или X2 или 2
Где:
- 1 = победа домашней команды ({home_team})
- 1X = ничья или победа домашней команды ({home_team})
- X = ничья
- X2 = ничья или победа гостевой команды ({away_team})
- 2 = победа гостевой команды ({away_team})"""
            
            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": os.getenv("OPENROUTER_SITE_URL", "http://localhost:5000")
            }
            
            # Список моделей для попыток (основная + три fallback)
            models_to_try = [self.ai_primary_model, self.ai_fallback_model1, self.ai_fallback_model2, self.ai_fallback_model3]
            
            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue
                    
                print(f"[Football AI] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model}")
                
                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "max_tokens": 2000,
                        "temperature": 0.3  # Низкая температура для более детерминированного ответа
                    }
                    
                    print(f"[Football AI] Отправка запроса к OpenRouter API (модель: {model})")
                    
                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=60
                    )
                    
                    if response.status_code == 200:
                        try:
                            data = response.json()
                            if 'choices' in data and len(data['choices']) > 0:
                                ai_response = data['choices'][0]['message']['content']
                                print(f"[Football AI] Получен ответ длиной {len(ai_response)} символов от модели {model}")
                                
                                # Парсим ответ - ищем один из вариантов: 1, 1X, X, X2, 2
                                bet_ai = self._parse_ai_prediction(ai_response)
                                
                                if bet_ai:
                                    print(f"[Football AI] Успешно распознан прогноз: {bet_ai}")
                                    return bet_ai, ai_response, model
                                else:
                                    # Даже если прогноз не распознан, возвращаем полный ответ для сохранения
                                    print(f"[Football AI] Не удалось распознать валидный прогноз в ответе, но сохраняем полный ответ: {ai_response[:200]}...")
                                    return None, ai_response, model
                            else:
                                print(f"[Football AI] Неверный формат ответа от OpenRouter API для модели {model}")
                                continue
                        except json.JSONDecodeError as e:
                            print(f"[Football AI] Ошибка парсинга JSON для модели {model}: {e}")
                            continue
                    else:
                        print(f"[Football AI] HTTP ошибка OpenRouter API для модели {model}: {response.status_code}")
                        try:
                            error_details = response.json()
                            print(f"[Football AI] Детали ошибки: {error_details}")
                            
                            # Если это ошибка 503 "No instances available", переходим к следующей модели
                            if response.status_code == 503 and "No instances available" in str(error_details):
                                print(f"[Football AI] Модель {model} недоступна (503), переходим к следующей")
                                continue
                        except:
                            print(f"[Football AI] Текст ошибки: {response.text[:500]}...")
                        continue
                        
                except requests.exceptions.Timeout:
                    print(f"[Football AI] Таймаут запроса к модели {model}")
                    continue
                except requests.exceptions.RequestException as e:
                    print(f"[Football AI] Ошибка запроса к модели {model}: {e}")
                    continue
                except Exception as e:
                    print(f"[Football AI] Неожиданная ошибка при запросе к модели {model}: {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue
            
            # Если все модели не дали валидного ответа
            print("[Football AI] Все модели не дали валидного прогноза")
            return None, None, None
            
        except Exception as e:
            print(f"[Football AI ERROR] Ошибка получения ИИ-прогноза: {e}")
            import traceback
            print(traceback.format_exc())
            return None, None, None
    
    def _parse_ai_prediction(self, ai_response: str) -> Optional[str]:
        """
        Парсит ответ ИИ и извлекает прогноз (1, 1X, X, X2, 2).
        
        Args:
            ai_response: Полный ответ от ИИ
        
        Returns:
            Прогноз ('1', '1X', 'X', 'X2', '2') или None если не найден
        """
        # Ищем один из вариантов в ответе (регистронезависимо)
        # Используем word boundary чтобы не захватывать часть других слов
        valid_predictions = ['1X', 'X2', '1', 'X', '2']
        
        # Сначала ищем двухсимвольные варианты (1X, X2), потом односимвольные
        for pred in valid_predictions:
            # Используем регулярное выражение для поиска точного совпадения
            pattern = r'\b' + re.escape(pred) + r'\b'
            if re.search(pattern, ai_response, re.IGNORECASE):
                return pred.upper()
        
        return None

    def _parse_ai_recommendation(self, ai_response: str) -> bool:
        """
        Парсит ответ ИИ и извлекает рекомендацию (СТАВИМ/ИГНОРИРУЕМ).
        
        Args:
            ai_response: Полный ответ от ИИ
        
        Returns:
            True если найдено "СТАВИМ", False если "ИГНОРИРУЕМ" или не найдено
        """
        # Ищем слово "СТАВИМ" (регистронезависимо)
        if re.search(r'\bСТАВИМ\b', ai_response, re.IGNORECASE):
            return True
        return False

    def _parse_bet_approve_decision(self, ai_response: Optional[str]) -> Optional[int]:
        """
        Парсит итоговую строку "Резюме: ОДОБРИТЬ/ОТКЛОНИТЬ" из ответа ИИ.

        Args:
            ai_response: Полный ответ от ИИ (может содержать Markdown)

        Returns:
            1 если найдено "Резюме: ОДОБРИТЬ", 0 если "Резюме: ОТКЛОНИТЬ", None если не удалось распознать
        """
        if not ai_response:
            return None

        match_result = re.search(r'Резюме\s*:\s*(ОДОБРИТЬ|ОТКЛОНИТЬ)', ai_response, re.IGNORECASE)
        if not match_result:
            return None

        decision = match_result.group(1).upper()
        if decision == 'ОДОБРИТЬ':
            return 1
        if decision == 'ОТКЛОНИТЬ':
            return 0
        return None

    def _get_bet_ai_decision(self, match: sqlite3.Row, stats: Dict) -> Tuple[Optional[bool], Optional[str]]:
        """
        Получает решение ИИ о том, стоит ли делать ставку (ДА/НЕТ) на основе статистики матча.

        Args:
            match: Запись матча из БД
            stats: Статистика на 60-й минуте (из stats_60min, с raw_data)

        Returns:
            Кортеж (is_yes, ai_reason):
            - is_yes: True если ИИ ответил ДА, False если НЕТ, None если ошибка
            - ai_reason: Полный ответ от ИИ или None
        """
        if not self.openrouter_api_key:
            print("[Football] OpenRouter API ключ не установлен, пропускаем ИИ-решение для bet")
            return None, None

        try:
            # Формируем промпт
            home_team = match['home_team']
            away_team = match['away_team']
            fav = match['fav']
            initial_odds = match['initial_odds'] if 'initial_odds' in match.keys() and match['initial_odds'] is not None else '-'
            last_odds = match['last_odds'] if 'last_odds' in match.keys() and match['last_odds'] is not None else '-'
            
            # Получаем текущий счет
            score = stats.get('score', {})
            home_score = score.get('home', 0)
            away_score = score.get('away', 0)
            
            # Берем сырую статистику из raw_data
            raw_stats = stats.get('raw_data', {})
            
            # Сериализуем статистику в JSON для промпта
            import json
            stats_json = json.dumps(raw_stats, ensure_ascii=False, indent=2)
            
            prompt = f"""Ты - футбольный аналитик. Сейчас начнется второй тайм матча фаворита с аутсайдером. 

Если фаворит не выигрывает, то есть ли статистическая аномалия, при которой фаворит выглядит лучше, чем показывает счёт?

Сделай два шага:

1) Сравни статистику команд и оцени, есть ли сильное доминирование фаворита 

2) Если доминирование сильное и системное и статистика показывает устойчивую аномалию, ответь - «ДА». Во всех прочих случаях ответь «НЕТ»

Информация о матче:
- Команды: {home_team} vs {away_team}
- Фаворит: {fav} (текущий коэффициент ставки на победу фаворита {last_odds})
- Текущий счет после первого тайма: {home_score} - {away_score}

Статистика первого тайма:
{stats_json}"""

            # Заголовки для OpenRouter API
            headers = {
                "Authorization": f"Bearer {self.openrouter_api_key}",
                "HTTP-Referer": "https://github.com",
                "X-Title": "Football Bet Analysis"
            }

            # Пробуем все доступные модели
            models_to_try = [self.ai_primary_model, self.ai_fallback_model1, self.ai_fallback_model2, self.ai_fallback_model3]

            for model_idx, model in enumerate(models_to_try):
                if not model:
                    continue

                print(f"[Football Bet AI] Пробуем модель {model_idx + 1}/{len(models_to_try)}: {model}")

                try:
                    payload = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "max_tokens": 500,
                        "temperature": 0.3  # Низкая температура для более детерминированного ответа
                    }

                    print(f"[Football Bet AI] Отправка запроса к OpenRouter API (модель: {model})")

                    response = requests.post(
                        f"{self.openrouter_api_url}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=60
                    )

                    if response.status_code == 200:
                        try:
                            data = response.json()
                            if 'choices' in data and len(data['choices']) > 0:
                                ai_response = data['choices'][0]['message']['content']
                                print(f"[Football Bet AI] Получен ответ длиной {len(ai_response)} символов от модели {model}")

                                # Парсим ответ - ищем ДА или НЕТ
                                is_yes = self._parse_bet_ai_response(ai_response)

                                if is_yes is not None:
                                    print(f"[Football Bet AI] Успешно распознан ответ: {'ДА' if is_yes else 'НЕТ'}")
                                    return is_yes, ai_response
                                else:
                                    print(f"[Football Bet AI] Не удалось распознать ДА/НЕТ в ответе: {ai_response[:200]}...")
                                    # Продолжаем с следующей моделью
                                    continue
                            else:
                                print(f"[Football Bet AI] Неверный формат ответа от OpenRouter API для модели {model}")
                                continue
                        except json.JSONDecodeError as e:
                            print(f"[Football Bet AI] Ошибка парсинга JSON для модели {model}: {e}")
                            continue
                    else:
                        print(f"[Football Bet AI] HTTP ошибка OpenRouter API для модели {model}: {response.status_code}")
                        try:
                            error_details = response.json()
                            print(f"[Football Bet AI] Детали ошибки: {error_details}")

                            # Если это ошибка 503 "No instances available", переходим к следующей модели
                            if response.status_code == 503 and "No instances available" in str(error_details):
                                print(f"[Football Bet AI] Модель {model} недоступна (503), переходим к следующей")
                                continue
                        except:
                            print(f"[Football Bet AI] Текст ошибки: {response.text[:500]}...")
                        continue

                except requests.exceptions.Timeout:
                    print(f"[Football Bet AI] Таймаут запроса к модели {model}")
                    continue
                except requests.exceptions.RequestException as e:
                    print(f"[Football Bet AI] Ошибка запроса к модели {model}: {e}")
                    continue
                except Exception as e:
                    print(f"[Football Bet AI] Неожиданная ошибка при запросе к модели {model}: {e}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            # Если все модели не дали валидного ответа
            print("[Football Bet AI] Все модели не дали валидного ответа")
            return None, None

        except Exception as e:
            print(f"[Football Bet AI ERROR] Ошибка получения ИИ-решения: {e}")
            import traceback
            print(traceback.format_exc())
            return None, None

    def _parse_bet_ai_response(self, ai_response: str) -> Optional[bool]:
        """
        Парсит ответ ИИ и извлекает ДА/НЕТ.

        Args:
            ai_response: Полный ответ от ИИ

        Returns:
            True если ДА, False если НЕТ, None если не найден
        """
        # Ищем ДА или НЕТ в ответе (регистронезависимо)
        # Используем word boundary чтобы не захватывать часть других слов
        response_upper = ai_response.upper().strip()
        
        # Ищем ДА (может быть написано как "ДА", "ДА.", "ДА!", "ДА," и т.д.)
        if re.search(r'\bДА\b', response_upper):
            return True
        
        # Ищем НЕТ (может быть написано как "НЕТ", "НЕТ.", "НЕТ!", "НЕТ," и т.д.)
        if re.search(r'\bНЕТ\b', response_upper):
            return False
        
        return None

# === Функции для работы с подписками Telegram ===

# Временное сопоставление "токен -> user_id" только в памяти процесса,
# чтобы UI мог узнать статус по токену, не храня токены в БД
_football_token_bindings: Dict[str, str] = {}

def bind_token_to_user(token: str, user_id: str) -> None:
    """Связывает одноразовый токен с user_id в памяти процесса."""
    if token:
        _football_token_bindings[token] = str(user_id)

def add_football_subscription(user_id: str) -> bool:
    """
    Добавляет подписку пользователя на уведомления о футболе.
    
    Args:
        token: Временный токен, сгенерированный на странице
        user_id: ID пользователя Telegram (chat_id)
    
    Returns:
        True если подписка добавлена успешно, False в случае ошибки
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        # Включаем подписку через UPSERT по user_id (без токена)
        cursor.execute("""
            INSERT INTO football_telegram_subscriptions (user_id, created_at, is_active)
            VALUES (?, CURRENT_TIMESTAMP, 1)
            ON CONFLICT(user_id) DO UPDATE SET
                is_active=1,
                created_at=CURRENT_TIMESTAMP
        """, (user_id,))
        
        conn.commit()
        print(f"[Football] Подписка активирована: user_id={user_id}")
        return True
        
    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка добавления подписки: {e}")
        return False
    finally:
        if conn:
            conn.close()


def remove_football_subscription(user_id: str) -> bool:
    """
    Удаляет подписку пользователя на уведомления о футболе.
    
    Args:
        user_id: ID пользователя Telegram (chat_id)
    
    Returns:
        True если подписка удалена успешно, False в случае ошибки
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE football_telegram_subscriptions
            SET is_active = 0
            WHERE user_id = ? AND is_active = 1
        """, (user_id,))
        
        conn.commit()
        affected = cursor.rowcount
        print(f"[Football] Подписка удалена (идемпотентно): user_id={user_id}, affected={affected}")
        # Идемпотентность: даже если уже был отписан (affected=0), считаем операцию успешной
        return True
        
    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка удаления подписки: {e}")
        return False
    finally:
        if conn:
            conn.close()


def get_football_subscribers() -> List[str]:
    """
    Получает список всех активных подписчиков на уведомления о футболе.
    
    Returns:
        Список user_id активных подписчиков
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT DISTINCT user_id FROM football_telegram_subscriptions
            WHERE is_active = 1
        """)
        
        rows = cursor.fetchall()
        return [row[0] for row in rows]
        
    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка получения подписчиков: {e}")
        return []
    finally:
        if conn:
            conn.close()


def is_football_subscribed_by_token(token: str) -> bool:
    """
    Проверяет, есть ли активная подписка по токену.
    
    Args:
        token: Временный токен
    
    Returns:
        True если есть активная подписка с таким токеном, False иначе
    """
    # Токены больше не храним в БД. Если токен был использован, он будет
    # связан в памяти процесса с конкретным user_id. Проверяем по привязке.
    user_id = _football_token_bindings.get(token)
    if not user_id:
        return False
    return is_football_subscribed(user_id)


def is_football_subscribed(user_id: str) -> bool:
    """
    Проверяет, подписан ли пользователь на уведомления о футболе.
    
    Args:
        user_id: ID пользователя Telegram (chat_id)
    
    Returns:
        True если пользователь подписан, False иначе
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) FROM football_telegram_subscriptions
            WHERE user_id = ? AND is_active = 1
        """, (user_id,))
        
        count = cursor.fetchone()[0]
        return count > 0
        
    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка проверки подписки: {e}")
        return False
    finally:
        if conn:
            conn.close()

# === Функции для APScheduler ===

def collect_tomorrow_matches_task():
    """Задача для планировщика - сбор матчей на завтра."""
    try:
        manager = get_manager()
        count = manager.collect_tomorrow_matches()
        print(f"[Football] Задача сбора завершена: {count} матчей")
        return count
    except Exception as e:
        print(f"[Football] Ошибка в задаче сбора: {e}")
        import traceback
        print(traceback.format_exc())
        return 0


def check_matches_and_collect_task():
    """Задача для планировщика - проверка матчей и сбор статистики."""
    try:
        manager = get_manager()
        manager.check_matches_and_collect()
    except Exception as e:
        print(f"[Football] Ошибка в задаче проверки: {e}")
        import traceback
        print(traceback.format_exc())


def check_matches_60min_task():
    """Задача для планировщика - детектор 60-й минуты и обновление статусов (без финального счета)."""
    try:
        manager = get_manager()
        manager.check_matches_60min_and_status()
    except Exception as e:
        print(f"[Football] Ошибка в задаче детектора 60-й минуты: {e}")
        import traceback
        print(traceback.format_exc())

def thesportsdb_update_scores_task():
    """Задача для планировщика - обновление текущих счетов через TheSportsDB для матчей in_progress."""
    try:
        manager = get_manager()
        n = manager.update_inprogress_scores_from_thesportsdb()
        if n:
            print(f"[Football] Обновлены текущие счета (TheSportsDB): {n}")
    except Exception as e:
        print(f"[Football] Ошибка при обновлении счетов из TheSportsDB: {e}")
        import traceback
        print(traceback.format_exc())


def get_all_matches(filter_fav: bool = True) -> List[Dict[str, Any]]:
    """
    Получает матчи для UI.
    
    Args:
        filter_fav: Если True, возвращает только матчи с фаворитом (fav != 'NONE').
                    Если False, возвращает все матчи.

    Returns:
        Список матчей
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()

        # Исключаем большие поля: bet_ai_full_response, bet_ai_reason (не используются в шаблоне)
        # Оставляем stats_60min, так как он используется для tooltip
        if filter_fav:
            cursor.execute("""
                SELECT id, fixture_id, sofascore_event_id, home_team, away_team, fav, fav_team_id,
                       match_date, match_time, initial_odds, last_odds, live_odds, live_odds_1, live_odds_x, live_odds_2,
                       status, stats_60min, bet, bet_ai, bet_ai_odds, bet_ai_model_name,
                       bet_alt_code, bet_alt_odds, bet_alt_confirm,
                       final_score_home, final_score_away, fav_won, sport_key,
                       created_at, updated_at
                FROM matches
                WHERE fav != 'NONE'
                ORDER BY match_date DESC, match_time DESC
            """)
        else:
            cursor.execute("""
                SELECT id, fixture_id, sofascore_event_id, home_team, away_team, fav, fav_team_id,
                       match_date, match_time, initial_odds, last_odds, live_odds, live_odds_1, live_odds_x, live_odds_2,
                       status, stats_60min, bet, bet_ai, bet_ai_odds, bet_ai_model_name,
                       bet_alt_code, bet_alt_odds, bet_alt_confirm,
                       final_score_home, final_score_away, fav_won, sport_key,
                       created_at, updated_at
                FROM matches
                ORDER BY match_date DESC, match_time DESC
            """)

        rows = cursor.fetchall()
        return [dict(row) for row in rows]

    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка получения матчей: {e}")
        return []
    finally:
        if conn:
            conn.close()


def get_api_limits() -> Dict[str, Any]:
    """
    Получает текущие лимиты API для UI.

    Returns:
        Словарь с информацией о лимитах API
    """
    try:
        manager = get_manager()
        return {
            'requests_remaining': manager.requests_remaining,
            'requests_used': manager.requests_used,
            'requests_last_cost': manager.requests_last_cost
        }
    except Exception as e:
        print(f"[Football ERROR] Ошибка получения лимитов API: {e}")
        return {
            'requests_remaining': None,
            'requests_used': None,
            'requests_last_cost': None
        }


def _is_prediction_win(prediction: str, actual_result: str) -> bool:
    """
    Проверяет, выиграл ли прогноз на основе фактического результата.
    
    Args:
        prediction: Прогноз (1, X, 2, 1X, X2)
        actual_result: Фактический результат (1, X, 2)
    
    Returns:
        True если прогноз выиграл, False иначе
    """
    if not prediction or not actual_result:
        return False
    
    pred_upper = prediction.upper()
    
    if pred_upper == '1':
        return actual_result == '1'
    elif pred_upper == 'X':
        return actual_result == 'X'
    elif pred_upper == '2':
        return actual_result == '2'
    elif pred_upper == '1X':
        return actual_result == '1' or actual_result == 'X'
    elif pred_upper == 'X2':
        return actual_result == 'X' or actual_result == '2'
    elif pred_upper == '12':
        return actual_result == '1' or actual_result == '2'
    
    return False


def _is_alternative_bet_win(bet_alt_code: str, home_score: int, away_score: int) -> bool:
    """
    Проверяет, выиграла ли альтернативная ставка.
    
    Args:
        bet_alt_code: Код альтернативной ставки
        home_score: Счет домашней команды
        away_score: Счет гостевой команды
    
    Returns:
        True если ставка выиграла, False иначе
    """
    if not bet_alt_code or home_score is None or away_score is None:
        return False
    
    # Убираем пробелы и переводим в верхний регистр
    code = bet_alt_code.strip().upper()
    
    # Определяем фактический результат
    if home_score > away_score:
        actual_result = '1'
    elif home_score == away_score:
        actual_result = 'X'
    else:
        actual_result = '2'
    
    # Проверяем простые ставки 1X2
    if code in ['1', 'X', '2', '1X', 'X2', '12']:
        return _is_prediction_win(code, actual_result)
    
    # Гандикап: Ф1-1.5, Ф1+0.5, Ф2-2.5, Ф2+1.0
    handicap_match = re.match(r'^Ф([12])([+-]?)(\d+\.?\d*)$', code)
    if handicap_match:
        team = handicap_match.group(1)  # 1 или 2
        sign = handicap_match.group(2)  # + или -
        value = float(handicap_match.group(3))
        
        if team == '1':
            adjusted_home = home_score + (value if sign == '+' else -value)
            return adjusted_home > away_score
        else:  # team == '2'
            adjusted_away = away_score + (value if sign == '+' else -value)
            return adjusted_away > home_score
    
    # Тотал: Б, М, T2.5Б, T2.5М, Т2.5Б, Т2.5М и т.д.
    # Может быть формат: Б2.5, М2.5, T2.5Б, T2.5М, Т2.5Б, Т2.5М (латинская или кириллическая Т)
    total_goals = home_score + away_score
    
    # Проверяем формат с префиксом T (латинская) или Т (кириллическая): T2.5Б, T2.5М, Т2.5Б, Т2.5М
    if code.startswith('T') or code.startswith('Т'):  # Проверяем и латинскую, и кириллическую Т
        # Извлекаем число и букву Б/М
        # Формат: T2.5Б или T2.5М или Т2.5Б или Т2.5М
        try:
            # Убираем первый символ (T или Т) и парсим остальное
            rest = code[1:]
            if rest.endswith('Б') or rest.endswith('М'):
                over_under = rest[-1]
                threshold_str = rest[:-1]
                try:
                    threshold = float(threshold_str)
                    if over_under == 'Б':
                        return total_goals > threshold
                    else:  # М
                        return total_goals < threshold
                except ValueError as e:
                    print(f"[Football ERROR] Ошибка парсинга числа в тотале: {code}, threshold_str={threshold_str}, {e}")
        except Exception as e:
            print(f"[Football ERROR] Ошибка парсинга тотала с префиксом T/Т: {code}, {e}")
    
    # Старый формат без префикса T: Б2.5, М2.5
    if code.startswith('Б') or code.startswith('М'):
        over_under = code[0]
        threshold_str = code[1:]
        try:
            threshold = float(threshold_str)
            if over_under == 'Б':
                return total_goals > threshold
            else:  # М
                return total_goals < threshold
        except ValueError:
            # Пробуем регулярное выражение как запасной вариант
            total_match = re.match(r'^([БМ])(\d+\.?\d*)$', code)
            if total_match:
                over_under = total_match.group(1)
                threshold = float(total_match.group(2))
                if over_under == 'Б':
                    return total_goals > threshold
                else:  # М
                    return total_goals < threshold
    
    return False


def _parse_total_bet_code(bet_alt_code: str) -> Optional[Tuple[float, str]]:
    """
    Парсит код тотала и возвращает (threshold, over_under).
    
    Args:
        bet_alt_code: Код ставки (например, "T2.5Б", "Т4.5М")
    
    Returns:
        Tuple (threshold, over_under) где over_under = 'Б' или 'М', или None если не тотал
    """
    if not bet_alt_code:
        return None
    
    code = bet_alt_code.strip().upper()
    
    # Проверяем формат с префиксом T (латинская) или Т (кириллическая)
    if code.startswith('T') or code.startswith('Т'):
        rest = code[1:]
        if rest.endswith('Б') or rest.endswith('М'):
            over_under = rest[-1]
            threshold_str = rest[:-1]
            try:
                threshold = float(threshold_str)
                return (threshold, over_under)
            except ValueError:
                return None
    
    # Старый формат без префикса T
    if code.startswith('Б') or code.startswith('М'):
        over_under = code[0]
        threshold_str = code[1:]
        try:
            threshold = float(threshold_str)
            return (threshold, over_under)
        except ValueError:
            return None
    
    return None


def _calculate_expected_odds_range(total_goals: int, threshold: float, over_under: str) -> Tuple[float, float]:
    """
    Определяет ожидаемый диапазон коэффициента на основе разницы голов.
    Пессимистичный подход - возвращает верхние границы диапазонов.
    
    Args:
        total_goals: Текущее количество голов на 60-й минуте
        threshold: Линия тотала
        over_under: 'Б' (больше) или 'М' (меньше)
    
    Returns:
        Tuple (min_odds, max_odds) - ожидаемый диапазон коэффициента
    """
    if over_under == 'Б':
        goals_needed = threshold - total_goals + 0.5  # Сколько голов нужно для прохода
    else:  # М
        goals_needed = total_goals - threshold + 0.5  # Сколько голов уже "лишних"
    
    # Пессимистичные диапазоны (верхние границы)
    if goals_needed <= 0:
        # Ставка уже прошла или почти прошла
        return (1.01, 1.05)
    elif goals_needed <= 0.5:
        # Нужен 1 гол
        return (1.10, 1.25)
    elif goals_needed <= 1.0:
        # Нужно 1-1.5 гола
        return (1.25, 1.50)
    elif goals_needed <= 1.5:
        # Нужно 1.5-2 гола
        return (1.50, 1.80)
    elif goals_needed <= 2.0:
        # Нужно 2-2.5 гола
        return (1.80, 2.20)
    else:
        # Нужно 3+ гола
        return (2.20, 3.00)


def _recalculate_total_odds_pessimistic(total_goals: int, threshold: float, over_under: str, goals_per_minute: float) -> float:
    """
    Пересчитывает коэффициент для тотала пессимистично на основе счета и темпа.
    Учитывает, что при высоком темпе вероятность забить несколько голов выше.
    
    Args:
        total_goals: Текущее количество голов на 60-й минуте
        threshold: Линия тотала
        over_under: 'Б' (больше) или 'М' (меньше)
        goals_per_minute: Темп игры (голы в минуту)
    
    Returns:
        Пересчитанный коэффициент (пессимистично)
    """
    if over_under == 'Б':
        # Для Over: нужно чтобы total_goals > threshold
        # Если уже прошло (total_goals > threshold) - минимальный коэффициент
        if total_goals > threshold:
            return 1.03  # Ставка уже прошла
        # Сколько голов нужно для прохода (нужно чтобы было > threshold)
        goals_needed = threshold - total_goals + 0.5
    else:  # М (Under)
        # Для Under: нужно чтобы total_goals < threshold
        # Если уже не прошло (total_goals >= threshold) - максимальный коэффициент
        if total_goals >= threshold:
            return 2.50  # Ставка уже не пройдет
        # Для Under: считаем сколько голов "в запасе" до провала
        # Например, при 0 голах и линии 1.5 - можно забить еще максимум 1 гол
        goals_remaining_allowed = threshold - total_goals - 0.5  # Максимум голов, которые можно забить
    
    # Прогнозируем количество голов за оставшиеся 30 минут на основе темпа
    # Пессимистично: уменьшаем темп на 20% (учитываем усталость, тактику)
    predicted_goals_30min = goals_per_minute * 30 * 0.80
    
    # Базовый коэффициент на основе прогноза и нужного количества голов
    if over_under == 'Б':
        # Для Over: если прогноз >= нужного количества - очень низкий коэффициент
        if predicted_goals_30min >= goals_needed:
            # Прогноз покрывает нужное количество - очень низкий коэффициент
            # Чем выше темп, тем ниже коэффициент (выше вероятность)
            if goals_needed <= 0.5:
                # Нужен 1 гол, прогноз его покрывает
                if goals_per_minute >= 0.083:  # 5+ голов за 60 мин
                    base_odds = 1.05
                elif goals_per_minute >= 0.067:  # 4+ голов за 60 мин
                    base_odds = 1.08
                else:
                    base_odds = 1.12
            elif goals_needed <= 1.0:
                # Нужно 1-1.5 гола, прогноз покрывает
                if goals_per_minute >= 0.10:  # 6+ голов за 60 мин
                    base_odds = 1.08
                elif goals_per_minute >= 0.083:  # 5+ голов за 60 мин
                    base_odds = 1.12
                elif goals_per_minute >= 0.067:  # 4+ голов за 60 мин
                    base_odds = 1.15
                elif goals_per_minute >= 0.05:  # 3+ голов за 60 мин
                    base_odds = 1.18
                else:
                    base_odds = 1.22
            elif goals_needed <= 1.5:
                # Нужно 1.5-2 гола, прогноз покрывает
                if goals_per_minute >= 0.10:  # 6+ голов за 60 мин
                    base_odds = 1.20
                elif goals_per_minute >= 0.083:  # 5+ голов за 60 мин
                    base_odds = 1.28
                else:
                    base_odds = 1.35
            elif goals_needed <= 2.0:
                # Нужно 2-2.5 гола, прогноз покрывает
                if goals_per_minute >= 0.10:  # 6+ голов за 60 мин
                    base_odds = 1.30
                elif goals_per_minute >= 0.083:  # 5+ голов за 60 мин
                    base_odds = 1.40
                else:
                    base_odds = 1.50
            else:
                # Нужно 3+ гола, прогноз покрывает
                if goals_per_minute >= 0.12:  # 7+ голов за 60 мин
                    base_odds = 1.50
                else:
                    base_odds = 1.80
        else:
            # Прогноз не покрывает - коэффициент выше, но не слишком
            deficit = goals_needed - predicted_goals_30min
            if deficit <= 0.5:
                base_odds = 1.25
            elif deficit <= 1.0:
                base_odds = 1.50
            elif deficit <= 1.5:
                base_odds = 1.85
            else:
                base_odds = 2.30
    else:  # М (Under)
        # Для Under: риск выше - достаточно одного гола, чтобы провалить ставку
        # При нулевом темпе коэффициенты должны быть ВЫШЕ, независимо от запаса
        # Потому что вероятность забить голы есть, но не гарантирована
        if predicted_goals_30min <= goals_remaining_allowed:
            # Прогноз в пределах разрешенного
            # Считаем "запас безопасности" - насколько прогноз ниже разрешенного
            safety_margin = goals_remaining_allowed - predicted_goals_30min
            
            # При нулевом темпе: коэффициенты ВЫШЕ, так как риск провала есть
            if goals_per_minute == 0 or goals_per_minute < 0.001:
                # Нулевой темп - высокая неопределенность
                if safety_margin >= 2.0:
                    # Очень большой запас (можно забить 2+ гола) - но риск есть
                    # T2.5M при 0 голах: можно забить 2 гола → коэффициент 1.75
                    base_odds = 1.75
                elif safety_margin >= 1.5:
                    # Большой запас (можно забить 1.5+ гола)
                    base_odds = 1.60
                elif safety_margin >= 1.0:
                    # Средний запас (можно забить 1 гол) - T1.5M при 0 голах
                    base_odds = 1.25
                elif safety_margin >= 0.5:
                    # Маленький запас (можно забить 0.5 гола)
                    base_odds = 1.40
                else:
                    # Очень маленький запас - высокий риск
                    base_odds = 1.60
            else:
                # Ненулевой темп - используем логику с учетом запаса и темпа
                if safety_margin >= 2.0:
                    # Очень большой запас (прогноз на 2+ гола ниже разрешенного)
                    if goals_per_minute < 0.05:
                        base_odds = 1.06
                    else:
                        base_odds = 1.04
                elif safety_margin >= 1.5:
                    # Большой запас (прогноз на 1.5+ гола ниже)
                    if goals_per_minute < 0.05:
                        base_odds = 1.14
                    else:
                        base_odds = 1.12
                elif safety_margin >= 1.0:
                    # Средний запас (прогноз на 1+ гол ниже)
                    if goals_per_minute < 0.05:
                        base_odds = 1.22
                    elif goals_per_minute < 0.067:
                        base_odds = 1.20
                    else:
                        base_odds = 1.18
                elif safety_margin >= 0.5:
                    # Маленький запас (прогноз на 0.5+ гола ниже)
                    if goals_per_minute < 0.05:
                        base_odds = 1.35
                    elif goals_per_minute < 0.067:
                        base_odds = 1.30
                    else:
                        base_odds = 1.28
                else:
                    # Очень маленький запас - высокий риск
                    if goals_per_minute < 0.05:
                        base_odds = 1.55
                    else:
                        base_odds = 1.50
        else:
            # Прогноз превышает разрешенное - ставка под угрозой, очень высокий коэффициент
            excess = predicted_goals_30min - goals_remaining_allowed
            if excess <= 0.5:
                base_odds = 2.00
            elif excess <= 1.0:
                base_odds = 2.40
            else:
                base_odds = 2.80
    
    # Ограничиваем диапазон
    return max(1.01, min(3.00, round(base_odds, 2)))


def recalculate_alt_bet_confirm():
    """
    Пересчитывает bet_alt_confirm для всех матчей по алгоритму:
    Если bet_alt_odds <= bet_ai_odds и bet_alt_odds > 1.10, то bet_alt_confirm=1, иначе 0
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        # Находим все матчи с bet_alt_code и bet_alt_odds
        cursor.execute("""
            SELECT id, fixture_id, bet_alt_code, bet_alt_odds, bet_ai_odds
            FROM matches
            WHERE bet_alt_code IS NOT NULL 
              AND bet_alt_code != ''
              AND bet_alt_odds IS NOT NULL
        """)
        
        rows = cursor.fetchall()
        updated_count = 0
        
        for row in rows:
            match_id = row['id']
            fixture_id = row['fixture_id']
            bet_alt_odds = row['bet_alt_odds']
            bet_ai_odds = row['bet_ai_odds']
            
            # Вычисляем bet_alt_confirm по алгоритму
            if bet_ai_odds is not None and bet_alt_odds <= bet_ai_odds and bet_alt_odds > 1.10:
                new_confirm = 1
            else:
                new_confirm = 0
            
            # Обновляем bet_alt_confirm
            cursor.execute("""
                UPDATE matches 
                SET bet_alt_confirm = ?
                WHERE id = ?
            """, (new_confirm, match_id))
            
            updated_count += 1
        
        conn.commit()
        print(f"[Football] Пересчет bet_alt_confirm завершен. Обновлено матчей: {updated_count}")
        return {'updated': updated_count}
        
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"[Football ERROR] Ошибка пересчета bet_alt_confirm: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}
    finally:
        if conn:
            conn.close()


def recalculate_alt_bet_odds_for_totals():
    """
    Пересчитывает коэффициенты для альтернативных ставок (тоталы) на основе статистики 60-й минуты.
    Обновляет только те матчи, где текущий коэффициент явно не соответствует ожидаемому диапазону.
    Пессимистичный подход - завышает риски.
    """
    conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        # Находим все матчи с stats_60min и bet_alt_code
        cursor.execute("""
            SELECT id, fixture_id, bet_alt_code, bet_alt_odds, stats_60min
            FROM matches
            WHERE stats_60min IS NOT NULL 
              AND stats_60min != ''
              AND bet_alt_code IS NOT NULL
              AND bet_alt_code != ''
        """)
        
        rows = cursor.fetchall()
        updated_count = 0
        reset_count = 0
        
        for row in rows:
            match_id = row['id']
            fixture_id = row['fixture_id']
            bet_alt_code = row['bet_alt_code']
            current_odds = row['bet_alt_odds']
            stats_60min_str = row['stats_60min']
            
            # Парсим код тотала - ВАЖНО: пересчитываем ТОЛЬКО тоталы, гандикапы и другие типы ставок пропускаем
            total_info = _parse_total_bet_code(bet_alt_code)
            if not total_info:
                continue  # Не тотал (гандикап, 1X2 и т.д.), пропускаем
            
            threshold, over_under = total_info
            
            # Парсим статистику
            try:
                stats = json.loads(stats_60min_str) if isinstance(stats_60min_str, str) else stats_60min_str
            except Exception:
                # Не удалось распарсить - сбрасываем в 1
                cursor.execute("""
                    UPDATE matches SET bet_alt_odds = 1.0 WHERE id = ?
                """, (match_id,))
                conn.commit()
                reset_count += 1
                continue
            
            # Извлекаем счет
            if not stats or 'score' not in stats:
                # Нет счета - сбрасываем в 1
                cursor.execute("""
                    UPDATE matches SET bet_alt_odds = 1.0 WHERE id = ?
                """, (match_id,))
                conn.commit()
                reset_count += 1
                continue
            
            score = stats.get('score', {})
            home_score = score.get('home')
            away_score = score.get('away')
            
            if home_score is None or away_score is None:
                # Нет счета - сбрасываем в 1
                cursor.execute("""
                    UPDATE matches SET bet_alt_odds = 1.0 WHERE id = ?
                """, (match_id,))
                conn.commit()
                reset_count += 1
                continue
            
            total_goals = int(home_score) + int(away_score)
            
            # Вычисляем темп игры (голы в минуту)
            goals_per_minute = total_goals / 60.0
            
            # Пересчитываем ВСЕ коэффициенты, независимо от того, находятся ли они в диапазоне или нет
            # Потому что новая логика расчета может дать более точные значения
            new_odds = _recalculate_total_odds_pessimistic(total_goals, threshold, over_under, goals_per_minute)
            
            # Обновляем в БД
            cursor.execute("""
                UPDATE matches SET bet_alt_odds = ? WHERE id = ?
            """, (new_odds, match_id))
            conn.commit()
            updated_count += 1
        
        return {
            'updated': updated_count,
            'reset': reset_count,
            'total_processed': len(rows)
        }
        
    except Exception as e:
        print(f"[Football ERROR] Ошибка пересчета коэффициентов: {e}")
        import traceback
        print(traceback.format_exc())
        return None
    finally:
        if conn:
            conn.close()


def export_matches_to_excel(date_filter: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, 
                             match_type: str = 'fav', timezone_offset: Optional[int] = None) -> Optional[BytesIO]:
    """
    Экспортирует матчи в Excel файл с результатами ставок.
    Экспортирует только то, что показано на странице (с учетом фильтра).
    
    Args:
        date_filter: Тип фильтра ('all', 'today', 'yesterday', 'tomorrow', 'range')
        date_from: Начальная дата для фильтра 'range' (формат YYYY-MM-DD)
        date_to: Конечная дата для фильтра 'range' (формат YYYY-MM-DD)
        match_type: Тип матчей ('fav' - с фаворитом, 'all' - все матчи)
        timezone_offset: Смещение часового пояса пользователя в минутах (например, 180 для GMT+3)
    
    Returns:
        BytesIO объект с Excel файлом или None в случае ошибки
    """
    if not OPENPYXL_AVAILABLE:
        print("[Football ERROR] openpyxl не установлен, экспорт в Excel недоступен")
        return None
    
        conn = None
    try:
        conn = get_football_db_connection()
        cursor = conn.cursor()
        
        # Формируем условие фильтра по типу матчей
        if match_type == 'fav':
            match_condition = "fav != 'NONE'"
        else:  # 'all'
            match_condition = "1=1"  # Все матчи
        
        # Формируем условие фильтра по дате с учетом часового пояса пользователя
        date_condition = ""
        
        if date_filter and date_filter != 'all':
            from datetime import datetime, timedelta, timezone
            
            # Получаем текущее время в UTC
            utc_now = datetime.now(timezone.utc)
            
            # Применяем смещение часового пояса пользователя (если указано)
            if timezone_offset is not None:
                user_tz = timezone(timedelta(minutes=timezone_offset))
                user_now = utc_now.astimezone(user_tz)
            else:
                # Если смещение не указано, используем UTC
                user_now = utc_now
            
            # Получаем дату в часовом поясе пользователя
            user_date = user_now.date()
            
            if date_filter == 'today':
                date_condition = f" AND match_date = '{user_date}'"
            elif date_filter == 'yesterday':
                yesterday = user_date - timedelta(days=1)
                date_condition = f" AND match_date = '{yesterday}'"
            elif date_filter == 'tomorrow':
                tomorrow = user_date + timedelta(days=1)
                date_condition = f" AND match_date = '{tomorrow}'"
            elif date_filter == 'range' and date_from and date_to:
                date_condition = f" AND match_date >= '{date_from}' AND match_date <= '{date_to}'"
        
        # Формируем полное условие WHERE
        where_condition = match_condition + date_condition
        
        # Получаем матчи с учетом фильтров
        query = f"""
            SELECT 
                fixture_id,
                home_team,
                away_team,
                fav,
                match_date,
                match_time,
                initial_odds,
                last_odds,
                live_odds,
                status,
                bet,
                bet_ai,
                bet_ai_odds,
                bet_alt_code,
                bet_alt_odds,
                bet_alt_confirm,
                final_score_home,
                final_score_away,
                fav_won
            FROM matches
            WHERE {where_condition}
            ORDER BY match_date DESC, match_time DESC
        """
        cursor.execute(query)
        
        rows = cursor.fetchall()
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Матчи"
        
        # Заголовки (как на странице)
        headers = [
            "Дата",
            "Время",
            "Домашняя команда",
            "Гостевая команда",
            "Фаворит",
            "Начальные коэффициенты",
            "Последние коэффициенты",
            "Коэффициент на 60 мин",
            "Прогноз ИИ",
            "Коэф. ИИ",
            "Результат прогноза ИИ",
            "Ставка",
            "Альтернативная ставка",
            "Коэф. Alt",
            "Alt Bet",
            "Результат Alt ставки",
            "Финальный счет",
            "Результат для фаворита",
            "Статус"
        ]
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для результатов
        win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый
        loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный
        draw_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Записываем данные
        for row_num, row in enumerate(rows, 2):
            # Распаковываем данные
            (fixture_id, home_team, away_team, fav, match_date, match_time,
             initial_odds, last_odds, live_odds, status, bet,
             bet_ai, bet_ai_odds, bet_alt_code, bet_alt_odds, bet_alt_confirm,
             final_score_home, final_score_away, fav_won) = row
            
            # Определяем фактический результат матча
            actual_result = None
            if status == 'finished' and final_score_home is not None and final_score_away is not None:
                home_score = int(final_score_home)
                away_score = int(final_score_away)
                if home_score > away_score:
                    actual_result = '1'
                elif home_score == away_score:
                    actual_result = 'X'
                else:
                    actual_result = '2'
            
            # Результат прогноза ИИ
            bet_ai_result = ""
            if bet_ai and actual_result:
                if _is_prediction_win(bet_ai, actual_result):
                    bet_ai_result = "Выиграл"
                else:
                    bet_ai_result = "Проиграл"
            elif bet_ai:
                bet_ai_result = "Не завершен"
            
            # Результат альтернативной ставки
            bet_alt_result = ""
            if bet_alt_code:
                if status == 'finished' and final_score_home is not None and final_score_away is not None:
                    home_score = int(final_score_home)
                    away_score = int(final_score_away)
                    result = _is_alternative_bet_win(bet_alt_code, home_score, away_score)
                    if result:
                        bet_alt_result = "Выиграл"
                    else:
                        bet_alt_result = "Проиграл"
                else:
                    bet_alt_result = "Не завершен"
            
            # Результат для фаворита
            fav_result = ""
            if status == 'finished':
                if final_score_home is not None and final_score_away is not None:
                    home_score = int(final_score_home)
                    away_score = int(final_score_away)
                    if home_score == away_score:
                        fav_result = "Ничья"
                    elif fav_won == 1:
                        fav_result = "Выиграл"
                    elif fav_won == 0:
                        fav_result = "Проиграл"
                    else:
                        fav_result = "Не определен"
                else:
                    fav_result = "Нет счета"
            else:
                fav_result = "Не завершен"
            
            # Форматируем финальный счет
            final_score = ""
            if final_score_home is not None and final_score_away is not None:
                final_score = f"{final_score_home}-{final_score_away}"
            
            # Форматируем коэффициенты
            def format_odds(odds):
                if odds is None:
                    return ""
                try:
                    return f"{float(odds):.2f}"
                except:
                    return str(odds)
            
            # Форматируем статус
            status_text = {
                'scheduled': 'Запланирован',
                'in_progress': 'Идет',
                'finished': 'Завершен'
            }.get(status, status)
            
            # Форматируем ставку
            bet_text = "Да" if (bet and float(bet) >= 1) else ""
            
            # Форматируем Alt Bet
            alt_bet_text = "Да" if (bet_alt_confirm and int(bet_alt_confirm) == 1) else ""
            
            # Подготавливаем данные для записи
            data = [
                match_date or "",
                match_time or "",
                home_team or "",
                away_team or "",
                fav or "",
                format_odds(initial_odds),
                format_odds(last_odds),
                format_odds(live_odds),
                bet_ai or "",
                format_odds(bet_ai_odds),
                bet_ai_result,
                bet_text,
                bet_alt_code or "",
                format_odds(bet_alt_odds),
                alt_bet_text,
                bet_alt_result,
                final_score,
                fav_result,
                status_text
            ]
            
            # Записываем данные
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Применяем цветовую подсветку для результатов
                if col_num == 11:  # Результат прогноза ИИ
                    if value == "Выиграл":
                        cell.fill = win_fill
                    elif value == "Проиграл":
                        cell.fill = loss_fill
                elif col_num == 16:  # Результат Alt ставки
                    if value == "Выиграл":
                        cell.fill = win_fill
                    elif value == "Проиграл":
                        cell.fill = loss_fill
                elif col_num == 18:  # Результат для фаворита
                    if value == "Выиграл":
                        cell.fill = win_fill
                    elif value == "Проиграл":
                        cell.fill = loss_fill
                    elif value == "Ничья":
                        cell.fill = draw_fill
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"[Football] Экспортировано {len(rows)} матчей в Excel")
        return output
        
    except sqlite3.Error as e:
        print(f"[Football ERROR] Ошибка экспорта в Excel (SQLite): {e}")
        return None
    except Exception as e:
        print(f"[Football ERROR] Ошибка экспорта в Excel: {e}")
        import traceback
        print(traceback.format_exc())
        return None
    finally:
        if conn:
            conn.close()

